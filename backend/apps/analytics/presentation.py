"""
MWT.ONE · apps.analytics.presentation — motor de presentación server-side (Ola 3.10 ampliada).

Extiende `chart_svg.py` con las demás categorías del motor de presentación:
  · P2 · Tablas renderizadas (SVG con branding MWT) + versión Markdown.
  · P3 · Reportes (Markdown largo) — el PDF se genera en el endpoint con
         reportlab cuando `formato="pdf"`.
  · P5 · Exportaciones (XLSX con openpyxl / CSV).

Seguridad:
  · Recibe SOLO datos puros (dicts/lists), nunca URLs ni HTML.
  · Todo texto se escapa (XML/Markdown) para evitar inyección.
  · La redacción por rol ya la aplicó el MCP ANTES de llamar aquí.
"""
from __future__ import annotations

import csv as _csv
import html
import io
from typing import Any

# ── Paleta MWT (alineada con --brand-primary del frontend) ──────────────────
_MWT_BRAND = "#013A57"
_MWT_ACCENT = "#0B7285"
_MWT_BORDER = "#DEE2E6"
_MWT_HEAD_BG = "#013A57"
_MWT_ROW_ALT = "#F1F3F5"
_MWT_TEXT = "#212529"
_MWT_MUTED = "#868E96"

_TABLE_MAX_ROWS = 500


def _esc(v: Any) -> str:
    """Escapa un valor para XML/SVG/HTML (previene inyección)."""
    return html.escape(str(v if v is not None else ""), quote=True)


def _fmt(v: Any) -> str:
    """Formatea un valor para display: números con coma de miles, fechas raw."""
    if v is None:
        return ""
    if isinstance(v, float):
        if v == int(v):
            return f"{int(v):,}"
        return f"{v:,.2f}"
    if isinstance(v, int):
        return f"{v:,}"
    return str(v)


# ═══════════════════════════════════════════════════════════════════════ #
# P2 · Tablas renderizadas (SVG) + Markdown
# ═══════════════════════════════════════════════════════════════════════ #
def render_tabla_svg(
    columnas: list,
    filas: list,
    *,
    titulo: str = "",
    resaltar: str | None = None,
    width: int = 900,
) -> str:
    """Genera un SVG de tabla con el branding MWT.

    `columnas`: [{key, label, tipo?}]  (tipo: 'num'|'money'|'date'|None).
    `filas`: list[dict] (las filas YA deben estar redactadas por el MCP).
    `resaltar`: clave a resaltar en negrita (opcional).
    """
    columnas = list(columnas or [])[:20]
    filas = list(filas or [])[: _TABLE_MAX_ROWS]
    n_cols = len(columnas)
    if not columnas or not filas:
        return (
            '<?xml version="1.0" encoding="UTF-8"?>\n'
            f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="90" '
            f'viewBox="0 0 {width} 90" role="img"><rect width="100%" height="100%" '
            f'fill="#ffffff"/><text x="{width/2}" y="46" text-anchor="middle" '
            f'font-family="sans-serif" font-size="14" fill="{_MWT_MUTED}">'
            f'{_esc("Sin datos para la tabla")}</text></svg>'
        )

    col_w = width / n_cols
    row_h = 30
    head_h = 34
    pad = 8
    height = head_h + row_h * len(filas) + 12

    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" '
        f'viewBox="0 0 {width} {height}" role="img">',
        f'<rect width="100%" height="100%" fill="#ffffff"/>',
    ]
    if titulo:
        parts.append(
            f'<text x="{pad}" y="20" font-family="sans-serif" font-size="15" '
            f'font-weight="600" fill="{_MWT_BRAND}">{_esc(titulo)}</text>'
        )
    y0 = 34

    # Cabecera
    parts.append(f'<rect x="0" y="{y0}" width="{width}" height="{head_h}" fill="{_MWT_HEAD_BG}"/>')
    for ci, col in enumerate(columnas):
        x = ci * col_w + pad
        parts.append(
            f'<text x="{x}" y="{y0 + 22}" font-family="sans-serif" font-size="12" '
            f'font-weight="600" fill="#ffffff">{_esc(col.get("label") or col.get("key") or "")}</text>'
        )

    # Filas
    for ri, row in enumerate(filas):
        y = y0 + head_h + ri * row_h
        if ri % 2 == 1:
            parts.append(f'<rect x="0" y="{y}" width="{width}" height="{row_h}" fill="{_MWT_ROW_ALT}"/>')
        parts.append(f'<line x1="0" y1="{y}" x2="{width}" y2="{y}" stroke="{_MWT_BORDER}"/>')
        for ci, col in enumerate(columnas):
            key = col.get("key")
            val = row.get(key) if isinstance(row, dict) else ""
            txt = _fmt(val)
            if key and key == resaltar:
                parts.append(
                    f'<text x="{ci * col_w + pad}" y="{y + 20}" font-family="sans-serif" '
                    f'font-size="12" font-weight="600" fill="{_MWT_TEXT}">{_esc(txt)}</text>'
                )
            else:
                parts.append(
                    f'<text x="{ci * col_w + pad}" y="{y + 20}" font-family="sans-serif" '
                    f'font-size="12" fill="{_MWT_TEXT}">{_esc(txt)}</text>'
                )

    parts.append(f'<line x1="0" y1="{y0 + head_h + row_h * len(filas)}" x2="{width}" '
                 f'y2="{y0 + head_h + row_h * len(filas)}" stroke="{_MWT_BORDER}"/>')
    parts.append("</svg>")
    return "".join(parts)


def render_tabla_markdown(columnas: list, filas: list) -> str:
    """Versión Markdown de la misma tabla (para que la IA la interprete)."""
    columnas = list(columnas or [])[:20]
    filas = list(filas or [])[: _TABLE_MAX_ROWS]
    if not columnas:
        return ""
    headers = [str(c.get("label") or c.get("key") or "") for c in columnas]
    keys = [c.get("key") for c in columnas]
    out = ["| " + " | ".join(headers) + " |", "|" + "---|" * len(columnas)]
    for row in filas:
        cells = [_esc(_fmt(row.get(k))) for k in keys]
        out.append("| " + " | ".join(cells) + " |")
    return "\n".join(out)


# ═══════════════════════════════════════════════════════════════════════ #
# P3 · Reportes (Markdown) — el PDF lo arma el endpoint con reportlab
# ═══════════════════════════════════════════════════════════════════════ #
def render_reporte_markdown(titulo: str, secciones: list) -> str:
    """Arma un reporte Markdown largo a partir de secciones.

    `secciones`: [{titulo, tipo: "texto"|"tabla"|"chart", data?}]
      - tipo=texto: data es str.
      - tipo=tabla: data es {columnas, filas}.
      - tipo=chart: data es el markdown/descripcion (la imagen va aparte).
    """
    out = [f"# {titulo}", ""]
    for sec in secciones or []:
        st = (sec.get("tipo") or "texto").lower()
        out.append(f"## {sec.get('titulo') or ''}")
        out.append("")
        if st == "tabla":
            d = sec.get("data") or {}
            out.append(render_tabla_markdown(d.get("columnas") or [], d.get("filas") or []))
        elif st == "chart":
            out.append(str(sec.get("data") or "(imagen del chart en la respuesta)"))
        else:
            out.append(str(sec.get("data") or ""))
        out.append("")
    return "\n".join(out)


# ═══════════════════════════════════════════════════════════════════════ #
# P5 · Exportaciones (XLSX / CSV)
# ═══════════════════════════════════════════════════════════════════════ #
def export_xlsx_bytes(nombre: str, hojas: list) -> bytes:
    """Genera un workbook XLSX en memoria (openpyxl). Devuelve bytes.

    `hojas`: [{nombre, columnas: [{key, label}], filas: [dict]}]
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    head_fill = PatternFill(start_color="013A57", end_color="013A57", fill_type="solid")
    head_font = Font(color="FFFFFF", bold=True)

    for sheet in hojas or []:
        ws = wb.create_sheet(title=(sheet.get("nombre") or "Hoja1")[:31])
        columnas = sheet.get("columnas") or []
        filas = sheet.get("filas") or []
        keys = [c.get("key") for c in columnas]
        # Header
        for ci, c in enumerate(columnas, start=1):
            cell = ws.cell(row=1, column=ci, value=(c.get("label") or c.get("key") or ""))
            cell.fill = head_fill
            cell.font = head_font
        # Datos
        for ri, row in enumerate(filas, start=2):
            for ci, k in enumerate(keys, start=1):
                ws.cell(row=ri, column=ci, value=row.get(k))
        # Ancho razonable
        for ci in range(1, len(columnas) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_csv_bytes(columnas: list, filas: list) -> bytes:
    """Genera CSV (UTF-8 con BOM para Excel). Devuelve bytes."""
    buf = io.StringIO()
    writer = _csv.writer(buf)
    headers = [str(c.get("label") or c.get("key") or "") for c in (columnas or [])]
    keys = [c.get("key") for c in (columnas or [])]
    writer.writerow(headers)
    for row in filas or []:
        writer.writerow([row.get(k) for k in keys])
    return ("\ufeff" + buf.getvalue()).encode("utf-8")
