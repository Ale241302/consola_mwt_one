"""
=====================================================================
MWT.ONE · apps.expedientes.views_simplified_wizard
Agente responsable: [AG-BACKEND]

Sprint Wizard Simplificado · 2026-04-29.

Endpoints del nuevo wizard de 3 pasos:
  · POST /api/expedientes/parse-template/    → CSV/Excel → JSON validado vs CPA
  · POST /api/catalog/request-assignment/    → email al account manager
                                                con fallback a info@mwt.one

NO toca el endpoint principal POST /api/expedientes/. La relajación de
campos opcionales (marca/mode/currency/freight_mode) ya vive en el
serializer existente con required=False (ver `views_wizard.py` o
`serializers.py` según corresponda).

POL_VISIBILIDAD: ZERO datos financieros en respuestas de estos endpoints.
=====================================================================
"""
from __future__ import annotations

import csv
import io
import logging
import os
import re
import uuid
from datetime import date

from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.db import connection
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

log = logging.getLogger(__name__)

# Configuración por defecto (overrideable vía env)
FALLBACK_TO = os.environ.get("CATALOG_REQUEST_FALLBACK_TO", "info@mwt.one")
DEFAULT_FROM = os.environ.get("DEFAULT_FROM_EMAIL", "info@mwt.one")
ADMIN_BASE_URL = os.environ.get("MWT_ADMIN_BASE_URL", "https://consola.mwt.one")


# =====================================================================
# Helpers
# =====================================================================
def _norm_sku(s):
    return (s or "").strip().upper()


def _norm_size(s):
    return (s or "").strip().upper()


def _safe_int(v, default=0):
    try:
        return int(float(str(v).strip()))
    except (TypeError, ValueError):
        return default


def _read_template_rows(file_bytes: bytes, filename: str) -> list[dict]:
    """Lee CSV o XLSX/XLS y devuelve lista de dicts {sku, talla, cantidad}.

    Espera 3 columnas FIJAS por orden o por nombre (case-insensitive):
       SKU | Talla | Cantidad
    """
    name = (filename or "").lower()
    rows = []

    # ── XLSX / XLS via openpyxl ──
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError("openpyxl no instalado en el backend.")
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        ws = wb.active
        header = None
        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if r_idx == 0:
                header = [(c or "").strip().lower() if isinstance(c, str) else str(c or "").lower()
                          for c in row]
                continue
            d = _row_to_dict(row, header)
            if d:
                rows.append(d)
        return rows

    # ── CSV ──
    text = None
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = file_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise RuntimeError("No pude decodificar el archivo (utf-8 / latin-1).")

    # Saltar directiva Excel `sep=;` si está presente en la primera línea.
    # Microsoft Excel respeta este prefijo para forzar el separador
    # independientemente del locale (MX/CO/PE/ES usan `;`). El parser CSV
    # de Python no la entiende, por eso la quitamos antes de pasarla al
    # Sniffer para que detecte el delimitador real.
    text_for_parsing = text
    first_line = text.split("\n", 1)[0].strip().lower() if text else ""
    if first_line.startswith("sep="):
        text_for_parsing = text.split("\n", 1)[1] if "\n" in text else ""

    # Detectar separador
    sniffer = csv.Sniffer()
    sample = text_for_parsing[:2048]
    try:
        dialect = sniffer.sniff(sample, delimiters=",;|\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text_for_parsing), dialect=dialect)
    header = None
    for r_idx, row in enumerate(reader):
        if not row or all((c or "").strip() == "" for c in row):
            continue
        if header is None:
            header = [(c or "").strip().lower() for c in row]
            continue
        d = _row_to_dict(row, header)
        if d:
            rows.append(d)
    return rows


def _row_to_dict(row, header):
    """Map flexible: nombre conocido O posición.

    Sprint 2026-05-02 (AG-03): la primera columna acepta SKU **o nombre
    del producto** **o ref proveedor**. El resolver `_resolve_input_to_sku`
    en la vista lo mapea a SKU canónico.
    """
    if not row:
        return None
    if header and any(h in ("sku", "talla", "cantidad", "size", "qty", "quantity",
                            "nombre", "name", "producto", "product", "ref")
                     for h in header):
        idx_sku   = next(
            (i for i, h in enumerate(header)
                if h in ("sku", "codigo", "código", "code",
                         "nombre", "name", "producto", "product",
                         "ref", "ref_proveedor", "ref proveedor", "supplier ref")),
            0,
        )
        idx_size  = next((i for i, h in enumerate(header) if h in ("talla", "size", "tamaño")), 1)
        idx_qty   = next((i for i, h in enumerate(header) if h in ("cantidad", "qty", "quantity", "unidades", "units")), 2)
    else:
        idx_sku, idx_size, idx_qty = 0, 1, 2

    try:
        # IMPORTANTE: NO normalizamos a UPPER aquí porque podemos perder
        # información del nombre del producto (ej. "Bota Plena Flor" se
        # vuelve "BOTA PLENA FLOR" lo cual sigue matcheando, pero el resolver
        # más adelante decide cuándo es nombre vs SKU). Solo trim().
        raw_input = (row[idx_sku] or "").strip() if isinstance(row[idx_sku], str) else str(row[idx_sku] or "").strip()
        size = _norm_size(row[idx_size])
        qty  = _safe_int(row[idx_qty], 0)
    except IndexError:
        return None
    if not raw_input or qty <= 0:
        return None
    return {"raw_input": raw_input, "talla": size, "cantidad": qty}


# =====================================================================
# Sprint 2026-05-02 (AG-03) · Resolvers para parse-template
# =====================================================================
# El usuario CEO pidió que la plantilla del expediente acepte:
#   1. Primera columna  → SKU **o** Nombre del producto **o** Ref Proveedor
#   2. Columna talla    → cualquier sistema (BRA, US, UK, EU, CM) con o sin
#                          prefijo explícito. La salida siempre es la talla
#                          base canónica MWT (EU).
#
# Sin estos resolvers, el matchmaker fallaba para PDFs/CSV de clientes que
# nunca usan nuestro SKU interno.
# =====================================================================
def _build_product_index() -> dict:
    """Indexa productos.producto activos por SKU, Nombre y Ref Proveedor.
    Todos en MAYÚSCULAS para lookup case-insensitive."""
    by_sku = {}
    by_nombre = {}
    by_ref = {}
    try:
        with connection.cursor() as c:
            c.execute("""
                SELECT
                    id::text,
                    UPPER(COALESCE(sku, ''))                                AS sku,
                    UPPER(COALESCE(nombre, ''))                             AS nombre,
                    UPPER(COALESCE(especificaciones->>'ref_proveedor','')) AS ref_proveedor
                  FROM productos.producto
                 WHERE COALESCE(is_active, TRUE) = TRUE
            """)
            for pid, sku, nombre, ref in c.fetchall():
                row = {"id": pid, "sku": sku, "nombre": nombre, "ref": ref}
                if sku:    by_sku[sku] = row
                if nombre: by_nombre[nombre] = row
                if ref:    by_ref[ref] = row
    except Exception as e:
        log.warning("[parse_template] product index build failed: %s", e)
    return {"by_sku": by_sku, "by_nombre": by_nombre, "by_ref": by_ref}


def _resolve_input_to_sku(raw_input: str, idx: dict):
    """Resuelve un string libre (puede ser SKU, Nombre o Ref Proveedor) al
    SKU canónico de productos.producto. Devuelve (sku, nombre, strategy)
    o (None, None, 'UNRESOLVED')."""
    s = (raw_input or "").strip().upper()
    if not s:
        return None, None, "EMPTY"
    if s in idx["by_sku"]:
        p = idx["by_sku"][s]
        return p["sku"], p["nombre"], "SKU_EXACT"
    if s in idx["by_ref"]:
        p = idx["by_ref"][s]
        return p["sku"], p["nombre"], "REF_PROVEEDOR_EXACT"
    if s in idx["by_nombre"]:
        p = idx["by_nombre"][s]
        return p["sku"], p["nombre"], "NOMBRE_EXACT"
    # Fuzzy nombre — solo si el input es razonablemente largo (evita falsos
    # positivos sobre tokens cortos como "37")
    if len(s) >= 4:
        for nombre, p in idx["by_nombre"].items():
            if s in nombre or nombre in s:
                return p["sku"], p["nombre"], "NOMBRE_FUZZY"
    return None, None, "UNRESOLVED"


# Sistemas de medida soportados explícitamente en el resolver de tallas.
# Mapea prefijo (escrito por el usuario) → columna en ops.tallas.
_TALLA_SYSTEM_MAP = {
    'BR':    'br',  'BRA':  'br',  'BRASIL': 'br',
    'EU':    'eu',  'EUR':  'eu',  'EUROPA': 'eu',
    'US':    'us_men', 'USA':  'us_men', 'USM':  'us_men', 'USMEN': 'us_men',
    'USW':   'us_women', 'USWOMEN': 'us_women',
    'UK':    'uk_men', 'UKM':  'uk_men', 'UKMEN': 'uk_men',
    'UKW':   'uk_women', 'UKWOMEN': 'uk_women',
    'CM':    'cm',
    'TALLA': None,  # genérico, sin sistema → asume EU
    'SIZE':  None,
    'T':     None,
}
_RE_TALLA_PREFIX = re.compile(
    r'^(BR|BRA|BRASIL|EU|EUR|EUROPA|US|USA|USM|USMEN|USW|USWOMEN|'
    r'UK|UKM|UKMEN|UKW|UKWOMEN|CM|TALLA|SIZE|T)\s*[\.\-:]?\s*(.+)$',
    re.IGNORECASE,
)


def _build_talla_index() -> dict:
    """Catálogo de tallas con dos índices:

      · by_system[col][value] = talla_base
            Para lookup explícito cuando el usuario escribió un prefijo
            (ej. "BRA 37" → buscar en by_system['br']['37']).

      · by_value[value] = talla_base
            Lookup permisivo SIN prefijo. EU tiene prioridad — si "35"
            existe como EU y como BR (de talla_base 37), gana el EU.
    """
    by_value = {}
    by_system = {
        'eu': {}, 'us_men': {}, 'us_women': {},
        'uk_men': {}, 'uk_women': {}, 'br': {}, 'cm': {},
    }
    canonical_set = set()
    rows = []
    try:
        with connection.cursor() as c:
            c.execute("""
                SELECT talla_base, eu, us_men, us_women, uk_men, uk_women, br, cm
                  FROM ops.tallas
                 WHERE COALESCE(is_active, TRUE) = TRUE
                   AND tipo_producto = 'calzado'
            """)
            rows = c.fetchall()
    except Exception as e:
        log.warning("[parse_template] talla index build failed: %s", e)
        return {"by_value": by_value, "by_system": by_system, "canonical": canonical_set}

    # Pass 1 — canonical (talla_base + EU). EU = talla_base por convención
    # del seed B2_seed_tallas_calzado, así que estos dos valores ganan
    # cualquier colisión futura con BR/US/UK.
    for talla_base, eu, us_men, us_women, uk_men, uk_women, br, cm in rows:
        base = (talla_base or "").upper().strip()
        if not base:
            continue
        canonical_set.add(base)
        by_value[base] = base
        if eu:       by_system['eu'][eu.upper().strip()] = base
        if us_men:   by_system['us_men'][us_men.upper().strip()] = base
        if us_women: by_system['us_women'][us_women.upper().strip()] = base
        if uk_men:   by_system['uk_men'][uk_men.upper().strip()] = base
        if uk_women: by_system['uk_women'][uk_women.upper().strip()] = base
        if br:       by_system['br'][br.upper().strip()] = base
        if cm:       by_system['cm'][cm.upper().strip()] = base

    # Pass 2 — by_value para sistemas no-EU (sólo si la clave aún está libre)
    for talla_base, eu, us_men, us_women, uk_men, uk_women, br, cm in rows:
        base = (talla_base or "").upper().strip()
        if not base:
            continue
        for v in (us_men, us_women, uk_men, uk_women, br, cm):
            if v:
                k = v.upper().strip()
                if k not in by_value:
                    by_value[k] = base

    return {
        "by_value":  by_value,
        "by_system": by_system,
        "canonical": canonical_set,
    }


def _resolve_talla(raw: str, idx: dict):
    """Mapea una talla libre al canonical EU. Devuelve (talla_base, strategy).

    Casos:
      · "37"        → ("37", "DIRECT_EU")        — sin prefijo, asume EU
      · "BRA 37"    → ("39", "BR")               — BR 37 = EU 39 (vía matriz)
      · "US 9.5"    → ("43", "US_MEN")           — US 9.5 = EU 43
      · "UK 9"      → ("43", "UK_MEN")           — UK 9 = EU 43
      · "EU 43"     → ("43", "EU")               — EU explícito
      · "TALLA 42"  → ("42", "DIRECT_EU")        — prefijo genérico, asume EU
      · "M"         → ("M",  "PASSTHROUGH")      — talla alfa, no-numeric
      · "UNICA"     → ("UNICA", "PASSTHROUGH")   — caso especial
    """
    if not raw:
        return None, "EMPTY"
    s = raw.strip().upper()

    # Caso 1: prefijo de sistema explícito
    m = _RE_TALLA_PREFIX.match(s)
    if m:
        sys_raw = m.group(1).upper()
        value = m.group(2).strip()
        col = _TALLA_SYSTEM_MAP.get(sys_raw)
        if col is None:
            # Prefijo genérico tipo "TALLA 42" → asume canónica
            resolved = idx["by_value"].get(value, value)
            return resolved, "DIRECT_EU"
        # Lookup en el sistema específico
        target = idx["by_system"].get(col, {}).get(value)
        if target:
            return target, col.upper()
        # Sistema reconocido pero valor no en matriz → passthrough con valor crudo
        return value, f"{col.upper()}_PASSTHROUGH"

    # Caso 2: sin prefijo → lookup permisivo (EU prioritario)
    if s in idx["by_value"]:
        return idx["by_value"][s], "DIRECT_EU"

    # Caso 3: passthrough (M, L, UNICA, etc.)
    return s, "PASSTHROUGH"


def _resolve_account_manager_email(client_id):
    """Devuelve email del Account Manager del cliente, o (None) si no se puede.

    Lógica:
      1. cliente.responsable_id → core.users.email (si la tabla existe)
      2. fallback NULL → caller usa FALLBACK_TO.
    """
    if not client_id:
        return None, None
    try:
        with connection.cursor() as c:
            c.execute("""
                SELECT cl.responsable_id, cl.razon_social, cl.contacto_nombre,
                       cl.contacto_email
                  FROM clientes.cliente cl
                 WHERE cl.id = %s
            """, [str(client_id)])
            row = c.fetchone()
            if not row:
                return None, None
            resp_id, razon_social, contacto_nombre, contacto_email = row

            am_email = None
            am_name = None
            if resp_id:
                # core.users (Sprint M3)
                try:
                    c.execute("""
                        SELECT email, full_name FROM core.users
                         WHERE id = %s AND is_active = TRUE
                    """, [str(resp_id)])
                    user_row = c.fetchone()
                    if user_row:
                        am_email, am_name = user_row[0], user_row[1]
                except Exception:
                    # core.users puede no existir en builds viejos
                    pass

            return (am_email or contacto_email or None,
                    am_name or contacto_nombre or razon_social or "")
    except Exception:
        log.exception("[catalog/request_assignment] resolve AM falló")
        return None, None


# =====================================================================
# /api/expedientes/parse-template/
# =====================================================================
class ParseTemplateView(APIView):
    """POST multipart con `file` + form `client_id`.

    Devuelve:
      {
        "lines": [
          {"sku":"...", "talla":"...", "cantidad":12,
           "is_assigned": true|false,
           "product_label":"...",
           "producto_id":"...",
           "row":  3 },
          ...
        ],
        "summary": {
          "total_rows":         42,
          "valid_rows":         40,
          "assigned_rows":      37,
          "unassigned_rows":     3,
          "client_id":         "...",
          "client_label":      "..."
        },
        "unassigned_skus": ["NIK-AIR-001", "ADI-PRO-002"],
        "errors": []
      }
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        f = request.FILES.get("file") or request.FILES.get("upload")
        if not f:
            return Response({"detail": "Falta el archivo (`file`)."}, status=400)
        if f.size > 10 * 1024 * 1024:
            return Response({"detail": "Archivo > 10MB."}, status=413)

        client_id = (request.data.get("client_id") or "").strip()
        if not client_id:
            return Response({"detail": "client_id requerido para validar CPA."}, status=400)

        # Parse
        try:
            rows = _read_template_rows(f.read(), f.name)
        except Exception as e:
            return Response({"detail": f"No pude parsear el archivo: {e}"}, status=400)

        if not rows:
            return Response({
                "lines":            [],
                "summary":          {"total_rows": 0, "valid_rows": 0,
                                     "assigned_rows": 0, "unassigned_rows": 0,
                                     "client_id": client_id, "client_label": ""},
                "unassigned_skus":  [],
                "errors":           ["Archivo vacío o sin líneas válidas."],
            }, status=200)

        # Snapshot del cliente
        client_label = ""
        try:
            with connection.cursor() as c:
                c.execute(
                    "SELECT razon_social FROM clientes.cliente WHERE id = %s",
                    [client_id],
                )
                r = c.fetchone()
                if r:
                    client_label = r[0] or ""
        except Exception:
            pass

        # ── Sprint 2026-05-02: resolución IA-light ──
        # 1. Catálogo de productos por SKU/Nombre/Ref Proveedor
        # 2. Catálogo de tallas con sistemas explícitos (BR/US/UK/EU/CM)
        product_idx = _build_product_index()
        talla_idx   = _build_talla_index()

        # Resolvemos cada fila: input crudo → SKU canónico, talla cruda → talla base
        for r in rows:
            sku, label, sku_strategy = _resolve_input_to_sku(r["raw_input"], product_idx)
            r["original_input"]   = r["raw_input"]
            r["sku"]              = sku or r["raw_input"].upper()  # fallback al input crudo si no resuelve
            r["product_label"]    = label
            r["sku_strategy"]     = sku_strategy
            r["sku_resolved"]     = bool(sku)

            base, talla_strategy  = _resolve_talla(r["talla"], talla_idx)
            r["original_talla"]   = r["talla"]
            r["talla"]            = base or r["talla"]
            r["talla_strategy"]   = talla_strategy
            r["talla_resolved"]   = (talla_strategy not in ("PASSTHROUGH", "EMPTY"))

        # SKUs únicos (ya canónicos) para el lookup CPA
        unique_skus = list({r["sku"] for r in rows if r.get("sku")})
        assignment_skus = set()
        product_meta = {}
        if unique_skus:
            try:
                with connection.cursor() as c:
                    # CPA → solo SKUs activos asignados al cliente
                    c.execute("""
                        SELECT brand_sku
                          FROM pricing.client_assignment
                         WHERE client_id = %s
                           AND is_active = TRUE
                           AND (valid_to IS NULL OR valid_to >= CURRENT_DATE)
                           AND brand_sku = ANY(%s)
                    """, [client_id, unique_skus])
                    assignment_skus = {row[0] for row in c.fetchall()}
            except Exception:
                log.exception("[parse_template] CPA lookup falló")

            try:
                with connection.cursor() as c:
                    c.execute("""
                        SELECT id, sku, nombre
                          FROM productos.producto
                         WHERE sku = ANY(%s)
                    """, [unique_skus])
                    for row in c.fetchall():
                        product_meta[row[1]] = {
                            "producto_id":   str(row[0]),
                            "product_label": row[2] or "",
                        }
            except Exception:
                pass

        out_lines = []
        unassigned = []
        for i, r in enumerate(rows):
            is_assigned = r["sku"] in assignment_skus
            if not is_assigned and r["sku"] not in unassigned:
                unassigned.append(r["sku"])
            meta = product_meta.get(r["sku"]) or {}
            out_lines.append({
                "row":             i + 2,           # +2 = header + 1-based
                "sku":             r["sku"],
                "talla":           r["talla"],
                "cantidad":        r["cantidad"],
                "is_assigned":     is_assigned,
                "producto_id":     meta.get("producto_id"),
                "product_label":   meta.get("product_label") or r.get("product_label"),
                # Trazabilidad de la resolución (para que el frontend pueda
                # mostrar badges / warnings sobre qué se resolvió y cómo).
                "original_input":  r["original_input"],
                "original_talla":  r["original_talla"],
                "sku_resolved":    r["sku_resolved"],
                "sku_strategy":    r["sku_strategy"],
                "talla_resolved":  r["talla_resolved"],
                "talla_strategy":  r["talla_strategy"],
            })

        # Resumen de resolución (extra de las métricas CPA)
        resolution_summary = {
            "sku_resolved":   len([l for l in out_lines if l["sku_resolved"]]),
            "sku_unresolved": len([l for l in out_lines if not l["sku_resolved"]]),
            "talla_resolved": len([l for l in out_lines if l["talla_resolved"]]),
        }

        return Response({
            "lines":           out_lines,
            "summary": {
                "total_rows":      len(rows),
                "valid_rows":      len(out_lines),
                "assigned_rows":   len([l for l in out_lines if l["is_assigned"]]),
                "unassigned_rows": len([l for l in out_lines if not l["is_assigned"]]),
                "client_id":       client_id,
                "client_label":    client_label,
                "resolution":      resolution_summary,
            },
            "unassigned_skus": unassigned,
            "errors":          [],
        })


# =====================================================================
# /api/catalog/request-assignment/
# =====================================================================
class CatalogRequestAssignmentView(APIView):
    """POST {client_id, sku, talla?, cantidad?, client_email?}

    Resuelve el destinatario:
      1. responsable_id del cliente → core.users.email
      2. fallback: clientes.cliente.contacto_email
      3. fallback final: info@mwt.one (env CATALOG_REQUEST_FALLBACK_TO)

    Devuelve:
      {ok: true, sent_to: "...", subject: "...", fallback_used: bool}
    """
    permission_classes = [IsAuthenticated]

    EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

    def post(self, request):
        d = request.data or {}
        client_id = (d.get("client_id") or "").strip()
        sku       = _norm_sku(d.get("sku"))
        talla     = _norm_size(d.get("talla"))
        cantidad  = _safe_int(d.get("cantidad"), 0)
        client_email = (d.get("client_email") or "").strip()

        if not client_id or not sku:
            return Response({"detail": "client_id y sku requeridos."}, status=400)

        # Resolver destinatario
        am_email, am_name = _resolve_account_manager_email(client_id)
        fallback_used = False
        if not am_email or not self.EMAIL_RE.match(am_email):
            am_email = FALLBACK_TO
            am_name  = "Equipo MWT"
            fallback_used = True

        # Cliente metadata
        client_label = ""
        client_contact_email = ""
        try:
            with connection.cursor() as c:
                c.execute("""
                    SELECT razon_social, contacto_email FROM clientes.cliente WHERE id = %s
                """, [client_id])
                row = c.fetchone()
                if row:
                    client_label = row[0] or ""
                    client_contact_email = row[1] or ""
        except Exception:
            pass

        if not client_email:
            client_email = client_contact_email

        # Buscar producto / brand para el link
        product_label = ""
        producto_id = None
        brand_id = None
        try:
            with connection.cursor() as c:
                c.execute("""
                    SELECT id, nombre, marca_id FROM productos.producto WHERE sku = %s LIMIT 1
                """, [sku])
                pr = c.fetchone()
                if pr:
                    producto_id, product_label, brand_id = str(pr[0]), pr[1] or "", pr[2]
        except Exception:
            pass

        link_admin = (
            f"{ADMIN_BASE_URL}/marcas/{brand_id}/clientes/{client_id}/precios"
            if brand_id else
            f"{ADMIN_BASE_URL}/clientes/{client_id}"
        )

        # Componer email
        subject = f"[MWT.ONE] Solicitud de asignación de SKU {sku} — {client_label}"
        text_body = (
            f"Hola {am_name or ''},\n\n"
            f"El cliente {client_label}"
            + (f" ({client_email})" if client_email else "")
            + " ha solicitado acceso al siguiente producto:\n\n"
            f"  · SKU:      {sku}\n"
            f"  · Producto: {product_label or '—'}\n"
            f"  · Talla:    {talla or '—'}\n"
            f"  · Cantidad: {cantidad or '—'}\n\n"
            f"Para autorizar y asignar el SKU al cliente, abrí la siguiente vista:\n"
            f"  {link_admin}\n\n"
            f"— MWT.ONE · Catálogo Comercial"
        )
        html_body = f"""
        <div style="font-family:Inter,Arial,sans-serif;font-size:14px;color:#0B1E3A;
                    max-width:600px;margin:0 auto;padding:24px;
                    background:#F8FAFC;border-radius:12px">
          <div style="background:#0B1E3A;color:#fff;padding:16px 20px;border-radius:10px;
                      margin-bottom:18px">
            <div style="font-size:11px;color:#1DE394;letter-spacing:1.5px;font-weight:700">
              SOLICITUD DE ASIGNACIÓN DE SKU
            </div>
            <div style="font-size:18px;font-weight:700;margin-top:4px">{client_label}</div>
          </div>

          <p>Hola <strong>{am_name or ''}</strong>,</p>
          <p>Tu cliente <strong>{client_label}</strong>
             {f"(<a href='mailto:{client_email}'>{client_email}</a>)" if client_email else ""}
             solicita acceso al siguiente producto:</p>

          <table style="width:100%;border-collapse:collapse;margin:16px 0">
            <tr><td style="padding:8px;color:#64748B;font-size:11px;
                           text-transform:uppercase;letter-spacing:0.5px">SKU</td>
                <td style="padding:8px;font-weight:700;font-family:monospace">{sku}</td></tr>
            <tr><td style="padding:8px;color:#64748B;font-size:11px;
                           text-transform:uppercase;letter-spacing:0.5px">Producto</td>
                <td style="padding:8px">{product_label or '—'}</td></tr>
            <tr><td style="padding:8px;color:#64748B;font-size:11px;
                           text-transform:uppercase;letter-spacing:0.5px">Talla</td>
                <td style="padding:8px">{talla or '—'}</td></tr>
            <tr><td style="padding:8px;color:#64748B;font-size:11px;
                           text-transform:uppercase;letter-spacing:0.5px">Cantidad</td>
                <td style="padding:8px;font-weight:600">{cantidad or '—'}</td></tr>
          </table>

          <a href="{link_admin}"
             style="display:inline-block;background:#00B286;color:#fff;text-decoration:none;
                    padding:12px 22px;border-radius:8px;font-weight:700;letter-spacing:0.3px">
            Autorizar / Asignar SKU →
          </a>

          <p style="margin-top:24px;color:#64748B;font-size:12px">
            Este mensaje fue generado automáticamente por MWT.ONE. No respondas a este correo.
          </p>
        </div>
        """

        try:
            msg = EmailMultiAlternatives(
                subject = subject,
                body    = text_body,
                from_email = DEFAULT_FROM,
                to      = [am_email],
                reply_to = [client_email] if client_email else None,
            )
            msg.attach_alternative(html_body, "text/html")
            msg.send(fail_silently=False)
        except Exception as e:
            log.exception("[catalog/request_assignment] envío SMTP falló")
            return Response({
                "detail": f"No pude enviar el email: {type(e).__name__}: {e}",
                "sent_to": am_email,
                "fallback_used": fallback_used,
            }, status=500)

        return Response({
            "ok":            True,
            "sent_to":       am_email,
            "sent_to_name":  am_name,
            "subject":       subject,
            "fallback_used": fallback_used,
            "link_admin":    link_admin,
        })
