# backend/apps/expedientes/document_matchmaker.py
"""
=====================================================================
MWT.ONE · apps.expedientes.document_matchmaker
Agente responsable: [AG-BACKEND]

Sprint Document Matchmaker · 2026-04-29.
Update 2026-05-02 (AG-03): el extractor OC ahora captura múltiples
señales de identidad del producto (client_part_number, supplier_ref,
product_label) y cae a un resolver que mapea contra productos.producto
por SKU / Nombre / Ref Proveedor — porque los clientes NUNCA codifican
con nuestro SKU interno.

Patrón A (SonDel)   → sólo "Part Nº" tipo 75BPR29-CLIMM-CPAP-37
                      → base 75BPR29-CLIMM-CPAP matchea productos.nombre.
Patrón B (Sonepar)  → "ARTICULO" + "REF. PROVEEDOR" tipo 60B19M-CPAP-MIN-CP-38
                      → base 60B19M-CPAP-MIN-CP matchea
                        productos.especificaciones->>'ref_proveedor'.

Salida estructurada GARANTIZADA: si la IA falla o devuelve algo raro,
caemos a un payload determinístico vacío con `error` poblado.
=====================================================================
"""
from __future__ import annotations

import base64
import json
import logging
import os
import re
from typing import Optional

from django.db import connection

log = logging.getLogger(__name__)

OCR_MODEL = os.environ.get("OPENAI_OCR_MODEL", "gpt-5-nano")

# Sufijo de talla típico: -37, -38, -XL, _M, etc. al final del código.
_RE_TALLA_SUFFIX = re.compile(r"[-_]([0-9]{2,3}|[A-Z]{1,3})$", re.IGNORECASE)


# ─────────────────────────────────────────────────────────────────────
# System prompts
# ─────────────────────────────────────────────────────────────────────
SYSTEM_PROMPT_OC = """Eres un extractor de Órdenes de Compra (OC) del cliente.
Lees el documento y devuelves un JSON ESTRICTO con la lista de líneas.

CONTEXTO CRÍTICO: el cliente NUNCA codifica con el SKU interno del proveedor.
Cada cliente usa su propia convención. Tu trabajo es extraer TODAS las
señales de identidad presentes para que el resolver del backend pueda
mapearlas contra el catálogo.

DETECCIÓN DEL SHAPE DE COLUMNAS (PRIORIDAD ALTA — leer con atención):

  · CASO A — UNA columna de código (típico SonDel, SAP Business One):
    Headers tipo "Part Nº" o "Articulo" o "Code".
       client_part_number = ese código
       supplier_ref       = null

  · CASO B — DOS columnas de código consecutivas (típico Sonepar, sistemas
    con cross-referencing). Headers tipo:

        ARTICULO              REF. PROVEEDOR       DESCRIPCION
        27BM60B19M-CPAP-...   60B19M-CPAP-MIN-CP   BOTA META NE...

    EL HEADER PUEDE VENIR CORTADO POR LINE-WRAP, ej:
        "POS. CANTIDAD UND. ARTICULO REF. DESCRIPCION PRECIO IMPORTE"
        "                                PROVEEDOR  UNITARIO"
    Ahí "REF." (línea 1) + "PROVEEDOR" (línea 2) forman UN solo header
    "REF. PROVEEDOR" que indica COLUMNA SEPARADA del proveedor.

    HEURÍSTICA DEFINITIVA: si en cada fila ves DOS códigos alfanuméricos
    similares lado a lado antes de la descripción, eso es CASO B.
    Casi siempre el segundo está CONTENIDO dentro del primero como
    sufijo (ej. "27BM60B19M-CPAP-MIN-CP-38" CONTIENE
    "60B19M-CPAP-MIN-CP-38" — el cliente prefijó "27BM").

    En CASO B DEBES extraer:
       client_part_number = PRIMER código de la fila
       supplier_ref       = SEGUNDO código de la fila

Esquema obligatorio:
{
  "document_kind": "OC",
  "client_po_number": "<si aparece>",
  "client_name":      "<si aparece>",
  "issued_date":      "YYYY-MM-DD",
  "currency":         "USD",
  "lines": [
    {
      "client_part_number": "<código que aparece en la columna 'Part Nº' / 'Articulo' / 'Code' del cliente — string completo tal cual aparece>",
      "supplier_ref":       "<si existe una columna 'REF Proveedor' / 'Supplier Ref' / 'REF.' aparte, ponlo aquí; si no, null>",
      "base_code":          "<el code SIN el sufijo de talla — ej. de '75BPR29-CLIMM-CPAP-37' devolver '75BPR29-CLIMM-CPAP'>",
      "talla":              "<talla numérica o letra extraída del sufijo o de columna 'Size' — ej. '37','38','XL'>",
      "product_label":      "<descripción si aparece>",
      "qty":                <entero>,
      "unit_price":         <decimal o null>,
      "confidence":         0..100
    }, ...
  ],
  "raw_text": "<transcripción literal, max 2000 chars>"
}

Reglas duras:
  1. CERO INVENTOS. Si un campo no aparece, omitirlo (NO lo pongas en null si no aparece — omítelo).
  2. Si NO hay tabla de productos, devolver lines=[].
  3. client_part_number / supplier_ref / base_code / talla en MAYÚSCULAS sin espacios.
  4. base_code: si no puedes inferirlo con seguridad (>90%), devuelve null.
  5. talla: si no puedes inferirla, devuelve null. Convierte "T-38" / "Talla 38" a "38".
  6. Devolver SOLO el JSON, sin texto adicional, sin markdown.
"""


SYSTEM_PROMPT_PROFORMA = """Eres un extractor de Proformas MWT (documento comercial interno).

ESTRUCTURA TÍPICA — cada producto en la proforma se rinde como un slot
con esta forma:

  ┌──────────────────────────────────────────────────────────────────┐
  │ Código:       701935 (SKU MWT canónico)                            │
  │ Referencia:   60B19M-CPAP-MIN-CP (REF del proveedor)               │
  │ Descripción:  "Bota meta NE p-comp..."                              │
  │ Color:        NEGRO                                                 │
  │ Precio $:     19.35    Cantidad total: 110    Total línea: $2,128  │
  │                                                                     │
  │ MATRIZ DE TALLAS (las 3 filas se repiten contiguas):                │
  │    Referencia BRA: 33  34  35  36  37  38  39  40  41  42  43 ...    │
  │    Referencia EU:  35  36  37  38  39  40  41  42  43  44  45 ...    │
  │    Referencia USA: 4.5 5.5 6.5  7   8  8.5 9.5 10 11 12 13 ...        │
  │    Qty:            0   0   0   0   10  10  10  10  30  30  10 ...     │
  └──────────────────────────────────────────────────────────────────┘

  La proforma puede tener varios slots (1, 2, 3, ... 11). Slots vacíos
  (sin código y todas las qty=0) deben ser IGNORADOS — no devolverlos.

TU TRABAJO: para cada slot con datos, EXPANDIR la matriz a una línea
por talla con qty>0. Si una talla tiene qty=0, no la devuelvas.

Esquema obligatorio:
{
  "document_kind":     "PROFORMA",
  "proforma_number":   "<código MWT, ej. 2414-2026>",
  "client_po_number":  "<si aparece, ej. 110022220>",
  "client_name":       "<si aparece, ej. Sonepar Colombia>",
  "issued_date":       "YYYY-MM-DD",
  "currency":          "USD",
  "groups": [
    {
      "sap_number":  null,
      "lines": [
        {
          "sku":           "<código MWT canónico, ej. 701935>",
          "supplier_ref":  "<referencia proveedor sin talla, ej. 60B19M-CPAP-MIN-CP>",
          "product_label": "<descripción>",
          "talla":         "<EU canónica, ej. 39, 40, 41>",
          "qty":           <entero>,
          "unit_price":    <decimal o null>,
          "confidence":    0..100
        }, ...
      ]
    }, ...
  ],
  "raw_text": "<transcripción literal, max 2000 chars>"
}

Reglas duras:
  1. CERO INVENTOS. Si un campo no aparece, omitirlo.
  2. EXPANDIR la matriz: cada talla con qty>0 es una línea separada.
  3. Talla canónica = EU. Si la cantidad está en la columna BR 37,
     la talla EU correspondiente es 39 (la fila contigua de la matriz).
  4. SKU / supplier_ref / talla en MAYÚSCULAS sin espacios.
  5. Slots vacíos (sin Código y qty totalmente en 0) → ignorar.
  6. Si la proforma agrupa varias OCs/SAPs, usar groups múltiples.
     Si solo es una orden, un único group con sap_number=null.
  7. Devolver SOLO el JSON, sin texto adicional, sin markdown.
"""


SYSTEM_PROMPT_SAP = """Eres un extractor de Confirmaciones SAP del proveedor.
Lees el documento y devuelves un JSON ESTRICTO.

Una confirmación SAP puede traer una o varias órdenes. Cada orden tiene
un sap_number (también llamado "Sales Order" o "Auftrag") y una lista
de líneas con material, talla y cantidad confirmada.

Esquema obligatorio:
{
  "document_kind": "SAP",
  "supplier_name": "<si aparece>",
  "issued_date":   "YYYY-MM-DD",
  "groups": [
    {
      "sap_number":   "<código SAP / SO #>",
      "po_reference": "<MWT PO si aparece>",
      "delivery_date":"YYYY-MM-DD",
      "lines": [
        {"sku":"...", "product_label":"...", "talla":"...",
         "qty_confirmed":N, "qty_open":N, "confidence":0..100}, ...
      ]
    }, ...
  ],
  "raw_text": "<transcripción literal, max 2000 chars>"
}

Reglas:
  1. CERO INVENTOS.
  2. SKU / talla en MAYÚSCULAS sin espacios.
  3. Si no hay sap_number, usar null y agrupar todas las líneas ahí.
  4. Devolver SOLO el JSON.
"""


PROMPT_BY_TYPE = {
    "ART-01_OC":       SYSTEM_PROMPT_OC,
    "ART-02_PROFORMA": SYSTEM_PROMPT_PROFORMA,
    "ART-04_SAP":      SYSTEM_PROMPT_SAP,
}


# ─────────────────────────────────────────────────────────────────────
# Llamada IA
# ─────────────────────────────────────────────────────────────────────
def _empty_result(document_kind, error=None):
    return {
        "document_kind": document_kind,
        "lines":         [],
        "groups":        [],
        "raw_text":      "",
        "error":         error,
        "model":         OCR_MODEL,
    }


def extract_document(file_bytes: bytes, filename: str, content_type: str,
                     document_type: str, expediente_id=None) -> dict:
    """Punto de entrada del extractor.

    Sprint 2026-05-02 (AG-03): para PROFORMA delegamos a un MÓDULO
    SEPARADO (proforma_extractor.py) que usa vision API. Esto aísla
    completamente el path de proforma del path de OC/Factura/Otros —
    cualquier cambio en el extractor de proforma NO PUEDE afectar la OC.

    `expediente_id` se usa SOLO para PROFORMA: el extractor lo usa para
    cargar las líneas BD del expediente y pasárselas al AI como contexto
    de anclaje. La OC y SAP NO consumen este parámetro — el path queda
    intacto y no se ve afectado por nada que pase con la proforma.
    """
    # ── PROFORMA → módulo dedicado, aislado ────────────────────────
    if document_type == "ART-02_PROFORMA":
        from .proforma_extractor import extract_proforma
        return extract_proforma(file_bytes, filename, content_type,
                                expediente_id=expediente_id)

    # ── Resto (OC, SAP, Otros) → path original ─────────────────────
    api_key = os.environ.get("OPENAI_API_KEY")
    kind_map = {
        "ART-01_OC": "OC", "ART-02_PROFORMA": "PROFORMA", "ART-04_SAP": "SAP",
    }
    document_kind = kind_map.get(document_type, "OTHER")

    if not api_key:
        return _empty_result(document_kind, "OPENAI_API_KEY no configurada en el servidor.")

    is_pdf   = content_type == "application/pdf" or filename.lower().endswith(".pdf")
    is_image = content_type.startswith("image/") or any(
        filename.lower().endswith(ext) for ext in (".jpg", ".jpeg", ".png", ".webp")
    )
    is_excel = filename.lower().endswith((".xlsx", ".xlsm", ".xls"))
    is_csv   = filename.lower().endswith(".csv")

    if not (is_pdf or is_image or is_excel or is_csv):
        return _empty_result(document_kind, f"Tipo de archivo no soportado: {content_type}")

    text_payload = None
    if is_excel:
        try:
            from openpyxl import load_workbook
            import io as _io
            wb = load_workbook(_io.BytesIO(file_bytes), data_only=True, read_only=True)
            ws = wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append("\t".join(str(c) if c is not None else "" for c in row))
            text_payload = "\n".join(rows)[:18000]
        except Exception as e:
            return _empty_result(document_kind, f"No pude leer el Excel: {e}")
    elif is_csv:
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                text_payload = file_bytes.decode(enc)[:18000]
                break
            except UnicodeDecodeError:
                continue
        if text_payload is None:
            return _empty_result(document_kind, "No pude decodificar el CSV.")
    elif is_pdf:
        # Sprint 2026-05-02 (AG-03): los PDFs de clientes (SonDel, Sonepar,
        # MARLUVAS) son text-native, no escaneos. Extraemos texto con pypdf
        # y lo mandamos como prompt de chat.completions — mucho más rápido,
        # barato, y compatible con cualquier modelo de chat. Antes mandábamos
        # el binario a `responses.create()` con `input_file`, que rechaza
        # PDFs con `Invalid MIME type. Only image types are supported`.
        # Si pypdf no logra extraer texto (PDF escaneado, OCR-only), caemos
        # al path de vision como fallback.
        try:
            from pypdf import PdfReader
            import io as _io
            reader = PdfReader(_io.BytesIO(file_bytes))
            pages = []
            for page in reader.pages:
                t = (page.extract_text() or "").strip()
                if t:
                    pages.append(t)
            if pages:
                text_payload = "\n\n--- PAGE BREAK ---\n\n".join(pages)[:18000]
                log.info("[matchmaker] pypdf extrajo %d páginas (%d chars) de %s",
                         len(pages), len(text_payload), filename)
        except Exception as e:
            log.warning("[matchmaker] pypdf extracción falló (%s); fallback a vision", e)
            # text_payload queda None → cae al path PDF/Imagen (vision).

    try:
        from openai import OpenAI
    except ImportError:
        return _empty_result(document_kind, "Paquete `openai` no instalado en el backend.")

    # Sprint 2026-05-02 (AG-03): timeout=45s + max_retries=1 para que la
    # llamada falle limpio antes de los 120s del worker de gunicorn. Con
    # los defaults (timeout=60, max_retries=2), una API lenta podía estirar
    # la request a 60×3=180s y matar el worker con SIGKILL → 500 mudo.
    client = OpenAI(api_key=api_key, timeout=45.0, max_retries=1)
    system_prompt = PROMPT_BY_TYPE.get(document_type, SYSTEM_PROMPT_OC)

    raw_text = None
    try:
        if text_payload is not None:
            chat = client.chat.completions.create(
                model = OCR_MODEL,
                messages = [
                    {"role": "system", "content": system_prompt},
                    {"role": "user",   "content":
                        "Analiza el siguiente documento (Excel/CSV) y devuelve solo JSON:\n\n"
                        + text_payload},
                ],
                response_format = {"type": "json_object"},
            )
            raw_text = chat.choices[0].message.content
        else:
            b64 = base64.b64encode(file_bytes).decode("ascii")
            data_url = f"data:{content_type};base64,{b64}"
            user_prompt = "Analiza el siguiente documento y devuelve solo JSON."
            try:
                resp = client.responses.create(
                    model        = OCR_MODEL,
                    instructions = system_prompt,
                    input=[{
                        "role": "user",
                        "content": [
                            {"type": "input_text",  "text": user_prompt},
                            {"type": "input_image", "image_url": data_url}
                                if is_image else
                            {"type": "input_file",  "filename": filename, "file_data": data_url},
                        ],
                    }],
                    response_format={"type": "json_object"},
                )
                raw_text = resp.output_text
            except Exception:
                content_parts = [{"type": "text", "text": user_prompt},
                                 {"type": "image_url", "image_url": {"url": data_url}}]
                chat = client.chat.completions.create(
                    model    = OCR_MODEL,
                    messages = [{"role": "system", "content": system_prompt},
                                {"role": "user",   "content": content_parts}],
                    response_format = {"type": "json_object"},
                )
                raw_text = chat.choices[0].message.content
    except Exception as e:
        log.exception("[matchmaker] OpenAI call failed")
        return _empty_result(document_kind, f"OpenAI API error: {type(e).__name__}: {e}")

    try:
        data = json.loads(raw_text)
    except json.JSONDecodeError as e:
        return _empty_result(document_kind, f"JSON inválido del modelo: {e}")

    out = {
        "document_kind": data.get("document_kind") or document_kind,
        "raw_text":      (data.get("raw_text") or "")[:2000] if isinstance(data.get("raw_text"), str) else "",
        "model":         OCR_MODEL,
        "error":         None,
    }
    for k in ("client_po_number", "client_name", "issued_date", "currency",
              "proforma_number", "supplier_name"):
        if data.get(k) is not None:
            out[k] = data.get(k)

    if document_kind == "OC":
        out["lines"]  = _normalize_oc_lines(data.get("lines"))
        out["groups"] = []
    else:
        groups = []
        for g in (data.get("groups") or []):
            sap = g.get("sap_number") or None
            lines = _normalize_lines(g.get("lines"), is_sap=(document_kind == "SAP"))
            groups.append({
                "sap_number":    sap,
                "po_reference":  g.get("po_reference") or None,
                "delivery_date": g.get("delivery_date") or None,
                "lines":         lines,
            })
        if not groups and data.get("lines"):
            groups = [{
                "sap_number":    None,
                "po_reference":  None,
                "delivery_date": None,
                "lines":         _normalize_lines(data["lines"], is_sap=(document_kind == "SAP")),
            }]
        out["groups"] = groups
        out["lines"]  = []
    return out


# ─────────────────────────────────────────────────────────────────────
# Normalización
# ─────────────────────────────────────────────────────────────────────
def _split_base_and_talla(code: str) -> tuple[Optional[str], Optional[str]]:
    """Parte 'AAA-BBB-37' en ('AAA-BBB','37'). Defensivo."""
    if not code:
        return (None, None)
    code = code.strip().upper().replace(" ", "")
    m = _RE_TALLA_SUFFIX.search(code)
    if not m:
        return (code, None)
    base = code[:m.start()]
    talla = m.group(1)
    return (base or None, talla or None)


def _normalize_oc_lines(rows):
    """Normalización específica del OC: capturamos múltiples señales de identidad
    y derivamos base_code / talla cuando la IA no lo hizo."""
    out = []
    for ln in (rows or []):
        if not isinstance(ln, dict):
            continue

        client_part = str(ln.get("client_part_number") or ln.get("sku") or "").strip().upper()[:64]
        supplier_ref = str(ln.get("supplier_ref") or "").strip().upper()[:64]
        base_code = str(ln.get("base_code") or "").strip().upper()[:64]
        talla = str(ln.get("talla") or "").strip().upper()[:16]

        # Fallback: si la IA no separó base/talla, lo intentamos sobre supplier_ref
        # primero (Patrón B) y luego sobre client_part_number (Patrón A).
        if not base_code or not talla:
            for candidate in (supplier_ref, client_part):
                if not candidate:
                    continue
                b, t = _split_base_and_talla(candidate)
                if not base_code and b: base_code = b
                if not talla    and t: talla    = t
                if base_code and talla: break

        if not (client_part or supplier_ref or base_code):
            continue  # sin nada con qué matchear

        try:
            qty = int(ln.get("qty") or 0)
        except (TypeError, ValueError):
            qty = 0

        out.append({
            "client_part_number": client_part or None,
            "supplier_ref":       supplier_ref or None,
            "base_code":          base_code or None,
            "talla":              talla or None,
            # `sku` se llena después por el resolver (mantenemos la clave para
            # compatibilidad con cross_match y ResolveMatchView._apply_add_line).
            "sku":                "",
            "product_label":      str(ln.get("product_label") or "")[:255],
            "qty":                qty,
            "qty_confirmed":      None,
            "qty_open":           None,
            "unit_price":         _safe_float(ln.get("unit_price")),
            "confidence":         round(max(0.0, min(100.0, float(ln.get("confidence") or 80))), 2),
            # Trazabilidad del match (lo poblará el resolver):
            "match_strategy":     None,
            "match_score":        0,
            "matched_producto_id": None,
        })
    return out


def _normalize_lines(rows, is_sap=False):
    """Normalización para Proforma/SAP. Sprint 2026-05-02 (AG-03):
    capturamos `supplier_ref`, `client_part_number`, `base_code` además
    del `sku` para que el resolver pueda mapear cuando el AI devuelve
    ref proveedor en lugar del SKU MWT canónico (caso típico de
    proformas que importan códigos del proveedor)."""
    out = []
    for ln in (rows or []):
        if not isinstance(ln, dict):
            continue
        sku          = str(ln.get("sku") or "").strip().upper()[:64]
        client_part  = str(ln.get("client_part_number") or "").strip().upper()[:64]
        supplier_ref = str(ln.get("supplier_ref") or "").strip().upper()[:64]
        base_code    = str(ln.get("base_code") or "").strip().upper()[:64]

        # Aceptamos la línea si tiene CUALQUIER identificador. El resolver
        # decidirá después cómo mapearla. Sin ningún identificador → skip.
        if not (sku or client_part or supplier_ref or base_code):
            continue
        try:
            qty_field = "qty_confirmed" if is_sap else "qty"
            qty = int(ln.get(qty_field) or ln.get("qty") or 0)
        except (TypeError, ValueError):
            qty = 0
        out.append({
            "sku":                sku,
            "client_part_number": client_part or None,
            "supplier_ref":       supplier_ref or None,
            "base_code":          base_code or None,
            "product_label":  str(ln.get("product_label") or "")[:255],
            "talla":          str(ln.get("talla") or "").strip().upper()[:16],
            "qty":            qty,
            "qty_confirmed":  qty if is_sap else None,
            "qty_open":       _safe_int(ln.get("qty_open")) if is_sap else None,
            "unit_price":     _safe_float(ln.get("unit_price")),
            "confidence":     round(max(0.0, min(100.0, float(ln.get("confidence") or 80))), 2),
            "match_strategy":     None,
            "match_score":        0,
            "matched_producto_id": None,
        })
    return out


def _safe_int(v):
    try: return int(v) if v not in (None, "") else None
    except (TypeError, ValueError): return None
def _safe_float(v):
    try: return float(v) if v not in (None, "") else None
    except (TypeError, ValueError): return None


# ─────────────────────────────────────────────────────────────────────
# Resolver: IA → SKU canónico vía productos.producto
# ─────────────────────────────────────────────────────────────────────
def _resolve_oc_lines_to_canonical(lines: list[dict], cliente_id=None) -> list[dict]:
    """Para cada línea OC extraída por la IA, resuelve el SKU canónico
    consultando productos.producto. Estrategias en cascada (la primera
    que matchea gana):

      1. SKU directo  — client_part_number == producto.sku
      2. REF proveedor — supplier_ref base == especificaciones->>'ref_proveedor'
      3. Nombre        — base_code == producto.nombre
      4. Fuzzy nombre  — ILIKE %base_code% sobre nombre (score 60)

    Cada línea recibe `sku`, `match_strategy`, `match_score`,
    `matched_producto_id` poblados. Si nada matchea, `sku` queda vacío
    y la UI lo marcará como "no encontrado" para resolución manual.
    """
    if not lines:
        return lines

    # Indexamos productos en memoria una sola vez para esta corrida.
    catalog = _load_catalog_index(cliente_id=cliente_id)

    for ln in lines:
        # Sprint 2026-05-02 (AG-03): el AI puede devolver el SKU MWT canónico
        # directamente en `sku` (caso PROFORMA — el documento es interno y
        # tiene el código MWT). Si está poblado y matchea, ganamos.
        existing_sku = (ln.get("sku") or "").upper()
        if existing_sku and existing_sku in catalog["by_sku"]:
            p = catalog["by_sku"][existing_sku]
            ln["sku"] = p["sku"]
            ln["matched_producto_id"] = p["id"]
            ln["match_strategy"] = "SKU_DIRECT"
            ln["match_score"] = 100
            if not ln.get("product_label"):
                ln["product_label"] = p.get("nombre") or ""
            continue

        client_part = (ln.get("client_part_number") or "").upper()
        supplier_ref = (ln.get("supplier_ref") or "").upper()
        base_code = (ln.get("base_code") or "").upper()

        # 0. ALIAS por cliente (Sprint 2026-05-05). Antes que cualquier
        #    otra heuristica probamos si el alias del producto para ESTE
        #    cliente coincide con la base extraida del documento. Esto
        #    cubre el caso "27BM60B19M-CPAP-MIN-CP-38" donde el cliente
        #    nombra el producto con su prefijo propio: el SKU/nombre del
        #    catalogo MWT no matchea, pero el alias si.
        alias_index = catalog.get("by_alias") or {}
        if alias_index:
            alias_keys = []
            if client_part:
                alias_keys.append(("CLIENT_PART", client_part))
                b, _ = _split_base_and_talla(client_part)
                if b and b != client_part:
                    alias_keys.append(("CLIENT_PART_BASE", b))
            if base_code and not any(k[1] == base_code for k in alias_keys):
                alias_keys.append(("BASE_CODE", base_code))
            if supplier_ref:
                b, _ = _split_base_and_talla(supplier_ref)
                if b and not any(k[1] == b for k in alias_keys):
                    alias_keys.append(("SUPPLIER_REF_BASE", b))
            alias_hit = None
            for source, key in alias_keys:
                if key and key in alias_index:
                    alias_hit = (source, alias_index[key])
                    break
            if alias_hit:
                source, p = alias_hit
                ln["sku"] = p["sku"]
                ln["matched_producto_id"] = p["id"]
                ln["match_strategy"] = f"ALIAS_EXACT/{source}"
                ln["match_score"] = 92
                if not ln.get("product_label"):
                    ln["product_label"] = p.get("nombre") or ""
                continue

        # 1. SKU vía client_part_number (caso OC del cliente — el primer
        #    código de la fila puede coincidir con un SKU MWT)
        if client_part and client_part in catalog["by_sku"]:
            p = catalog["by_sku"][client_part]
            ln["sku"] = p["sku"]
            ln["matched_producto_id"] = p["id"]
            ln["match_strategy"] = "SKU_EXACT"
            ln["match_score"] = 100
            if not ln.get("product_label"):
                ln["product_label"] = p.get("nombre") or ""
            continue

        # Construimos la lista de "claves base" a probar, en orden de prioridad:
        #   supplier_ref base (si lo hay) → base_code → client_part base
        candidates = []
        if supplier_ref:
            b, _ = _split_base_and_talla(supplier_ref)
            if b: candidates.append(("REF_PROVEEDOR", b))
        if base_code and (not candidates or candidates[0][1] != base_code):
            candidates.append(("BASE_CODE", base_code))
        if client_part:
            b, _ = _split_base_and_talla(client_part)
            if b and all(b != c[1] for c in candidates):
                candidates.append(("CLIENT_PART_BASE", b))

        matched = False
        for source, key in candidates:
            # 2. Ref proveedor
            if key in catalog["by_ref_proveedor"]:
                p = catalog["by_ref_proveedor"][key]
                ln["sku"] = p["sku"]
                ln["matched_producto_id"] = p["id"]
                ln["match_strategy"] = f"REF_PROVEEDOR/{source}"
                ln["match_score"] = 95
                if not ln.get("product_label"):
                    ln["product_label"] = p.get("nombre") or ""
                matched = True
                break
            # 3. Nombre exacto
            if key in catalog["by_nombre"]:
                p = catalog["by_nombre"][key]
                ln["sku"] = p["sku"]
                ln["matched_producto_id"] = p["id"]
                ln["match_strategy"] = f"NOMBRE_EXACT/{source}"
                ln["match_score"] = 90
                if not ln.get("product_label"):
                    ln["product_label"] = p.get("nombre") or ""
                matched = True
                break

        if matched:
            continue

        # 4. Fuzzy nombre — solo si tenemos algún base_code
        if base_code:
            fuzzy = _fuzzy_lookup_nombre(base_code)
            if fuzzy:
                ln["sku"] = fuzzy["sku"]
                ln["matched_producto_id"] = fuzzy["id"]
                ln["match_strategy"] = "NOMBRE_FUZZY"
                ln["match_score"] = 60
                if not ln.get("product_label"):
                    ln["product_label"] = fuzzy.get("nombre") or ""
                continue

        # 5. SUBSTRING fallback (Sprint 2026-05-02 / AG-03):
        # La red de seguridad cuando la IA se equivoca al separar las
        # columnas ARTICULO vs REF.PROVEEDOR. Si el client_part_number o
        # base_code CONTIENE alguna ref_proveedor o nombre del catálogo
        # como substring, lo aceptamos. Esto cubre el caso típico:
        #   AI extrae base_code = "27BM60B19M-CPAP-MIN-CP" (con prefijo
        #   "27BM" del cliente) cuando el catálogo tiene
        #   ref_proveedor = "60B19M-CPAP-MIN-CP". Substring match wins.
        substr_haystacks = []
        if base_code:
            substr_haystacks.append((base_code, "BASE_CODE"))
        cp = (ln.get("client_part_number") or "").upper()
        if cp:
            substr_haystacks.append((cp, "CLIENT_PART"))

        substring_matched = False
        for haystack, source in substr_haystacks:
            if len(haystack) < 6:
                continue
            # 4.0 Substring contra ALIAS del cliente (Sprint 2026-05-05).
            # Cubre cuando la IA pegó la talla al alias o le quedó algún
            # prefijo extra. Score 78 — mejor que ref_proveedor substring
            # porque el alias está fijado por el CEO específicamente para
            # este cliente.
            for alias_key, p in (catalog.get("by_alias") or {}).items():
                if alias_key and len(alias_key) >= 6 and alias_key in haystack:
                    ln["sku"] = p["sku"]
                    ln["matched_producto_id"] = p["id"]
                    ln["match_strategy"] = f"ALIAS_SUBSTRING/{source}"
                    ln["match_score"] = 78
                    if not ln.get("product_label"):
                        ln["product_label"] = p.get("nombre") or ""
                    substring_matched = True
                    break
            if substring_matched:
                break
            # Probamos primero por ref_proveedor (más confiable para OC)
            for ref_key, p in catalog["by_ref_proveedor"].items():
                if ref_key and len(ref_key) >= 6 and ref_key in haystack:
                    ln["sku"] = p["sku"]
                    ln["matched_producto_id"] = p["id"]
                    ln["match_strategy"] = f"REF_SUBSTRING/{source}"
                    ln["match_score"] = 75
                    if not ln.get("product_label"):
                        ln["product_label"] = p.get("nombre") or ""
                    substring_matched = True
                    break
            if substring_matched:
                break
            # Después por nombre
            for nombre_key, p in catalog["by_nombre"].items():
                if nombre_key and len(nombre_key) >= 6 and nombre_key in haystack:
                    ln["sku"] = p["sku"]
                    ln["matched_producto_id"] = p["id"]
                    ln["match_strategy"] = f"NOMBRE_SUBSTRING/{source}"
                    ln["match_score"] = 70
                    if not ln.get("product_label"):
                        ln["product_label"] = p.get("nombre") or ""
                    substring_matched = True
                    break
            if substring_matched:
                break

        if substring_matched:
            continue

        # No match: marcamos para resolución manual.
        ln["match_strategy"] = "UNRESOLVED"
        ln["match_score"] = 0

    return lines


def _load_catalog_index(cliente_id=None) -> dict:
    """Carga índices en memoria de productos.producto activos.

    Sprint 2026-05-05 (AG-03): si recibe `cliente_id`, también carga
    productos.product_client_alias activo para ese cliente y arma un
    `by_alias` con la misma forma que `by_sku` (alias UPPER → row del
    producto). Esto deja que el matchmaker resuelva OCs donde el cliente
    nombra el producto con su propio "alias comercial" en vez del SKU
    o nombre canónico MWT.
    """
    by_sku = {}
    by_nombre = {}
    by_ref_proveedor = {}
    by_alias = {}
    try:
        with connection.cursor() as c:
            c.execute("""
                SELECT
                    id::text,
                    UPPER(COALESCE(sku, ''))                                   AS sku,
                    UPPER(COALESCE(nombre, ''))                                AS nombre,
                    UPPER(COALESCE(especificaciones->>'ref_proveedor', ''))    AS ref_proveedor
                  FROM productos.producto
                 WHERE COALESCE(is_active, TRUE) = TRUE
            """)
            for pid, sku, nombre, ref in c.fetchall():
                row = {"id": pid, "sku": sku, "nombre": nombre, "ref_proveedor": ref}
                if sku:    by_sku[sku] = row
                if nombre: by_nombre[nombre] = row
                if ref:    by_ref_proveedor[ref] = row
    except Exception as e:
        log.warning("[matchmaker] no pude cargar catalogo de productos: %s", e)

    # ── Aliases por cliente (LOTE_SM_TICKETS · alias por cliente) ─────
    # Sólo carga si el caller nos da un cliente_id (sólo en OC, nunca
    # en proforma — proforma_extractor no consume este loader).
    if cliente_id:
        try:
            with connection.cursor() as c:
                c.execute("""
                    SELECT a.producto_id::text,
                           UPPER(COALESCE(a.alias, ''))     AS alias_upper,
                           UPPER(COALESCE(p.sku, ''))       AS sku,
                           UPPER(COALESCE(p.nombre, ''))    AS nombre,
                           UPPER(COALESCE(p.especificaciones->>'ref_proveedor', '')) AS ref
                      FROM productos.product_client_alias a
                      JOIN productos.producto p ON p.id = a.producto_id
                     WHERE a.is_active = TRUE
                       AND a.cliente_id = %s::uuid
                       AND COALESCE(p.is_active, TRUE) = TRUE
                """, [str(cliente_id)])
                for pid, alias_up, sku, nombre, ref in c.fetchall():
                    if not alias_up:
                        continue
                    by_alias[alias_up] = {
                        "id": pid,
                        "sku": sku,
                        "nombre": nombre,
                        "ref_proveedor": ref,
                        "alias": alias_up,
                    }
        except Exception as e:
            log.warning("[matchmaker] no pude cargar aliases del cliente %s: %s",
                        cliente_id, e)

    return {
        "by_sku":            by_sku,
        "by_nombre":         by_nombre,
        "by_ref_proveedor":  by_ref_proveedor,
        "by_alias":          by_alias,
    }


def _build_talla_equivalence_map() -> dict:
    """Sprint 2026-05-02 (AG-03): construye un mapa de equivalencia de tallas
    cross-convención (BR/EU/US/UK/CM) basado en ops.tallas.

    Cada fila de ops.tallas representa una talla física con valores en
    múltiples sistemas. Por ejemplo: (talla_base=39, eu=39, br=37, us_men=6-6.5,
    uk_men=5.5-6, cm=25.30). Todos esos valores son LA MISMA talla, sólo
    expresada distinto.

    Devuelve: {value → set de equivalentes (incluyendo a sí mismo)}.
    Cuando un valor aparece en MÚLTIPLES filas (ej. "37" como EU 37 y como
    BR 37 en filas distintas), los equivalentes se agregan de TODAS esas
    filas — el matcher acepta cualquier match.

    Caso real que esto resuelve: BD viene con BR (`37`-`43`, sufijos de
    códigos MARLUVAS) y la proforma extrae en EU canónica (`39`-`45`).
    Sin equivalencia, no matchean. Con equivalencia, la fila EU 39 incluye
    BR 37 → match.
    """
    equiv = {}
    try:
        with connection.cursor() as c:
            c.execute("""
                SELECT talla_base, eu, us_men, us_women, uk_men, uk_women, br, cm
                  FROM ops.tallas
                 WHERE COALESCE(is_active, TRUE) = TRUE
                   AND tipo_producto = 'calzado'
            """)
            for row in c.fetchall():
                vals = {str(v).strip().upper() for v in row if v and str(v).strip()}
                for v in vals:
                    equiv.setdefault(v, set()).update(vals)
    except Exception as e:
        log.warning("[matchmaker] talla equivalence map build failed: %s", e)
    return equiv


def _find_db_with_talla_equiv(sku, doc_talla, db_index, equiv_map):
    """Busca (sku, doc_talla) en db_index; si falla, prueba con tallas
    equivalentes según ops.tallas. Devuelve (key_que_matched, db_row) o
    (None, None) si nada matchea.

    El exact match siempre tiene prioridad — la equivalencia es fallback.
    """
    if not sku:
        return None, None
    key = (sku, doc_talla)
    if key in db_index:
        return key, db_index[key]
    # Fallback por equivalencia de talla
    for alt_talla in equiv_map.get(doc_talla, set()):
        if alt_talla == doc_talla:
            continue
        alt_key = (sku, alt_talla)
        if alt_key in db_index:
            return alt_key, db_index[alt_key]
    return None, None


def _find_best_db_match_proforma(sku, doc_talla, doc_qty, db_index, equiv_map):
    """Sprint 2026-05-02 (AG-03): match SMART para PROFORMA.

    Problema que resuelve: el AI suele extraer la proforma en EU canónica
    (39, 40, 41, ...) mientras que la BD del expediente tiene BR (37, 38,
    39, ...). El número '41' existe como EU 41 (= BR 39 físicamente) y
    también como BR 41 (= EU 43 físicamente). Match literal cruza '41'
    con '41' aunque sean tallas físicas distintas → reporta QTY_DIFF
    espurio (doc qty 10 vs BD qty 30).

    Estrategia: para cada doc line, recolectar TODOS los candidatos de
    BD (literal + equivalentes via ops.tallas) y elegir el que mejor
    matchea por CANTIDAD. Si una equivalencia coincide en qty y la
    literal no, la equivalencia gana — eso significa que el AI extrajo
    en otra convención que la BD pero ambas refieren a la misma talla
    física con la misma cantidad.

    Devuelve (matched_key, db_row, qty_matches_bool) o (None, None, False).
    """
    if not sku:
        return None, None, False

    # 1) Recolectar TODOS los candidatos (literal + equivalentes)
    candidate_keys = [(sku, doc_talla)]
    for alt_talla in equiv_map.get(doc_talla, set()):
        if alt_talla != doc_talla:
            candidate_keys.append((sku, alt_talla))

    candidates = [
        (k, db_index[k]) for k in candidate_keys if k in db_index
    ]
    if not candidates:
        return None, None, False

    # 2) Si solo hay 1 candidato, devolverlo (con flag de si qty matchea)
    if len(candidates) == 1:
        k, db = candidates[0]
        return k, db, int(db.get("qty", 0)) == int(doc_qty)

    # 3) Múltiples candidatos: preferir el que MATCHEA en qty.
    #    Esto desambigua "41 EU" vs "41 BR" cuando ambos están en BD.
    qty_doc = int(doc_qty or 0)
    for k, db in candidates:
        if int(db.get("qty", 0)) == qty_doc:
            return k, db, True

    # 4) Ningún candidato matchea en qty → devolver el LITERAL como
    #    fallback (mantiene comportamiento original — reporta QTY_DIFF)
    for k, db in candidates:
        if k[1] == doc_talla:
            return k, db, False
    # Si ni el literal está, devolver el primer candidato disponible
    k, db = candidates[0]
    return k, db, False


def _fuzzy_lookup_nombre(base_code: str) -> Optional[dict]:
    """Busca por ILIKE %base_code% en productos.producto.nombre.
    Devuelve el match con mejor score (longitud del match relativa)."""
    if not base_code or len(base_code) < 4:
        return None
    try:
        with connection.cursor() as c:
            c.execute("""
                SELECT id::text, sku, nombre
                  FROM productos.producto
                 WHERE COALESCE(is_active, TRUE) = TRUE
                   AND UPPER(COALESCE(nombre, '')) LIKE %s
                 ORDER BY LENGTH(nombre) ASC
                 LIMIT 1
            """, [f"%{base_code}%"])
            row = c.fetchone()
            if row:
                return {"id": row[0], "sku": row[1], "nombre": row[2]}
    except Exception as e:
        log.warning("[matchmaker] fuzzy lookup falló: %s", e)
    return None


# ─────────────────────────────────────────────────────────────────────
# Matchmaker — cruce IA vs líneas del expediente
# ─────────────────────────────────────────────────────────────────────
def cross_match(ai_payload: dict, expediente_id) -> dict:
    """Cruza el payload de la IA contra expedientes.linea del expediente."""
    kind = (ai_payload.get("document_kind") or "OC").upper()

    # ── Resolver IA→SKU canónico ─────────────
    # Sprint 2026-05-02 (AG-03): aplicamos el resolver a OC, PROFORMA y SAP.
    #   · OC del Cliente:  los códigos vienen del cliente, no son SKU MWT.
    #   · PROFORMA MWT:    suele traer SKU canónico, pero defendemos
    #                       por si solo viene supplier_ref.
    #   · SAP del proveedor: las confirmaciones traen ref del proveedor,
    #                       no el SKU MWT.
    # Sprint 2026-05-05 (AG-03): para OC consultamos el cliente del
    # expediente y se lo pasamos al resolver para que pueda usar
    # productos.product_client_alias en la cascada de match. PROFORMA y
    # SAP siguen sin alias (decisión: solo OC del cliente usa esa señal).
    cliente_id = None
    try:
        with connection.cursor() as c:
            c.execute(
                "SELECT client_id::text FROM expedientes.expediente WHERE id = %s::uuid",
                [str(expediente_id)],
            )
            row = c.fetchone()
            if row and row[0]:
                cliente_id = row[0]
    except Exception as e:
        log.warning("[matchmaker] cross_match: no pude leer client_id del expediente: %s", e)

    if kind == "OC":
        ai_payload["lines"] = _resolve_oc_lines_to_canonical(
            ai_payload.get("lines") or [], cliente_id=cliente_id,
        )
    elif kind in ("PROFORMA", "SAP"):
        for g in (ai_payload.get("groups") or []):
            g["lines"] = _resolve_oc_lines_to_canonical(g.get("lines") or [])

    db_lines = _load_expediente_lines(expediente_id)
    db_index = {}
    for l in db_lines:
        key = (l["sku"].upper(), (l["talla"] or "").upper())
        if key in db_index:
            db_index[key]["qty"] += l["qty"]
        else:
            db_index[key] = dict(l)

    # Sprint 2026-05-02 (AG-03): mapa de equivalencia de tallas. Usado como
    # fallback cuando (sku, doc_talla) no matchea exactamente — permite que
    # un doc con talla EU (ej. 39) matchee con BD que tiene BR (ej. 37) si
    # ambas son la misma fila de ops.tallas.
    talla_equiv = _build_talla_equivalence_map()

    discrepancies = []
    matched_keys = set()

    if kind == "OC":
        for ln in (ai_payload.get("lines") or []):
            sku = (ln.get("sku") or "").upper()
            talla = (ln.get("talla") or "").upper()

            # Línea no resuelta — UNRESOLVED, requiere intervención manual.
            if not sku:
                discrepancies.append({
                    "kind":             "UNRESOLVED_PRODUCT",
                    "sku":              None,
                    "talla":            talla or None,
                    "qty_doc":          ln.get("qty") or 0,
                    "qty_exp":          0,
                    "sap_doc":          None,
                    "sap_exp":          None,
                    "severity":         "ERROR",
                    "suggested_action": "MANUAL",
                    "product_label":    ln.get("product_label"),
                    "client_part_number": ln.get("client_part_number"),
                    "supplier_ref":     ln.get("supplier_ref"),
                    "base_code":        ln.get("base_code"),
                    "match_strategy":   ln.get("match_strategy"),
                    "match_score":      ln.get("match_score"),
                    "confidence":       ln.get("confidence"),
                    "unit_price":       ln.get("unit_price"),
                })
                continue

            key = (sku, talla)
            db = db_index.get(key)
            if not db:
                discrepancies.append({
                    "kind":             "MISSING_IN_EXPEDIENTE",
                    "sku":              sku,
                    "talla":            talla or None,
                    "qty_doc":          ln.get("qty") or 0,
                    "qty_exp":          0,
                    "sap_doc":          None,
                    "sap_exp":          None,
                    "severity":         "WARN",
                    "suggested_action": "ADD_LINE",
                    "product_label":    ln.get("product_label"),
                    "client_part_number": ln.get("client_part_number"),
                    "supplier_ref":     ln.get("supplier_ref"),
                    "base_code":        ln.get("base_code"),
                    "match_strategy":   ln.get("match_strategy"),
                    "match_score":      ln.get("match_score"),
                    "confidence":       ln.get("confidence"),
                    "unit_price":       ln.get("unit_price"),
                })
            else:
                matched_keys.add(key)
                if int(db["qty"]) != int(ln.get("qty") or 0):
                    discrepancies.append({
                        "kind":             "QTY_DIFF",
                        "sku":              sku,
                        "talla":            talla or None,
                        "qty_doc":          ln.get("qty") or 0,
                        "qty_exp":          int(db["qty"]),
                        "sap_doc":          None,
                        "sap_exp":          db.get("sap"),
                        "severity":         "WARN",
                        "suggested_action": "UPDATE_QTY",
                        "product_label":    ln.get("product_label") or db.get("product_label"),
                        "match_strategy":   ln.get("match_strategy"),
                        "match_score":      ln.get("match_score"),
                        "line_id":          db.get("line_id"),
                    })
        for key, db in db_index.items():
            if key in matched_keys:
                continue
            discrepancies.append({
                "kind":             "MISSING_IN_DOC",
                "sku":              db["sku"],
                "talla":            db["talla"] or None,
                "qty_doc":          0,
                "qty_exp":          int(db["qty"]),
                "severity":         "INFO",
                "suggested_action": "MANUAL",
                "product_label":    db.get("product_label"),
                "line_id":          db.get("line_id"),
            })
        groups_out = []

    else:
        # ── Proforma / SAP (sin cambios) ────────────────────
        groups_out = []
        for g in (ai_payload.get("groups") or []):
            sap = g.get("sap_number")
            g_disc = []
            for ln in g.get("lines") or []:
                sku_norm = (ln.get("sku") or "").upper()
                talla_norm = (ln.get("talla") or "").upper()
                qty_doc = int(ln.get("qty_confirmed") or ln.get("qty") or 0)

                # Sprint 2026-05-02 (AG-03): si el resolver no logró mapear
                # la línea a un SKU canónico, emitimos UNRESOLVED_PRODUCT
                # en vez de matchear contra (sku="", talla) — eso confundía
                # con MISSING_IN_EXPEDIENTE genuinos.
                if not sku_norm:
                    item = {
                        "kind":               "UNRESOLVED_PRODUCT",
                        "sku":                None,
                        "talla":              talla_norm or None,
                        "qty_doc":            qty_doc,
                        "qty_exp":            0,
                        "sap_doc":            sap,
                        "sap_exp":            None,
                        "severity":           "ERROR",
                        "suggested_action":   "MANUAL",
                        "product_label":      ln.get("product_label"),
                        "client_part_number": ln.get("client_part_number"),
                        "supplier_ref":       ln.get("supplier_ref"),
                        "base_code":          ln.get("base_code"),
                        "match_strategy":     ln.get("match_strategy"),
                        "match_score":        ln.get("match_score"),
                        "confidence":         ln.get("confidence"),
                    }
                    g_disc.append(item); discrepancies.append(item)
                    continue

                # Sprint 2026-05-02 (AG-03): MATCH LITERAL para PROFORMA.
                # Sin equivalencias. Si doc dice (sku=701935, talla=41) y BD
                # tiene esa key → match. Si no → discrepancia.
                # La responsabilidad de extraer la talla en el formato correcto
                # recae 100% en el prompt del AI (que debe leer fila BR).
                key = (sku_norm, talla_norm)
                db = db_index.get(key)
                matched_key = key
                if not db:
                    item = {
                        "kind":             "MISSING_IN_EXPEDIENTE",
                        "sku":              sku_norm,
                        "talla":            talla_norm or None,
                        "qty_doc":          qty_doc,
                        "qty_exp":          0,
                        "sap_doc":          sap,
                        "sap_exp":          None,
                        "severity":         "WARN",
                        "suggested_action": "ADD_LINE",
                        "product_label":    ln.get("product_label"),
                        "match_strategy":   ln.get("match_strategy"),
                        "match_score":      ln.get("match_score"),
                        "confidence":       ln.get("confidence"),
                        "unit_price":       ln.get("unit_price"),
                    }
                    g_disc.append(item); discrepancies.append(item)
                else:
                    matched_keys.add(matched_key)
                    delta_qty = int(db["qty"]) != qty_doc
                    sap_diff  = db.get("sap") and sap and (str(db["sap"]).strip() != str(sap).strip())
                    if delta_qty or sap_diff:
                        # Si el match fue por equivalencia, mostramos la talla
                        # canónica del BD (la real que vamos a actualizar) en
                        # lugar de la del doc — evita confundir al usuario.
                        item = {
                            "kind":             "SAP_MISMATCH" if sap_diff else "QTY_DIFF",
                            "sku":              sku_norm,
                            "talla":            db.get("talla") or talla_norm or None,
                            "qty_doc":          qty_doc,
                            "qty_exp":          int(db["qty"]),
                            "sap_doc":          sap,
                            "sap_exp":          db.get("sap"),
                            "severity":         "ERROR" if sap_diff else "WARN",
                            "suggested_action": "ATTACH_SAP" if sap_diff else "UPDATE_QTY",
                            "product_label":    ln.get("product_label") or db.get("product_label"),
                            "match_strategy":   ln.get("match_strategy"),
                            "match_score":      ln.get("match_score"),
                            "line_id":          db.get("line_id"),
                        }
                        g_disc.append(item); discrepancies.append(item)
            groups_out.append({
                "sap_number":     sap,
                "po_reference":   g.get("po_reference"),
                "delivery_date":  g.get("delivery_date"),
                "lines_count":    len(g.get("lines") or []),
                "discrepancies":  g_disc,
            })
        for key, db in db_index.items():
            if key in matched_keys:
                continue
            discrepancies.append({
                "kind":             "MISSING_IN_DOC",
                "sku":              db["sku"],
                "talla":            db["talla"] or None,
                "qty_doc":          0,
                "qty_exp":          int(db["qty"]),
                "sap_doc":          None,
                "sap_exp":          db.get("sap"),
                "severity":         "INFO",
                "suggested_action": "MANUAL",
                "product_label":    db.get("product_label"),
                "line_id":          db.get("line_id"),
            })

    lines_in_doc = (
        len(ai_payload.get("lines") or [])
        if kind == "OC"
        else sum(len(g.get("lines") or []) for g in (ai_payload.get("groups") or []))
    )
    lines_in_exp   = len(db_lines)
    matched        = len(matched_keys)
    discrepancies_count = len(discrepancies)
    coverage = round((matched / lines_in_exp) * 100, 2) if lines_in_exp > 0 else (
        100.0 if lines_in_doc == 0 else 0.0
    )
    perfect = (discrepancies_count == 0 and lines_in_doc > 0)

    # Métricas extras de resolución (OC, PROFORMA, SAP)
    # Sprint 2026-05-02 (AG-03): aplicamos al universo correcto de líneas.
    if kind == "OC":
        all_lines = ai_payload.get("lines") or []
    elif kind in ("PROFORMA", "SAP"):
        all_lines = []
        for g in (ai_payload.get("groups") or []):
            all_lines.extend(g.get("lines") or [])
    else:
        all_lines = []

    resolution_summary = None
    if all_lines:
        total = len(all_lines)
        resolved = sum(1 for l in all_lines if l.get("sku"))
        unresolved = total - resolved
        resolution_summary = {
            "lines_total":      total,
            "lines_resolved":   resolved,
            "lines_unresolved": unresolved,
            "resolution_pct":   round((resolved / total) * 100, 2) if total else 100.0,
        }

    return {
        "summary": {
            "perfect_match":       perfect,
            "coverage_pct":        coverage,
            "lines_in_doc":        lines_in_doc,
            "lines_in_expediente": lines_in_exp,
            "lines_matched":       matched,
            "discrepancies_count": discrepancies_count,
            "resolution":          resolution_summary,
        },
        "discrepancies": discrepancies,
        "groups":        groups_out,
    }


def _load_expediente_lines(expediente_id) -> list[dict]:
    """Carga líneas del expediente (sku, talla, qty, sap)."""
    rows = []
    try:
        with connection.cursor() as c:
            c.execute("""
                SELECT
                    l.id::text                        AS line_id,
                    COALESCE(l.sku, '')               AS sku,
                    COALESCE(l.size, '')              AS talla,
                    COALESCE(l.qty, 0)                AS qty,
                    COALESCE(l.sap, '')               AS sap,
                    COALESCE(p.nombre, l.sku, '')     AS product_label
                FROM   expedientes.linea l
                LEFT JOIN productos.producto p ON p.id = l.producto_id
                WHERE  l.expediente_id = %s
                  AND  COALESCE(l.is_active, TRUE) = TRUE
            """, [str(expediente_id)])
            cols = [c.description[i][0] for i in range(len(c.description))]
            for r in c.fetchall():
                rows.append(dict(zip(cols, r)))
    except Exception as e:
        log.warning("[matchmaker] no pude leer expedientes.linea: %s", e)
    return rows