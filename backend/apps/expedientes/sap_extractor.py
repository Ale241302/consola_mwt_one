# backend/apps/expedientes/sap_extractor.py
"""
=====================================================================
MWT.ONE · apps.expedientes.sap_extractor
Agente responsable: [AG-BACKEND]

Sprint 2026-05-04 · análisis IA + parser determinístico de
Confirmación SAP (Marluvas) — paralelo al matchmaker de OC/Proforma
pero ESPECIALIZADO para el flujo de "Agregar SAP" en el drawer
AddSAPConfirmationDrawer.

Se invoca desde:  POST /api/expedientes/{id}/analyze-sap-confirmation/

Para xlsx (formato exportado por SAP de Marluvas) usa parser
determinístico con openpyxl — es 100% confiable porque las columnas
son estables: "Documento de vendas" (SAP #), "Material" (SKU-talla),
"Descrição de material" (nombre + talla), "Pares Aberto" (qty).

Para PDF cae al `extract_document` con SYSTEM_PROMPT_SAP de la IA.

Salida estructurada:
{
  "ok":          True,
  "kind":        "xlsx_marluvas" | "pdf_ai",
  "filename":    "269486.xlsx",
  "sap_id":      "269486",
  "lineas":      [{ sku, talla, qty, descripcion, raw_material,
                    matched_line_id, match: {...} }, ...],
  "discrepancies": [{ kind, severity, sku, talla, qty_doc, qty_exp, ... }],
  "summary":     { lines_in_doc, lines_matched, lines_unmatched,
                   discrepancies_count, perfect_match }
}

Reglas duras (R1, R3, R4, política Confirmación SAP):
  · El SAP se determina del documento (col B "Documento de vendas").
    Si todas las filas tienen el mismo SAP, ese es el sap_id.
    Si hay varios, devolvemos el primero y reportamos `multiple_saps`.
  · El frontend luego permite al usuario filtrar SOLO las líneas que
    correspondan a `sap_id` y agregarlas al SAP que se está
    registrando — las que no estén en este SAP simplemente NO
    forman parte de este SAP (puede haber más SAPs en futuros uploads).
  · NUNCA inventamos líneas. Solo reportamos lo que el documento dice
    + el cruce contra expedientes.linea.
=====================================================================
"""
from __future__ import annotations

import io
import logging
import re
from typing import Optional

from django.db import connection

log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────
# Regex helpers
# ─────────────────────────────────────────────────────────────────────

# "700728-38" → ("700728", "38")
_RE_MATERIAL = re.compile(r"^([0-9A-Z]+)[-_]([0-9A-Z]{1,4})$", re.IGNORECASE)

# "75BPR29-MSMC-CPAP-ST, 38" → ("75BPR29-MSMC-CPAP-ST", "38")
_RE_DESC_TALLA = re.compile(r"^(.+?)[,\s]+([0-9A-Z]{1,4})\s*$", re.IGNORECASE)


def _split_material(material: str) -> tuple[Optional[str], Optional[str]]:
    """Parte 'SKU-TALLA' (ej. '700728-38') en ('700728', '38').

    Defensivo: si no encaja con el patrón devuelve (material, None).
    Soporta separador '-' o '_'.
    """
    if not material:
        return (None, None)
    s = str(material).strip().upper().replace(" ", "")
    m = _RE_MATERIAL.match(s)
    if m:
        return (m.group(1) or None, m.group(2) or None)
    # Fallback: split por '-' tomando lo último como talla si parece numérico
    if "-" in s:
        head, _, tail = s.rpartition("-")
        if tail and tail.isdigit():
            return (head or None, tail or None)
    return (s or None, None)


def _split_descripcion(desc: str) -> tuple[Optional[str], Optional[str]]:
    """Parte '75BPR29-MSMC-CPAP-ST, 38' en ('75BPR29-MSMC-CPAP-ST', '38').

    Útil para extraer el NOMBRE del producto (sin la talla suffix) y
    cruzarlo contra productos.producto.nombre.
    """
    if not desc:
        return (None, None)
    s = str(desc).strip()
    m = _RE_DESC_TALLA.match(s)
    if m:
        return (m.group(1).strip() or None, m.group(2).strip().upper() or None)
    return (s or None, None)


def _safe_int(v) -> int:
    try:
        if v is None or v == "":
            return 0
        return int(float(v))
    except (TypeError, ValueError):
        return 0


# ─────────────────────────────────────────────────────────────────────
# Parser determinístico — Marluvas xlsx
# ─────────────────────────────────────────────────────────────────────

# Headers conocidos del export Marluvas (case/locale-tolerant matcher).
_HEADER_ALIASES = {
    "sap":         ("documento de vendas", "documento de venda", "sales document",
                    "sales order", "auftrag"),
    "material":    ("material", "sku", "código material", "codigo material",
                    "product code"),
    "descripcion": ("descrição de material", "descripcion de material",
                    "description", "descricao de material", "material description"),
    "qty":         ("pares aberto", "qtd ordem", "qty open", "qty",
                    "cantidad", "quantity", "pares"),
}


def _norm_header(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def _detect_columns(header_row: list) -> dict:
    """Mapea headers reales → keys canónicas (sap, material, descripcion, qty).

    Devuelve {key_canon: col_index} solo para los headers que encontró.
    """
    found = {}
    for idx, raw in enumerate(header_row or []):
        h = _norm_header(raw)
        if not h:
            continue
        for canon, aliases in _HEADER_ALIASES.items():
            if canon in found:
                continue
            for a in aliases:
                if a == h or a in h:
                    found[canon] = idx
                    break
    return found


def _parse_xlsx_marluvas(file_bytes: bytes) -> dict:
    """Parsea xlsx exportado por SAP Marluvas.

    Estructura conocida:
      Row 1 = headers (incluye "Documento de vendas", "Material",
              "Descrição de material", "Pares Aberto", ...)
      Row 2..N = filas de producto
      Row N+1  = TOTAL (queda con material/descripcion vacíos pero qty
                 con la suma) → la filtramos por material vacío.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {
            "ok": False,
            "error": "openpyxl no instalado",
            "lineas": [],
            "sap_id": None,
        }

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        return {
            "ok": False,
            "error": f"No pude abrir el xlsx: {e}",
            "lineas": [],
            "sap_id": None,
        }

    # Iterar todas las hojas (Marluvas exporta a "Data" pero tolerar variantes)
    cols = {}
    rows_iter = None
    for ws in wb.worksheets:
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first:
            continue
        cmap = _detect_columns(list(first))
        if "material" in cmap and ("qty" in cmap or "descripcion" in cmap):
            cols = cmap
            rows_iter = list(ws.iter_rows(min_row=2, values_only=True))
            break

    if not cols or rows_iter is None:
        return {
            "ok": False,
            "error": "No encontré la hoja con los headers Marluvas (Material/Pares Aberto).",
            "lineas": [],
            "sap_id": None,
        }

    sap_set = set()
    lineas = []
    for row in rows_iter:
        if not row:
            continue
        # SAP #
        sap_val = None
        if "sap" in cols:
            v = row[cols["sap"]]
            if v not in (None, ""):
                sap_val = str(v).strip()
                # Normalizar SAP a string sin '.0' si vino como float
                if sap_val.endswith(".0"):
                    sap_val = sap_val[:-2]

        # Material → sku, talla
        material_raw = None
        sku, talla = (None, None)
        if "material" in cols:
            material_raw = row[cols["material"]]
            sku, talla = _split_material(material_raw)

        # Descripción → nombre (sin talla)
        descripcion_raw = None
        nombre, talla_desc = (None, None)
        if "descripcion" in cols:
            descripcion_raw = row[cols["descripcion"]]
            nombre, talla_desc = _split_descripcion(descripcion_raw)

        if not talla and talla_desc:
            talla = talla_desc

        # Qty
        qty = 0
        if "qty" in cols:
            qty = _safe_int(row[cols["qty"]])

        # Filtros: descartar fila TOTAL (sin material) y filas vacías
        if not material_raw and not descripcion_raw:
            continue
        if not sku and not nombre:
            continue

        if sap_val:
            sap_set.add(sap_val)

        lineas.append({
            "sap_doc":      sap_val,
            "sku":          (sku or "").upper() or None,
            "talla":        (talla or "").upper() or None,
            "qty":          qty,
            "descripcion":  (nombre or "").strip() or None,
            "raw_material": (str(material_raw or "") if material_raw else None),
        })

    sap_list = sorted(sap_set)
    sap_id = sap_list[0] if sap_list else None

    return {
        "ok":           True,
        "kind":         "xlsx_marluvas",
        "sap_id":       sap_id,
        "sap_count":    len(sap_list),
        "all_saps":     sap_list,
        "lineas":       lineas,
        "error":        None,
    }


# ─────────────────────────────────────────────────────────────────────
# Cruce contra expedientes.linea (semejante al cross_match pero
# adaptado al universo SAP, con foco en VALIDACIÓN: SKU, NOMBRE, TALLA,
# QTY).
# ─────────────────────────────────────────────────────────────────────

def _load_expediente_lines_with_name(expediente_id) -> list[dict]:
    """Carga líneas del expediente con nombre real del producto.

    Usado para validar que el `nombre` que viene en la descripción del
    SAP matchee con el catálogo MWT.
    """
    rows = []
    try:
        with connection.cursor() as c:
            c.execute("""
                SELECT
                    l.id::text                            AS line_id,
                    UPPER(COALESCE(l.sku, ''))            AS sku,
                    UPPER(COALESCE(l.size, ''))           AS talla,
                    COALESCE(l.qty, 0)::int               AS qty,
                    UPPER(COALESCE(l.sap, ''))            AS sap,
                    COALESCE(p.nombre, l.sku, '')         AS nombre,
                    UPPER(COALESCE(p.nombre, l.sku, ''))  AS nombre_upper
                FROM   expedientes.linea l
                LEFT JOIN productos.producto p ON p.id = l.producto_id
                WHERE  l.expediente_id = %s
                  AND  COALESCE(l.is_active, TRUE) = TRUE
            """, [str(expediente_id)])
            cols = [c.description[i][0] for i in range(len(c.description))]
            for r in c.fetchall():
                rows.append(dict(zip(cols, r)))
    except Exception as e:
        log.warning("[sap_extractor] no pude leer expedientes.linea: %s", e)
    return rows


def cross_match_sap(extracted: dict, expediente_id) -> dict:
    """Cruza el resultado del extractor SAP contra expedientes.linea.

    Reglas de match:
      · Clave primaria: (sku, talla).
      · Si matchea: validamos qty, validamos nombre (advertencia
        soft si difiere — no bloqueante).
      · Si no matchea: MISSING_IN_EXPEDIENTE (severidad WARN — el
        producto figura en la confirmación SAP pero no en el
        expediente; el operador decide si es un add legítimo o un
        error del documento).
      · Líneas del expediente que no aparecen en el documento NO
        emiten discrepancia: en el flujo SAP, no todas las líneas
        del expediente entran en cada SAP. Las ausentes solo se
        reportan a nivel summary (lines_pending_in_exp).

    Devuelve `extracted` enriquecido con `match` por línea +
    `discrepancies` global + `summary`.
    """
    if not extracted.get("ok"):
        return {**extracted, "discrepancies": [], "summary": {
            "lines_in_doc": 0, "lines_matched": 0, "lines_unmatched": 0,
            "discrepancies_count": 0, "perfect_match": False,
        }}

    db_lines = _load_expediente_lines_with_name(expediente_id)
    db_index = {}
    for l in db_lines:
        key = (l["sku"], l["talla"])
        if key in db_index:
            db_index[key]["qty"] += int(l["qty"])
        else:
            db_index[key] = dict(l)

    discrepancies = []
    matched_keys = set()

    for ln in extracted["lineas"]:
        sku = (ln.get("sku") or "").upper()
        talla = (ln.get("talla") or "").upper()
        qty_doc = int(ln.get("qty") or 0)
        nombre_doc = (ln.get("descripcion") or "").upper().strip()

        if not sku or not talla:
            ln["match"] = {
                "matched": False,
                "reason": "INCOMPLETE_KEY",
                "line_id": None,
            }
            discrepancies.append({
                "kind":             "INCOMPLETE_KEY",
                "severity":         "ERROR",
                "sku":              sku or None,
                "talla":            talla or None,
                "qty_doc":          qty_doc,
                "qty_exp":          0,
                "descripcion":      ln.get("descripcion"),
                "raw_material":     ln.get("raw_material"),
                "suggested_action": "MANUAL",
            })
            continue

        key = (sku, talla)
        db = db_index.get(key)
        if not db:
            ln["match"] = {
                "matched": False,
                "reason": "MISSING_IN_EXPEDIENTE",
                "line_id": None,
            }
            # Sprint 2026-05-06 (AG-03): MISSING_IN_EXPEDIENTE NO se sincroniza
            # al expediente. El producto extra confirmado por la fábrica
            # queda registrado en el SAP pero NO se inserta en
            # expedientes.linea. En su lugar, al confirmar producción se
            # dispara un email al cliente notificándole del extra.
            discrepancies.append({
                "kind":             "MISSING_IN_EXPEDIENTE",
                "severity":         "WARN",
                "sku":              sku,
                "talla":            talla,
                "qty_doc":          qty_doc,
                "qty_exp":          0,
                "descripcion":      ln.get("descripcion"),
                "raw_material":     ln.get("raw_material"),
                "sap_doc":          ln.get("sap_doc"),
                "suggested_action": "NOTIFY_CLIENT",
            })
            continue

        matched_keys.add(key)
        qty_exp = int(db["qty"])
        qty_diff = qty_doc != qty_exp
        nombre_exp = (db.get("nombre_upper") or "").strip()

        # Match de nombre: tolerante. Buscamos que `nombre_doc` y
        # `nombre_exp` compartan tokens significativos. Si difieren
        # 100% emitimos NAME_DIFF (severidad INFO — soft).
        name_match = True
        if nombre_doc and nombre_exp:
            # comparamos sin espacios/guiones/comas
            norm_doc = re.sub(r"[\s,\-_/]+", "", nombre_doc)
            norm_exp = re.sub(r"[\s,\-_/]+", "", nombre_exp)
            name_match = (
                norm_doc == norm_exp
                or norm_doc in norm_exp
                or norm_exp in norm_doc
            )

        ln["match"] = {
            "matched":     True,
            "line_id":     db["line_id"],
            "qty_exp":     qty_exp,
            "qty_diff":    qty_diff,
            "qty_delta":   qty_doc - qty_exp,
            "nombre_exp":  db.get("nombre"),
            "name_match":  name_match,
            "sap_exp":     db.get("sap") or None,
        }

        if qty_diff:
            discrepancies.append({
                "kind":             "QTY_DIFF",
                "severity":         "WARN",
                "sku":              sku,
                "talla":            talla,
                "qty_doc":          qty_doc,
                "qty_exp":          qty_exp,
                "descripcion":      ln.get("descripcion"),
                "nombre_exp":       db.get("nombre"),
                "line_id":          db["line_id"],
                "sap_doc":          ln.get("sap_doc"),
                "suggested_action": "UPDATE_QTY",
            })
        # Sprint 2026-05-06 (AG-03): NAME_DIFF deja de emitir discrepancia.
        # El SKU + talla + qty matchean → la línea está OK desde el punto de
        # vista operativo. La diferencia de nombre es solo cosmética (ej.
        # "75BPR29-CLI-MM-CPAP-EXP" vs "75BPR29-CLI-MM-E-CPAP" — el cliente
        # codifica con prefijo distinto pero referencia el mismo producto).
        # La info queda en `match.name_match=False` por si la UI quiere
        # mostrarlo como nota informativa, pero NO entra a la lista de
        # discrepancies y NO cuenta en `discrepancies_count`.

    lines_in_doc    = len(extracted["lineas"])
    lines_matched   = len(matched_keys)
    lines_unmatched = lines_in_doc - lines_matched
    perfect = (
        lines_in_doc > 0
        and lines_unmatched == 0
        and not any(d["severity"] in ("ERROR", "WARN") for d in discrepancies)
    )

    extracted["discrepancies"] = discrepancies
    extracted["summary"] = {
        "lines_in_doc":         lines_in_doc,
        "lines_matched":        lines_matched,
        "lines_unmatched":      lines_unmatched,
        "lines_in_expediente":  len(db_lines),
        "discrepancies_count":  len(discrepancies),
        "perfect_match":        perfect,
    }
    return extracted


# ─────────────────────────────────────────────────────────────────────
# Punto de entrada principal
# ─────────────────────────────────────────────────────────────────────

def analyze_sap_document(file_bytes: bytes, filename: str,
                         content_type: str, expediente_id) -> dict:
    """Analiza un documento de Confirmación SAP y lo cruza contra
    expedientes.linea del expediente.

    Estrategia:
      · xlsx / xls → parser determinístico Marluvas.
      · pdf        → IA via document_matchmaker (SYSTEM_PROMPT_SAP),
                     y reformateamos al schema canónico de este módulo.
      · csv        → parser determinístico (mismo schema headers).
      · cualquier otro → error.
    """
    fname = (filename or "").lower()
    is_xlsx = fname.endswith((".xlsx", ".xlsm", ".xls")) or \
        content_type in ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         "application/vnd.ms-excel")
    is_pdf  = fname.endswith(".pdf") or content_type == "application/pdf"
    is_csv  = fname.endswith(".csv") or content_type == "text/csv"

    if is_xlsx:
        result = _parse_xlsx_marluvas(file_bytes)
    elif is_csv:
        result = _parse_csv_marluvas(file_bytes)
    elif is_pdf:
        result = _parse_pdf_via_ai(file_bytes, filename, content_type)
    else:
        return {
            "ok": False,
            "error": f"Tipo de archivo no soportado: {content_type or fname}",
            "kind": None,
            "lineas": [],
            "sap_id": None,
            "discrepancies": [],
            "summary": {"lines_in_doc": 0, "lines_matched": 0,
                        "lines_unmatched": 0, "discrepancies_count": 0,
                        "perfect_match": False},
        }

    result["filename"] = filename
    if not result.get("ok"):
        result.setdefault("lineas", [])
        result.setdefault("discrepancies", [])
        result.setdefault("summary", {
            "lines_in_doc": 0, "lines_matched": 0, "lines_unmatched": 0,
            "discrepancies_count": 0, "perfect_match": False,
        })
        return result

    return cross_match_sap(result, expediente_id)


def _parse_csv_marluvas(file_bytes: bytes) -> dict:
    """CSV equivalente al xlsx Marluvas (mismas columnas, separador
    coma o tab)."""
    text = None
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = file_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        return {"ok": False, "error": "No pude decodificar el CSV.",
                "lineas": [], "sap_id": None}

    import csv as _csv
    reader = _csv.reader(io.StringIO(text), delimiter=(
        "\t" if text.count("\t") > text.count(",") else ","
    ))
    rows = list(reader)
    if not rows:
        return {"ok": False, "error": "CSV vacío.", "lineas": [], "sap_id": None}

    cols = _detect_columns(rows[0])
    if "material" not in cols:
        return {
            "ok": False,
            "error": "CSV sin headers Marluvas reconocibles.",
            "lineas": [], "sap_id": None,
        }

    sap_set = set(); lineas = []
    for row in rows[1:]:
        if not row:
            continue
        def _at(i):
            return row[i] if i < len(row) else None
        sap_val = _at(cols.get("sap")) if "sap" in cols else None
        sap_val = (str(sap_val).strip() if sap_val not in (None, "") else None)
        if sap_val and sap_val.endswith(".0"):
            sap_val = sap_val[:-2]
        material_raw = _at(cols.get("material")) if "material" in cols else None
        sku, talla = _split_material(material_raw)
        descripcion_raw = _at(cols.get("descripcion")) if "descripcion" in cols else None
        nombre, talla_desc = _split_descripcion(descripcion_raw)
        if not talla and talla_desc:
            talla = talla_desc
        qty = _safe_int(_at(cols.get("qty"))) if "qty" in cols else 0
        if not material_raw and not descripcion_raw:
            continue
        if not sku and not nombre:
            continue
        if sap_val:
            sap_set.add(sap_val)
        lineas.append({
            "sap_doc":      sap_val,
            "sku":          (sku or "").upper() or None,
            "talla":        (talla or "").upper() or None,
            "qty":          qty,
            "descripcion":  (nombre or "").strip() or None,
            "raw_material": (str(material_raw or "") if material_raw else None),
        })

    sap_list = sorted(sap_set)
    return {
        "ok":           True,
        "kind":         "csv_marluvas",
        "sap_id":       sap_list[0] if sap_list else None,
        "sap_count":    len(sap_list),
        "all_saps":     sap_list,
        "lineas":       lineas,
        "error":        None,
    }


def _parse_pdf_via_ai(file_bytes: bytes, filename: str,
                      content_type: str) -> dict:
    """Fallback PDF: usa el extractor IA del matchmaker
    (SYSTEM_PROMPT_SAP) y reformatea al schema canónico de este módulo.
    """
    try:
        from .document_matchmaker import extract_document
    except ImportError as e:
        return {
            "ok": False,
            "error": f"No pude importar matchmaker: {e}",
            "lineas": [], "sap_id": None,
        }

    ai = extract_document(
        file_bytes=file_bytes,
        filename=filename,
        content_type=content_type or "application/pdf",
        document_type="ART-04_SAP",
    )

    if ai.get("error"):
        return {
            "ok": False,
            "error": ai["error"],
            "kind": "pdf_ai",
            "lineas": [], "sap_id": None,
        }

    sap_set = set(); lineas = []
    for g in (ai.get("groups") or []):
        sap = g.get("sap_number") or None
        if sap:
            sap_set.add(str(sap).strip())
        for ln in (g.get("lines") or []):
            sku = (ln.get("sku") or "").upper().strip() or None
            talla = (ln.get("talla") or "").upper().strip() or None
            qty = _safe_int(ln.get("qty_confirmed") or ln.get("qty"))
            label = (ln.get("product_label") or "").strip() or None
            lineas.append({
                "sap_doc":      sap,
                "sku":          sku,
                "talla":        talla,
                "qty":          qty,
                "descripcion":  label,
                "raw_material": None,
            })

    sap_list = sorted(sap_set)
    return {
        "ok":        True,
        "kind":      "pdf_ai",
        "sap_id":    sap_list[0] if sap_list else None,
        "sap_count": len(sap_list),
        "all_saps":  sap_list,
        "lineas":    lineas,
        "error":     None,
        "ai_meta":   {"model": ai.get("model")},
    }
