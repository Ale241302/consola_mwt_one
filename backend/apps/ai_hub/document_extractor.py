"""
apps.ai_hub.document_extractor
==============================

Sprint 2026-05-11 · Fase 7 — Extracción genérica de campos desde un
documento usando IA.

Sprint 2026-05-11 fix · Migrado de Anthropic Claude a OpenAI gpt-5-nano
para alinearlo con el resto del proyecto (`document_matchmaker`,
`inbound_ocr`, `ocr_customs` ya usan OPENAI_API_KEY en producción).
Evita gestionar dos API keys distintas y consolida en un solo
proveedor que ya está validado en VPS.

Caso de uso (CEO): en `ArtifactFillModal` el operador sube un PDF/Excel/
Word/txt y la IA llena automáticamente los campos del template del
Builder. Este módulo es **agnóstico al dominio** — recibe los bytes y el
`structure_json` y devuelve un dict `{field_id: value, ...}`.

Patrón implementado (copiado de `document_matchmaker.extract_document`):
  · Text-native: pypdf / openpyxl / docx zip-XML / plain → chat.completions
    con `response_format={"type":"json_object"}`. Más rápido (5-10s).
  · Imagen / PDF escaneado: `responses.create()` con `input_file`
    (binario PDF) o `input_image` (imagen). Fallback a `chat.completions`
    con `image_url`.
  · Timeout 90s, max_retries 1 — protege workers gunicorn (timeout 120s).
"""
from __future__ import annotations

import base64
import io
import json
import logging
import os
from typing import Any

from django.conf import settings


log = logging.getLogger(__name__)


# ────────────────────────────────────────────────────────────
# Helpers de configuración
# ────────────────────────────────────────────────────────────
def _cfg(key: str, default: Any = None) -> Any:
    return getattr(settings, "AI_HUB", {}).get(key, default)


# Modelo por defecto — mismo que el resto del proyecto (matchmaker etc.).
EXTRACT_MODEL    = (os.environ.get("OPENAI_OCR_MODEL")
                    or os.environ.get("AI_DOCEXTRACT_MODEL")
                    or "gpt-5-nano")
EXTRACT_TIMEOUT  = float(os.environ.get("AI_DOCEXTRACT_TIMEOUT", "90"))
OPENAI_API_KEY   = (os.environ.get("OPENAI_API_KEY")
                    or _cfg("OPENAI_API_KEY", "")
                    or "")


# ────────────────────────────────────────────────────────────
# Extractores de texto por tipo de archivo
# ────────────────────────────────────────────────────────────
def _extract_text_pdf(file_bytes: bytes) -> str:
    """Intenta extraer texto plano de un PDF text-native con pypdf.
    Si el PDF es solo imagen (escaneado), devuelve ''."""
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(file_bytes))
        chunks = []
        for i, page in enumerate(reader.pages):
            if i >= 30:
                chunks.append(f"\n[... truncado: {len(reader.pages) - 30} páginas más ...]")
                break
            try:
                chunks.append(page.extract_text() or "")
            except Exception:
                continue
        return "\n".join(c for c in chunks if c.strip())
    except Exception as exc:
        log.warning("pypdf failed: %s", exc)
        return ""


def _extract_text_xlsx(file_bytes: bytes) -> str:
    """Extrae celdas de un xlsx/xlsm como TSV (una hoja por sección)."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        out = []
        for sh in wb.worksheets:
            out.append(f"### Hoja: {sh.title}")
            for row in sh.iter_rows(values_only=True):
                line = "\t".join("" if v is None else str(v) for v in row).rstrip()
                if line:
                    out.append(line)
        return "\n".join(out)[:50000]
    except Exception as exc:
        log.warning("openpyxl failed: %s", exc)
        return ""


def _extract_text_docx(file_bytes: bytes) -> str:
    """Word .docx → texto. Parseamos el XML interno del zip."""
    try:
        import zipfile
        import re
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
            xml = zf.read("word/document.xml").decode("utf-8", errors="ignore")
        parts = re.findall(r"<w:t[^>]*>([^<]*)</w:t>", xml)
        return "\n".join(parts)[:50000]
    except Exception as exc:
        log.warning("docx parse failed: %s", exc)
        return ""


def _extract_text_plain(file_bytes: bytes) -> str:
    """Texto plano (txt/csv) — decodifica utf-8 best-effort."""
    try:
        return file_bytes.decode("utf-8", errors="ignore")[:50000]
    except Exception:
        return ""


def _to_text_payload(file_bytes: bytes, mime: str, filename: str) -> tuple[str, str, bool]:
    """Devuelve (text_extracted, kind_label, is_image).

    Sprint 2026-05-11 fix · Umbral bajado de 100 → 20 caracteres.
    Casos como AWB de courier (FedEx, DHL) tienen muy poco texto
    embebido pero suficiente como para que el LLM razone — antes
    caían al path "pdf-image" que es problemático con OpenAI.
    Si pypdf extrae 0 caracteres, mandamos al menos los bytes como
    `input_file` (responses.create) — NUNCA como `image_url` que
    rechaza application/pdf con 400.
    """
    m = (mime or "").lower()
    name = (filename or "").lower()

    if m == "application/pdf" or name.endswith(".pdf"):
        text = _extract_text_pdf(file_bytes)
        if len((text or "").strip()) >= 20:
            return (text, "pdf-text", False)
        return ("", "pdf-image", False)  # PDF escaneado → responses.input_file
    if "spreadsheetml" in m or name.endswith((".xlsx", ".xlsm")):
        return (_extract_text_xlsx(file_bytes), "xlsx", False)
    if "wordprocessingml" in m or name.endswith(".docx"):
        return (_extract_text_docx(file_bytes), "docx", False)
    if m.startswith("text/") or name.endswith((".txt", ".csv", ".tsv")):
        return (_extract_text_plain(file_bytes), "plain", False)
    if m.startswith("image/") or name.endswith((".png", ".jpg", ".jpeg", ".webp")):
        return ("", "image", True)
    return (_extract_text_plain(file_bytes), "unknown", False)


# ────────────────────────────────────────────────────────────
# Construcción del prompt
# ────────────────────────────────────────────────────────────
def _flatten_fields(structure_json: dict) -> list[dict]:
    out = []
    for sec in (structure_json.get("sections") or []):
        for col in (sec.get("columns") or []):
            for f in (col.get("fields") or []):
                out.append({
                    "id":        f.get("id"),
                    "label":     f.get("label") or f.get("id"),
                    "type":      f.get("type") or "text",
                    "options":   f.get("options") or [],
                    "required":  bool(f.get("required")),
                    "helpText":  f.get("helpText") or "",
                })
    return out


def _build_schema_brief(fields: list[dict]) -> str:
    lines = ["FIELDS (id · type · label · valid options if applicable):"]
    for f in fields:
        opts = ""
        if f["type"] in ("select", "radio", "checkbox") and f["options"]:
            opt_labels = []
            for o in f["options"]:
                if isinstance(o, dict):
                    opt_labels.append(o.get("label") or o.get("id") or "")
                else:
                    opt_labels.append(str(o))
            opts = f"  · options=[{', '.join(repr(x) for x in opt_labels if x)}]"
        helptxt = f"  · hint={f['helpText'][:80]!r}" if f["helpText"] else ""
        req = " (required)" if f["required"] else ""
        lines.append(f"  - {f['id']} · {f['type']} · {f['label']!r}{req}{opts}{helptxt}")
    return "\n".join(lines)


SYSTEM_PROMPT = """You are an expert document data extractor for a B2B logistics platform.

Your task: read a document (PDF / Excel / Word / plain text / image) and
fill out the fields of a target form. The form schema includes the field
id, type, label and (for select/radio/checkbox) the list of valid option
labels.

Rules:
1. Return STRICT JSON {"extracted": {field_id: value, ...}, "confidence": {field_id: 0-100, ...}, "notes": "..."}.
2. Only include fields whose value you can support with evidence from the document. Omit unknown fields entirely.
3. For "select" and "radio" fields, the value MUST be one of the option labels exactly (case-insensitive match — return the canonical label as listed in options).
4. For "checkbox" fields, return true or false.
5. For "number" / "date" fields, return canonical form (numbers as JSON numbers; dates as ISO YYYY-MM-DD).
6. For "text" / "textarea" / "code" fields, return a plain string.
7. NEVER invent data. If unsure, omit the field.
8. confidence: integer 0-100 per field (your subjective confidence).
9. notes: optional short string (≤200 chars) explaining caveats.

Output STRICT JSON. No markdown, no commentary outside the JSON object."""


def _parse_strict_json(raw: str) -> dict:
    s = (raw or "").strip()
    if s.startswith("```"):
        s = s.lstrip("`")
        if s.lower().startswith("json"):
            s = s[4:].lstrip()
        if s.endswith("```"):
            s = s[:-3]
        s = s.strip()
    if not s.startswith("{"):
        i = s.find("{")
        if i >= 0:
            s = s[i:]
    if not s.endswith("}"):
        j = s.rfind("}")
        if j >= 0:
            s = s[:j + 1]
    return json.loads(s)


# ────────────────────────────────────────────────────────────
# Llamada al LLM (OpenAI · mismo patrón que document_matchmaker)
# ────────────────────────────────────────────────────────────
def _call_openai_text(*, system: str, user_text: str, model: str) -> str:
    """Path text-native: chat.completions con response_format JSON."""
    from openai import OpenAI
    client = OpenAI(api_key=OPENAI_API_KEY, timeout=EXTRACT_TIMEOUT, max_retries=1)
    chat = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": system},
            {"role": "user",   "content": user_text},
        ],
        response_format={"type": "json_object"},
    )
    return chat.choices[0].message.content or ""


def _call_openai_vision(*, system: str, user_text: str, model: str,
                       file_bytes: bytes, content_type: str, filename: str,
                       is_image: bool) -> str:
    """Path vision / file. Reglas estrictas:

    Sprint 2026-05-11 fix · NO convertimos PDF a image_url. La API de
    `chat.completions` con `image_url` rechaza application/pdf con
    "Invalid MIME type. Only image types are supported." (400). El CEO
    fue explícito: "no conviertas un pdf a imagen".

    Comportamiento:
      - PDF (escaneado o text-native sin texto extraído) → siempre
        `responses.create` con `input_file`. Si falla, propagamos el
        error tal cual.
      - Imagen real (png/jpg/etc) → primero `responses.create` con
        `input_image`. Si falla, fallback a `chat.completions` con
        `image_url` (sí acepta imágenes).
    """
    from openai import OpenAI
    client = OpenAI(api_key=OPENAI_API_KEY, timeout=EXTRACT_TIMEOUT, max_retries=1)
    b64 = base64.b64encode(file_bytes).decode("ascii")
    data_url = f"data:{content_type or 'application/octet-stream'};base64,{b64}"

    # ── Path imagen real ────────────────────────────────
    if is_image:
        try:
            resp = client.responses.create(
                model        = model,
                instructions = system,
                input=[{
                    "role": "user",
                    "content": [
                        {"type": "input_text",  "text": user_text},
                        {"type": "input_image", "image_url": data_url},
                    ],
                }],
                response_format={"type": "json_object"},
            )
            return resp.output_text or ""
        except Exception as exc:
            log.warning("responses.create (image) failed (%s); fallback chat.completions", exc)
        # Fallback solo válido para imágenes (no PDFs).
        content_parts = [
            {"type": "text",      "text": user_text},
            {"type": "image_url", "image_url": {"url": data_url}},
        ]
        chat = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": system},
                {"role": "user",   "content": content_parts},
            ],
            response_format={"type": "json_object"},
        )
        return chat.choices[0].message.content or ""

    # ── Path PDF/desconocido: SIEMPRE responses.create input_file ──
    # NO existe fallback porque chat.completions.image_url rechazaría
    # el PDF. Si la API responses falla, dejamos que la excepción se
    # propague para que extract_fields_from_document() la reporte
    # como _meta.error legible.
    pdf_filename = filename or "document.pdf"
    if not pdf_filename.lower().endswith(".pdf") and (content_type or "").lower() == "application/pdf":
        pdf_filename = pdf_filename + ".pdf"
    resp = client.responses.create(
        model        = model,
        instructions = system,
        input=[{
            "role": "user",
            "content": [
                {"type": "input_text", "text": user_text},
                {
                    "type":      "input_file",
                    "filename":  pdf_filename,
                    "file_data": data_url,
                },
            ],
        }],
        response_format={"type": "json_object"},
    )
    return resp.output_text or ""


# ────────────────────────────────────────────────────────────
# API pública
# ────────────────────────────────────────────────────────────
def extract_fields_from_document(*,
                                 file_bytes: bytes,
                                 mime_type: str,
                                 filename: str,
                                 structure_json: dict,
                                 model: str | None = None,
                                 max_tokens: int | None = None) -> dict:
    """Punto de entrada único. Ver docstring del módulo."""
    fields = _flatten_fields(structure_json or {})
    if not fields:
        return {
            "extracted": {}, "confidence": {}, "notes": "structure_json vacío",
            "_meta": {"error": "no fields in schema"},
        }

    if not OPENAI_API_KEY:
        return {
            "extracted": {}, "confidence": {}, "notes": "",
            "_meta": {"error": "OPENAI_API_KEY no configurado en el VPS"},
        }
    try:
        from openai import OpenAI  # noqa: F401
    except ImportError as exc:
        return {
            "extracted": {}, "confidence": {}, "notes": "",
            "_meta": {"error": f"openai SDK no disponible: {exc}"},
        }

    text_payload, kind, is_image = _to_text_payload(file_bytes, mime_type, filename)
    schema_brief = _build_schema_brief(fields)
    used_model   = model or EXTRACT_MODEL

    raw = ""
    try:
        if text_payload:
            user_text = (
                f"DOCUMENT (filename={filename!r}, kind={kind}):\n\n"
                f"{text_payload}\n\n---\n{schema_brief}\n\n"
                "Return strict JSON as instructed."
            )
            raw = _call_openai_text(system=SYSTEM_PROMPT, user_text=user_text,
                                    model=used_model)
        else:
            # PDF escaneado, imagen, o tipo desconocido sin texto.
            user_text = (
                f"Document attached (filename={filename!r}). Read it and fill the schema.\n\n"
                f"{schema_brief}\n\nReturn strict JSON as instructed."
            )
            raw = _call_openai_vision(
                system     = SYSTEM_PROMPT,
                user_text  = user_text,
                model      = used_model,
                file_bytes = file_bytes,
                content_type = mime_type or ("image/png" if is_image else "application/pdf"),
                filename   = filename,
                is_image   = is_image,
            )
    except Exception as exc:
        log.exception("extract_fields_from_document · OpenAI call failed")
        return {
            "extracted": {}, "confidence": {}, "notes": "",
            "_meta": {"error": f"OpenAI error: {type(exc).__name__}: {exc}",
                       "kind": kind, "model": used_model},
        }

    try:
        data = _parse_strict_json(raw)
    except Exception as exc:
        log.warning("extract_fields_from_document · JSON parse failed: %s | raw=%r",
                    exc, raw[:200])
        return {
            "extracted": {}, "confidence": {}, "notes": "",
            "_meta": {"error": f"JSON parse failed: {exc}",
                       "raw_preview": (raw or "")[:200],
                       "kind": kind, "model": used_model},
        }

    extracted  = data.get("extracted")  or {}
    confidence = data.get("confidence") or {}
    notes      = data.get("notes")      or ""

    # Normalización defensiva: sólo dejamos field_ids válidos del schema.
    valid_ids = {f["id"] for f in fields}
    extracted  = {k: v for k, v in extracted.items()  if k in valid_ids}
    confidence = {k: v for k, v in confidence.items() if k in valid_ids}

    return {
        "extracted":  extracted,
        "confidence": confidence,
        "notes":      str(notes)[:500],
        "_meta": {
            "model":             used_model,
            "kind":              kind,
            "fields_in_schema":  len(fields),
            "fields_extracted":  len(extracted),
        },
    }
