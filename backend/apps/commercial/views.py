"""
=====================================================================
MWT.ONE · apps.commercial.views
Agente responsable: [AG-BACKEND]

ViewSets + endpoint crítico `resolve_client_price`.

Recursos REST:
  · /api/commercial/pricelist-versions/     (CRUD)
  · /api/commercial/grade-items/            (CRUD · cost_usd enmascarado si ≠ CEO)
  · /api/commercial/client-assignments/     (CRUD · CPA)
  · /api/commercial/early-payment-policies/ (CRUD)
  · /api/commercial/early-payment-tiers/    (CRUD)
  · /api/commercial/commission-rules/       (CEO-ONLY · CRUD)
  · /api/commercial/catalogs/currencies/    (read-only)
  · /api/commercial/catalogs/sources/       (read-only)
  · /api/commercial/catalogs/commission-bases/ (read-only)

Endpoints de negocio:
  · POST /api/commercial/resolve_client_price/  → waterfall (CPA → MIN pricelist → EPP tier)
  · POST /api/commercial/pricelist-versions/<id>/bulk-upsert-items/
        → carga masiva de grade_items desde un Excel pre-parseado (JSON)

Reglas MWT:
  · CEO-ONLY: cost_usd / margenes / commission_rule completa NO llegan a CLIENT.
  · Idempotencia: CRUD usa _ensure_uuid() para auto-asignar UUID.
  · Soft-delete: DELETE hace is_active=False.
=====================================================================
"""
import uuid
import logging
from decimal import Decimal

import requests
from django.core.cache import cache
from django.db import transaction, connection
from django.db.models import Q, Case, When, Value, IntegerField
from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView

from .models import (
    PriceListVersion, GradeItem, ClientAssignment,
    EarlyPaymentPolicy, EarlyPaymentTier, CommissionRule,
    CurrencyCat, PriceListSourceCat, CommissionBaseCat,
)
from .serializers import (
    PriceListVersionSerializer, PriceListVersionListSerializer,
    GradeItemSerializer, GradeItemPublicSerializer,
    ClientAssignmentSerializer,
    EarlyPaymentPolicySerializer, EarlyPaymentTierSerializer,
    CommissionRuleSerializer,
    CurrencyCatSerializer, PriceListSourceCatSerializer,
    CommissionBaseCatSerializer,
    ResolveClientPriceInputSerializer, ResolveClientPriceOutputSerializer,
)

log = logging.getLogger(__name__)


# =====================================================================
# Helpers
# =====================================================================
CEO_ROLES = {"admin", "superadmin", "ceo"}


def _is_ceo(request) -> bool:
    """True si el JWT trae un role CEO-like (admin/superadmin/ceo)."""
    if not request or not getattr(request, "auth", None):
        return False
    role = (request.auth.get("role") or "").lower()
    return role in CEO_ROLES


# ─────────────────────────────────────────────────────────────
# Ola 1 · 1.4b — validación de ?client_id= contra el scope (P0-4)
# ─────────────────────────────────────────────────────────────
def _scope_client_ids(request):
    """Lista de client_ids permitidos al usuario, o None si bypass (ve todos).

    Reutiliza la lógica canónica de scoped_querysets (incluido el guard
    anti-bypass para tokens MCP con tenant quemado).
    """
    from apps.core.scoped_querysets import is_bypass, _scope_ids

    user = getattr(request, "user", None)
    if user is None or not getattr(user, "is_authenticated", False):
        return None
    if is_bypass(user):
        return None
    return _scope_ids(user)


def _validated_client_ids(request, requested: str | None) -> list | None:
    """Valida un `client_id` del query string contra el scope del usuario.

    Returns:
      - lista  → client_ids permitidos (si `requested`, la intersección o []).
      - None   → bypass (no restringir por cliente).
    """
    allowed = _scope_client_ids(request)
    if allowed is None:
        return None
    if not requested:
        return allowed
    rq = str(requested).lower()
    return [rq] if rq in allowed else []


def _scope_filter_qs(qs, request, field: str, requested: str | None):
    """Aplica el scope por cliente a un queryset (fail-closed a vacío)."""
    ids = _validated_client_ids(request, requested)
    if ids is None:
        return qs
    if not ids:
        return qs.none()
    return qs.filter(**{f"{field}__in": ids})


def _ensure_uuid(data: dict) -> dict:
    if not data.get("id"):
        data["id"] = str(uuid.uuid4())
    return data


def _request_data_copy(request):
    if hasattr(request.data, "copy"):
        return request.data.copy()
    return dict(request.data)


def _sum_size_multipliers(size_multipliers: dict) -> int:
    if not isinstance(size_multipliers, dict):
        return 0
    total = 0
    for v in size_multipliers.values():
        try:
            total += int(v)
        except (TypeError, ValueError):
            continue
    return total


# =====================================================================
# PriceListVersion ViewSet
# =====================================================================
class PriceListVersionViewSet(viewsets.ModelViewSet):
    required_module = "commercial"
    queryset = PriceListVersion.objects.filter(is_active=True)
    serializer_class = PriceListVersionSerializer

    def get_serializer_class(self):
        if self.action == "list":
            return PriceListVersionListSerializer
        return PriceListVersionSerializer

    def get_queryset(self):
        qs = PriceListVersion.objects.filter(is_active=True)
        brand_id = self.request.query_params.get("brand_id")
        source   = self.request.query_params.get("source")
        q        = self.request.query_params.get("q")
        if brand_id:
            qs = qs.filter(brand_id=brand_id)
        if source:
            qs = qs.filter(source=source)
        if q:
            qs = qs.filter(nombre__icontains=q) | qs.filter(codigo__icontains=q)
        return qs.order_by("-valid_from", "codigo")

    def list(self, request, *args, **kwargs):
        # Fable5 · batch: count de grade_items en UN query agrupado
        # (antes 1 query POR versión vía SerializerMethodField).
        from django.db.models import Count
        qs = self.filter_queryset(self.get_queryset())
        rows = list(qs)
        ids = [r.id for r in rows]
        counts = {str(i): 0 for i in ids}
        if ids:
            aggs = (GradeItem.objects
                    .filter(pricelist_version_id__in=ids, is_active=True)
                    .values("pricelist_version_id")
                    .annotate(n=Count("id")))
            for a in aggs:
                counts[str(a["pricelist_version_id"])] = int(a["n"] or 0)
        serializer = PriceListVersionListSerializer(
            rows, many=True,
            context={**self.get_serializer_context(), "batch_items_count": counts},
        )
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        data = _ensure_uuid(_request_data_copy(request))
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save(update_fields=["is_active", "updated_at"])

    # ── POST /api/commercial/pricelist-versions/<id>/bulk-upsert-items/ ─
    @action(detail=True, methods=["post"], url_path="bulk-upsert-items")
    def bulk_upsert_items(self, request, pk=None):
        """Carga masiva de grade_items desde un Excel ya parseado a JSON.

        Payload:
          {
            "items": [
              {
                "product_sku": "NIK-AIR-001",
                "product_name": "Air Max 001",
                "unit_price_usd": 42.50,
                "cost_usd": 18.20,                  # CEO-ONLY, opcional
                "size_multipliers": {"37": 2, "38": 4, "39": 6},
                "tags": ["ss26"],
                "metadata": {}
              },
              ...
            ],
            "replace_existing": true   # default false (upsert)
          }
        """
        try:
            plv = PriceListVersion.objects.get(pk=pk, is_active=True)
        except PriceListVersion.DoesNotExist:
            return Response({"detail": "PriceListVersion no encontrada."},
                            status=status.HTTP_404_NOT_FOUND)

        data = _request_data_copy(request)
        items = data.get("items") or []
        if not isinstance(items, list) or not items:
            return Response({"detail": "items[] vacío."},
                            status=status.HTTP_400_BAD_REQUEST)
        replace_existing = bool(data.get("replace_existing", False))

        created = 0
        updated = 0
        errors  = []

        with transaction.atomic():
            if replace_existing:
                GradeItem.objects.filter(
                    pricelist_version_id=plv.id, is_active=True
                ).update(is_active=False)

            for raw in items:
                sku = (raw.get("product_sku") or "").strip()
                if not sku:
                    errors.append({"row": raw, "error": "product_sku vacío"})
                    continue
                try:
                    unit_price = Decimal(str(raw.get("unit_price_usd", 0)))
                except Exception:
                    errors.append({"row": raw, "error": "unit_price_usd inválido"})
                    continue
                cost = raw.get("cost_usd")
                try:
                    cost_dec = Decimal(str(cost)) if cost is not None and cost != "" else None
                except Exception:
                    cost_dec = None

                size_multipliers = raw.get("size_multipliers") or {}
                if not isinstance(size_multipliers, dict):
                    size_multipliers = {}
                grade_total = _sum_size_multipliers(size_multipliers)

                existing = GradeItem.objects.filter(
                    pricelist_version_id=plv.id,
                    product_sku=sku,
                    is_active=True,
                ).first()

                if existing:
                    existing.product_name     = raw.get("product_name") or existing.product_name
                    existing.unit_price_usd   = unit_price
                    if cost_dec is not None:
                        existing.cost_usd     = cost_dec
                    existing.size_multipliers = size_multipliers
                    existing.grade_moq_total  = grade_total
                    existing.tags             = raw.get("tags") or existing.tags
                    existing.metadata         = raw.get("metadata") or existing.metadata
                    existing.save(update_fields=[
                        "product_name", "unit_price_usd", "cost_usd",
                        "size_multipliers", "grade_moq_total",
                        "tags", "metadata", "updated_at",
                    ])
                    updated += 1
                else:
                    GradeItem.objects.create(
                        id                   = uuid.uuid4(),
                        pricelist_version_id = plv.id,
                        brand_id             = plv.brand_id,
                        product_sku          = sku,
                        product_name         = raw.get("product_name") or "",
                        unit_price_usd       = unit_price,
                        cost_usd             = cost_dec,
                        size_multipliers     = size_multipliers,
                        grade_moq_total      = grade_total,
                        tags                 = raw.get("tags") or [],
                        metadata             = raw.get("metadata") or {},
                        is_active            = True,
                    )
                    created += 1

        return Response({
            "ok":      True,
            "created": created,
            "updated": updated,
            "errors":  errors,
        }, status=status.HTTP_200_OK)

    # ── GET /api/commercial/pricelist-versions/<id>/items/ ─────────────
    @action(detail=True, methods=["get"])
    def items(self, request, pk=None):
        """Lista los grade_items de esta pricelist (enmascarando cost si ≠ CEO)."""
        qs = GradeItem.objects.filter(
            pricelist_version_id=pk, is_active=True
        ).order_by("product_sku")
        ser_cls = GradeItemSerializer if _is_ceo(request) else GradeItemPublicSerializer
        return Response(ser_cls(qs, many=True).data)


# =====================================================================
# GradeItem ViewSet (cost_usd enmascarado si ≠ CEO)
# =====================================================================
class GradeItemViewSet(viewsets.ModelViewSet):
    required_module = "commercial"
    queryset = GradeItem.objects.filter(is_active=True)

    def get_serializer_class(self):
        return GradeItemSerializer if _is_ceo(self.request) else GradeItemPublicSerializer

    def get_queryset(self):
        qs = GradeItem.objects.filter(is_active=True)
        plv_id   = self.request.query_params.get("pricelist_version_id")
        brand_id = self.request.query_params.get("brand_id")
        sku      = self.request.query_params.get("product_sku")
        q        = self.request.query_params.get("q")
        if plv_id:
            qs = qs.filter(pricelist_version_id=plv_id)
        if brand_id:
            qs = qs.filter(brand_id=brand_id)
        if sku:
            qs = qs.filter(product_sku=sku)
        if q:
            qs = qs.filter(product_sku__icontains=q) | qs.filter(product_name__icontains=q)
        return qs.order_by("product_sku")

    def create(self, request, *args, **kwargs):
        data = _ensure_uuid(_request_data_copy(request))
        # si no-CEO intenta enviar cost_usd → se descarta silenciosamente
        if not _is_ceo(request):
            data.pop("cost_usd", None)
        # Recalcula grade_moq_total por seguridad (no confiar en cliente)
        sm = data.get("size_multipliers") or {}
        data["grade_moq_total"] = _sum_size_multipliers(sm)
        serializer = GradeItemSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        # Respuesta con el serializer público si ≠ CEO
        out_cls = GradeItemSerializer if _is_ceo(request) else GradeItemPublicSerializer
        return Response(out_cls(serializer.instance).data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        data = _request_data_copy(request)
        if not _is_ceo(request):
            data.pop("cost_usd", None)
        sm = data.get("size_multipliers")
        if sm is not None:
            data["grade_moq_total"] = _sum_size_multipliers(sm)
        serializer = GradeItemSerializer(instance, data=data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        out_cls = GradeItemSerializer if _is_ceo(request) else GradeItemPublicSerializer
        return Response(out_cls(serializer.instance).data)

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save(update_fields=["is_active", "updated_at"])


# =====================================================================
# ClientAssignment ViewSet (CPA)
# =====================================================================
class ClientAssignmentViewSet(viewsets.ModelViewSet):
    required_module = "commercial"
    queryset = ClientAssignment.objects.filter(is_active=True)
    serializer_class = ClientAssignmentSerializer

    def get_queryset(self):
        qs = ClientAssignment.objects.filter(is_active=True)
        client_id = self.request.query_params.get("client_id")
        brand_id  = self.request.query_params.get("brand_id")
        sku       = self.request.query_params.get("brand_sku")
        # Ola 1 · 1.4b — el ?client_id= se valida contra el scope del usuario.
        qs = _scope_filter_qs(qs, self.request, "client_id", client_id)
        if brand_id:
            qs = qs.filter(brand_id=brand_id)
        if sku:
            qs = qs.filter(brand_sku=sku)
        return qs.order_by("brand_sku")

    def create(self, request, *args, **kwargs):
        data = _ensure_uuid(_request_data_copy(request))
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save(update_fields=["is_active", "updated_at"])


# =====================================================================
# EarlyPaymentPolicy ViewSet
# =====================================================================
class EarlyPaymentPolicyViewSet(viewsets.ModelViewSet):
    required_module = "commercial"
    queryset = EarlyPaymentPolicy.objects.filter(is_active=True)
    serializer_class = EarlyPaymentPolicySerializer

    def get_queryset(self):
        qs = EarlyPaymentPolicy.objects.filter(is_active=True)
        client_id = self.request.query_params.get("client_id")
        brand_id  = self.request.query_params.get("brand_id")
        # Ola 1 · 1.4b — el ?client_id= se valida contra el scope del usuario.
        qs = _scope_filter_qs(qs, self.request, "client_id", client_id)
        if brand_id:
            qs = qs.filter(brand_id=brand_id)
        return qs.order_by("-valid_from", "codigo")

    def create(self, request, *args, **kwargs):
        data = _ensure_uuid(_request_data_copy(request))
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save(update_fields=["is_active", "updated_at"])

    # ── POST /api/commercial/early-payment-policies/<id>/replace-tiers/ ─
    @action(detail=True, methods=["post"], url_path="replace-tiers")
    def replace_tiers(self, request, pk=None):
        """Reemplaza atómicamente los tiers de una policy.

        Payload:
          { "tiers": [
              { "payment_days": 0,  "discount_pct": 5.0, "tier_label": "Contado" },
              { "payment_days": 30, "discount_pct": 2.5, "tier_label": "30 días" },
              { "payment_days": 60, "discount_pct": 0.0, "tier_label": "60 días" }
          ] }
        """
        try:
            policy = EarlyPaymentPolicy.objects.get(pk=pk, is_active=True)
        except EarlyPaymentPolicy.DoesNotExist:
            return Response({"detail": "Policy no encontrada."},
                            status=status.HTTP_404_NOT_FOUND)

        data = _request_data_copy(request)
        tiers = data.get("tiers") or []
        if not isinstance(tiers, list):
            return Response({"detail": "tiers[] inválido."},
                            status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            EarlyPaymentTier.objects.filter(
                policy_id=policy.id, is_active=True
            ).update(is_active=False)

            created = []
            for i, raw in enumerate(tiers):
                try:
                    days = int(raw.get("payment_days", 0))
                    pct  = Decimal(str(raw.get("discount_pct", 0)))
                except Exception:
                    continue
                tier = EarlyPaymentTier.objects.create(
                    id           = uuid.uuid4(),
                    policy_id    = policy.id,
                    payment_days = days,
                    discount_pct = pct,
                    tier_label   = raw.get("tier_label") or f"{days} días",
                    orden        = i,
                    metadata     = raw.get("metadata") or {},
                    is_active    = True,
                )
                created.append(tier)

        return Response({
            "ok": True,
            "policy_id": str(policy.id),
            "tiers": EarlyPaymentTierSerializer(created, many=True).data,
        })


# =====================================================================
# EarlyPaymentTier ViewSet (CRUD bajo nivel)
# =====================================================================
class EarlyPaymentTierViewSet(viewsets.ModelViewSet):
    required_module = "commercial"
    queryset = EarlyPaymentTier.objects.filter(is_active=True)
    serializer_class = EarlyPaymentTierSerializer

    def get_queryset(self):
        qs = EarlyPaymentTier.objects.filter(is_active=True)
        policy_id = self.request.query_params.get("policy_id")
        # Ola 1 · 1.4b — los tiers se scopean vía la policy a la que
        # pertenecen (policy.client_id ∈ scope del usuario).
        allowed_clients = _scope_client_ids(self.request)
        if allowed_clients is not None:
            qs = qs.filter(policy_id__in=(
                EarlyPaymentPolicy.objects.filter(is_active=True)
                .filter(client_id__in=allowed_clients)
                .values_list("id", flat=True)
            ))
        if policy_id:
            qs = qs.filter(policy_id=policy_id)
        return qs.order_by("payment_days")

    def create(self, request, *args, **kwargs):
        data = _ensure_uuid(_request_data_copy(request))
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save(update_fields=["is_active", "updated_at"])


# =====================================================================
# CommissionRule ViewSet            [CEO-ONLY]
# =====================================================================
class CommissionRuleViewSet(viewsets.ModelViewSet):
    """CRUD CEO-ONLY. Si not _is_ceo → 403 sobre TODOS los métodos."""
    required_module = "commercial"
    queryset = CommissionRule.objects.filter(is_active=True)
    serializer_class = CommissionRuleSerializer

    def initial(self, request, *args, **kwargs):
        super().initial(request, *args, **kwargs)
        if not _is_ceo(request):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("CommissionRule: CEO-ONLY.")

    def get_queryset(self):
        qs = CommissionRule.objects.filter(is_active=True)
        brand_id  = self.request.query_params.get("brand_id")
        client_id = self.request.query_params.get("client_id")
        base      = self.request.query_params.get("commission_base")
        if brand_id:
            qs = qs.filter(brand_id=brand_id)
        if client_id:
            qs = qs.filter(client_id=client_id)
        if base:
            qs = qs.filter(commission_base=base)
        return qs.order_by("-valid_from", "codigo")

    def create(self, request, *args, **kwargs):
        data = _ensure_uuid(_request_data_copy(request))
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save(update_fields=["is_active", "updated_at"])


# =====================================================================
# Catálogos read-only
# =====================================================================
class CurrencyCatViewSet(viewsets.ReadOnlyModelViewSet):
    required_module = "commercial"
    queryset = CurrencyCat.objects.filter(is_active=True)
    serializer_class = CurrencyCatSerializer


class PriceListSourceCatViewSet(viewsets.ReadOnlyModelViewSet):
    required_module = "commercial"
    queryset = PriceListSourceCat.objects.filter(is_active=True).order_by("orden")
    serializer_class = PriceListSourceCatSerializer


class CommissionBaseCatViewSet(viewsets.ReadOnlyModelViewSet):
    required_module = "commercial"
    queryset = CommissionBaseCat.objects.filter(is_active=True).order_by("orden")
    serializer_class = CommissionBaseCatSerializer


# =====================================================================
# Endpoint crítico: resolve_client_price (WATERFALL)
# =====================================================================
from rest_framework.views import APIView


# ---------------------------------------------------------------------
# Helper reutilizable: compute_client_price()
#
# Misma lógica que ResolveClientPriceView.post pero CALLABLE desde
# otros módulos (ej. expedientes.views.ExpedienteViewSet.create) sin
# depender de un Request HTTP. Devuelve el `final_price` como Decimal
# o None si no se puede resolver.
#
# Sin enriquecimientos CEO (cost/margen/commission) — solo el precio.
# ---------------------------------------------------------------------
def compute_client_price(client_id, brand_id, product_sku, days_req=0):
    """Aplica el waterfall (CPA → PriceListVersion → EPP) y devuelve el
    precio final como Decimal. Ignora errores: cualquier fallo devuelve
    None y deja que el caller use su fallback."""
    from django.utils import timezone
    if not client_id or not brand_id or not product_sku:
        return None
    try:
        days_req = int(days_req or 0)

        # 1) CPA override
        cpa = ClientAssignment.objects.filter(
            client_id=client_id, brand_id=brand_id,
            brand_sku=product_sku, is_active=True,
        ).order_by("-valid_from").first()

        base_price = None
        if cpa:
            base_price = Decimal(str(cpa.cached_client_price))
        else:
            # 2) PriceListVersion + GradeItem
            today = timezone.now().date()
            winning_plv = (
                PriceListVersion.objects.filter(
                    brand_id=brand_id, is_active=True,
                    valid_from__lte=today,
                ).filter(
                    Q(valid_to__isnull=True) | Q(valid_to__gte=today)
                ).annotate(
                    _has_end=Case(
                        When(valid_to__isnull=False, then=Value(0)),
                        default=Value(1),
                        output_field=IntegerField(),
                    )
                ).order_by("_has_end", "-created_at").first()
            )
            if winning_plv:
                grade_item = GradeItem.objects.filter(
                    pricelist_version_id=winning_plv.id,
                    product_sku=product_sku,
                    is_active=True,
                ).first()
                if grade_item:
                    base_price = Decimal(str(grade_item.unit_price_usd))

        if base_price is None:
            return None

        # 3) Early Payment tier (descuento)
        discount_pct = Decimal("0")
        policy = EarlyPaymentPolicy.objects.filter(
            client_id=client_id, brand_id=brand_id, is_active=True,
        ).order_by("-valid_from").first()
        if policy:
            tier = EarlyPaymentTier.objects.filter(
                policy_id=policy.id, is_active=True,
                payment_days__gte=days_req,
            ).order_by("payment_days").first()
            if tier is None:
                tier = EarlyPaymentTier.objects.filter(
                    policy_id=policy.id, is_active=True,
                ).order_by("-payment_days").first()
            if tier:
                discount_pct = Decimal(str(tier.discount_pct))

        final_price = (base_price * (Decimal("100") - discount_pct) / Decimal("100"))
        return final_price.quantize(Decimal("0.0001"))
    except Exception:
        return None


class ResolveClientPriceView(APIView):
    """POST /api/commercial/resolve_client_price/

    Waterfall:
      1. CPA override: pricing.client_assignment WHERE (client_id, brand_id, brand_sku)
         → si existe: base_price = cached_client_price · source = 'CPA'
      2. MIN(unit_price_usd) sobre pricing.grade_item WHERE is_active y
         pricelist_version_id ∈ versiones ACTIVAS de brand_id
         → base_price = min · source = 'PRICELIST'
      3. Aplica el tier MÁS CERCANO por arriba de early_payment_policy
         para (client_id, brand_id) donde payment_days >= requested_payment_days
         → discount_pct · source = 'CPA+EPP' o 'PRICELIST+EPP'

    Retorna:
      { ok, client_id, brand_id, product_sku, base_price, discount_applied,
        final_price, grade_moq, size_multipliers, source, notes[] }
      Si is_ceo también: cost_usd, margen_usd, margen_pct,
                         commission_pct, commission_base.

    Este endpoint NUNCA expone cost/margen a usuarios no-CEO — aún cuando el
    GradeItem los tiene, el resultado los OMITE antes de serializar.
    """
    required_module = "commercial"

    def post(self, request):
        ser = ResolveClientPriceInputSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        p = ser.validated_data

        client_id  = p["client_id"]
        brand_id   = p["brand_id"]
        sku        = p["product_sku"]
        days_req   = p.get("requested_payment_days", 0) or 0
        currency   = p.get("currency", "USD")

        notes = []
        base_price = None
        grade_moq = None
        size_mult = {}
        source = "NONE"
        grade_item = None

        # ── 1. CPA override ─────────────────────────────────────
        cpa = ClientAssignment.objects.filter(
            client_id=client_id,
            brand_id=brand_id,
            brand_sku=sku,
            is_active=True,
        ).order_by("-valid_from").first()

        # ── Selección de la PLV ganadora ──
        # Reglas (en este orden):
        #   1. Solo PLVs ACTIVE Y vigentes hoy (valid_from <= today AND
        #      (valid_to IS NULL OR valid_to >= today)).
        #   2. Closed-window (valid_to IS NOT NULL) gana sobre open-window
        #      (valid_to IS NULL). Razón: si el operador especificó fecha_fin,
        #      es una ventana intencional con prioridad sobre listas "por
        #      defecto"/sin caducidad.
        #   3. Empate → created_at DESC (la más reciente cargada gana).
        #
        # Antes la lógica era MIN(unit_price_usd), lo cual elegía la lista
        # más barata sin importar qué tan vieja fuera — incorrecto cuando
        # hay varias listas históricas activas.
        today = timezone.now().date()
        winning_plv = (
            PriceListVersion.objects.filter(
                brand_id=brand_id, is_active=True,
                valid_from__lte=today,
            ).filter(
                Q(valid_to__isnull=True) | Q(valid_to__gte=today)
            ).annotate(
                _has_end=Case(
                    When(valid_to__isnull=False, then=Value(0)),
                    default=Value(1),
                    output_field=IntegerField(),
                )
            ).order_by("_has_end", "-created_at").first()
        )
        grade_item = None
        if winning_plv:
            grade_item = GradeItem.objects.filter(
                pricelist_version_id=winning_plv.id,
                product_sku=sku,
                is_active=True,
            ).first()

        if cpa:
            base_price = Decimal(str(cpa.cached_client_price))
            source = "CPA"
            notes.append(f"CPA override activo desde {cpa.valid_from}.")
        elif grade_item:
            base_price = Decimal(str(grade_item.unit_price_usd))
            source = "PRICELIST"
            notes.append(
                f"Precio resuelto por priority-order sobre pricelist_versions vigentes "
                f"(closed-window > open-window > created_at DESC). "
                f"PLV ganadora: {winning_plv.codigo} "
                f"({winning_plv.valid_from}..{winning_plv.valid_to or '∞'})."
            )

        if grade_item:
            grade_moq = grade_item.grade_moq_total
            size_mult = grade_item.size_multipliers or {}

        if base_price is None:
            # No hay precio resoluble
            out = {
                "ok": False,
                "client_id": str(client_id),
                "brand_id":  str(brand_id),
                "product_sku": sku,
                "currency": currency,
                "base_price": None,
                "discount_applied": Decimal("0"),
                "final_price": None,
                "grade_moq": grade_moq,
                "size_multipliers": size_mult,
                "source": source,
                "notes": notes + ["Sin CPA ni pricelist activa para (brand_id, product_sku)."],
            }
            return Response(out, status=status.HTTP_404_NOT_FOUND)

        # ── 2. Early Payment Tier ───────────────────────────────
        discount_pct = Decimal("0")
        policy = EarlyPaymentPolicy.objects.filter(
            client_id=client_id, brand_id=brand_id, is_active=True,
        ).order_by("-valid_from").first()

        if policy:
            tier = EarlyPaymentTier.objects.filter(
                policy_id=policy.id,
                is_active=True,
                payment_days__gte=days_req,
            ).order_by("payment_days").first()
            if tier is None:
                # fallback: el tier con MAYOR payment_days disponible
                tier = EarlyPaymentTier.objects.filter(
                    policy_id=policy.id, is_active=True,
                ).order_by("-payment_days").first()
                if tier:
                    notes.append(
                        f"No hay tier ≥ {days_req} días; se usa tier mayor disponible ({tier.payment_days}d)."
                    )
            if tier:
                discount_pct = Decimal(str(tier.discount_pct))
                source = f"{source}+EPP"
                notes.append(f"Tier aplicado: {tier.payment_days}d → {tier.discount_pct}%.")
        else:
            notes.append("Sin EarlyPaymentPolicy para (client_id, brand_id).")

        # ── 3. Final price ──────────────────────────────────────
        final_price = (base_price * (Decimal("100") - discount_pct) / Decimal("100")).quantize(
            Decimal("0.0001")
        )

        out = {
            "ok": True,
            "client_id": str(client_id),
            "brand_id":  str(brand_id),
            "product_sku": sku,
            "currency": currency,
            "base_price": base_price,
            "discount_applied": discount_pct,
            "final_price": final_price,
            "grade_moq": grade_moq,
            "size_multipliers": size_mult,
            "source": source,
            "notes": notes,
        }

        # ── 4. CEO-ONLY enrichments (cost / margen / commission) ──
        if _is_ceo(request) and grade_item and grade_item.cost_usd is not None:
            cost = Decimal(str(grade_item.cost_usd))
            margen_usd = final_price - cost
            if final_price > 0:
                margen_pct = (margen_usd / final_price * Decimal("100")).quantize(Decimal("0.001"))
            else:
                margen_pct = None

            out["cost_usd"]   = cost
            out["margen_usd"] = margen_usd
            out["margen_pct"] = margen_pct

            # Commission Rule aplicable (client-specific primero, luego global por marca)
            rule = (
                CommissionRule.objects.filter(
                    brand_id=brand_id, client_id=client_id, is_active=True,
                ).order_by("-valid_from").first()
                or
                CommissionRule.objects.filter(
                    brand_id=brand_id, client_id__isnull=True, is_active=True,
                ).order_by("-valid_from").first()
            )
            if rule:
                out["commission_pct"]  = Decimal(str(rule.commission_pct))
                out["commission_base"] = rule.commission_base
                notes.append(f"Commission rule: {rule.codigo} ({rule.commission_base}).")
            out["notes"] = notes

        out_ser = ResolveClientPriceOutputSerializer(out)
        return Response(out_ser.data, status=status.HTTP_200_OK)


# =====================================================================
# COMEX pricing waterfall · endpoint de la calculadora del Excel v6
# =====================================================================
from .models import PricingConstant, PaymentIndex, GradeItem  # noqa: E402


class ResolveComexPriceView(APIView):
    """POST /api/commercial/resolve_price/

    Implementa la fórmula EXACTA de la hoja 'Calculadora' del Excel
    'Tabela de preços COMEX 2026 v6', celda J18:

        precio_final = VLOOKUP(sku, 'Tabela de Preços', col_10)      ← precio_base
                     × (1.0183 ^ (100 × comisión_pct))                ← factor_comisión
                     × VLOOKUP(días, 'Tabela de indices', col_3, 0)   ← factor_índice_ME

    Payload:
        {
          "sku":            "701407",            // obligatorio
          "comision_pct":   0.08,                // opcional (default 0) — decimal, 0.08 = 8%
          "dias_pago":      28,                  // opcional (default 0)
          "mercado":        "ME"                 // "MI" | "ME" (default ME)
        }

    Respuesta:
        {
          "sku":           "701407",
          "price_base_usd":  5.0013,
          "commission_pct":  0.08,
          "commission_factor": 1.15565,
          "payment_days":    28,
          "payment_market":  "ME",
          "payment_factor":  1.009,
          "price_final_usd": 5.8312,
          "breakdown": [
            "precio_base(701407) = 5.0013 USD",
            "commission_factor = 1.0183 ^ (100 × 0.08) = 1.15565",
            "payment_factor(28d, ME) = 1.009",
            "final = 5.0013 × 1.15565 × 1.009 = 5.8312 USD"
          ]
        }

    Errores:
      · 404 · SKU no encontrado o sin pricelist activa.
      · 404 · No existe índice para esos `dias_pago` (sugerir los cercanos).
      · 400 · comision_pct fuera de [0, 1].
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        sku           = (request.data.get("sku") or "").strip()
        comision_pct  = request.data.get("comision_pct")
        dias_pago     = request.data.get("dias_pago")
        mercado       = (request.data.get("mercado") or "ME").upper()

        if not sku:
            return Response({"detail": "sku es obligatorio."}, status=400)

        try:
            comision_pct = Decimal(str(comision_pct)) if comision_pct is not None else Decimal("0")
        except Exception:
            return Response({"detail": "comision_pct debe ser decimal."}, status=400)

        if comision_pct < 0 or comision_pct > 1:
            return Response(
                {"detail": "comision_pct fuera de rango. Usar decimal entre 0 y 1 (ej. 0.08 = 8%)."},
                status=400,
            )

        try:
            dias_pago = int(dias_pago) if dias_pago is not None else 0
        except Exception:
            return Response({"detail": "dias_pago debe ser entero."}, status=400)

        if mercado not in ("MI", "ME"):
            return Response({"detail": "mercado debe ser 'MI' o 'ME'."}, status=400)

        # ── 1. precio base del SKU (GradeItem activo + más reciente) ─────
        gi = (
            GradeItem.objects.filter(product_sku=sku, is_active=True)
            .order_by("-updated_at")
            .first()
        )
        if not gi:
            return Response(
                {"detail": f"SKU '{sku}' no encontrado en pricelists activas."},
                status=404,
            )
        price_base_usd = Decimal(str(gi.unit_price_usd))

        # ── 2. factor de comisión (base ^ (100 × pct)) ───────────────────
        base_const = PricingConstant.objects.filter(
            slug="base_commission_rate", is_active=True,
        ).first()
        base = Decimal(str(base_const.value)) if base_const else Decimal("1.0183")
        # Decimal ^ Decimal no es natural — convertimos via float y
        # volvemos a Decimal con 6 decimales para mantener determinismo.
        import math
        commission_factor = Decimal(
            f"{math.pow(float(base), float(100 * comision_pct)):.6f}"
        )

        # ── 3. factor de plazo (índice MI/ME) ────────────────────────────
        pi = PaymentIndex.objects.filter(dias=dias_pago, is_active=True).first()
        if not pi:
            # Sugerir los plazos cercanos disponibles.
            near = list(
                PaymentIndex.objects.filter(is_active=True)
                .values_list("dias", flat=True)
                .order_by("dias")
            )
            return Response(
                {"detail": f"No existe índice para {dias_pago} días.",
                 "dias_disponibles": near},
                status=404,
            )
        payment_factor = Decimal(str(pi.factor_me if mercado == "ME" else pi.factor_mi))

        # ── 4. precio final ──────────────────────────────────────────────
        price_final = (price_base_usd * commission_factor * payment_factor).quantize(Decimal("0.0001"))

        breakdown = [
            f"precio_base({sku}) = {price_base_usd:.4f} USD",
            f"commission_factor = {base} ^ (100 × {comision_pct}) = {commission_factor}",
            f"payment_factor({dias_pago}d, {mercado}) = {payment_factor}",
            f"final = {price_base_usd:.4f} × {commission_factor} × {payment_factor} = {price_final} USD",
        ]

        return Response({
            "sku":               sku,
            "price_base_usd":    str(price_base_usd),
            "commission_pct":    str(comision_pct),
            "commission_factor": str(commission_factor),
            "payment_days":      dias_pago,
            "payment_market":    mercado,
            "payment_factor":    str(payment_factor),
            "price_final_usd":   str(price_final),
            "currency":          "USD",
            "breakdown":         breakdown,
        }, status=200)


class PaymentIndexListView(APIView):
    """GET /api/commercial/payment_index/ · lista los 34 índices del Excel."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = PaymentIndex.objects.filter(is_active=True).order_by("dias")
        return Response([
            {
                "dias":       p.dias,
                "factor_mi":  str(p.factor_mi),
                "factor_me":  str(p.factor_me),
                "descripcion": p.descripcion,
            }
            for p in qs
        ], status=200)


class PricingConstantListView(APIView):
    """GET /api/commercial/pricing_constants/ · constantes editables."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = PricingConstant.objects.filter(is_active=True).order_by("slug")
        return Response([
            {
                "slug":        c.slug,
                "nombre":      c.nombre,
                "descripcion": c.descripcion,
                "value":       str(c.value),
                "unit":        c.unit,
            }
            for c in qs
        ], status=200)


# =====================================================================
# Sprint M3c · BrandClientPricingAssignment
# =====================================================================
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser  # noqa: E402
from .models import BrandClientPricingAssignment  # noqa: E402
from .serializers import (                         # noqa: E402
    BrandClientPricingAssignmentSerializer,
    BrandClientPricingAssignmentListSerializer,
)


_ADMIN_ROLES = {"superadmin", "admin"}


def _is_admin(user) -> bool:
    if getattr(user, "is_superuser", False):
        return True
    return (getattr(user, "role", "") or "").lower() in _ADMIN_ROLES


class BrandClientPricingAssignmentViewSet(viewsets.ModelViewSet):
    """CRUD de asignaciones cliente↔marca con archivo + modificadores.

      GET  /api/commercial/brand-client-pricing/?brand_id=<uuid>
      GET  /api/commercial/brand-client-pricing/?cliente_id=<uuid>
      POST /api/commercial/brand-client-pricing/
      POST /api/commercial/brand-client-pricing/<id>/upload-file/
      DELETE /api/commercial/brand-client-pricing/<id>/  → soft delete
    """
    required_module = "commercial"
    queryset = BrandClientPricingAssignment.objects.filter(is_active=True)
    serializer_class = BrandClientPricingAssignmentSerializer
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    def get_serializer_class(self):
        if self.action == "list":
            return BrandClientPricingAssignmentListSerializer
        return BrandClientPricingAssignmentSerializer

    def get_queryset(self):
        qs = BrandClientPricingAssignment.objects.filter(is_active=True)
        brand_id   = self.request.query_params.get("brand_id")
        cliente_id = self.request.query_params.get("cliente_id")
        if brand_id:
            qs = qs.filter(brand_id=brand_id)
        if cliente_id:
            qs = qs.filter(cliente_id=cliente_id)
        return qs.order_by("-updated_at")

    def create(self, request, *args, **kwargs):
        """Crea una asignación cliente↔marca de pricing.

        TODO_LOS_CAMPOS_OPCIONALES excepto brand_id y cliente_id.
        - id: lo generamos server-side si no llega
        - fecha_inicio: default = CURRENT_DATE si llega null/vacío
        - modificadores (sobre_precio, pronto_pago, volumen, fechas): null OK
        - snapshot del cliente: best-effort, tolerante a NULLs y errores SQL
        """
        try:
            data = dict(request.data)

            # ── 1. Defaults server-side ──
            if not data.get("id"):
                data["id"] = str(uuid.uuid4())
            if not data.get("fecha_inicio"):
                data["fecha_inicio"] = timezone.now().date().isoformat()

            # Validación mínima (los únicos 2 obligatorios)
            brand_id   = data.get("brand_id")
            cliente_id = data.get("cliente_id")
            if not brand_id or not cliente_id:
                return Response(
                    {"detail": "brand_id y cliente_id son obligatorios."},
                    status=400,
                )

            # ── 2. Snapshot del cliente (best-effort) ──
            snap_comision = snap_dias = snap_limit = None
            try:
                with connection.cursor() as cur:
                    cur.execute("""
                        SELECT comision_pct, dias_credito, credito_aprobado
                          FROM clientes.cliente
                         WHERE id = %s AND is_active = TRUE
                         LIMIT 1
                    """, [cliente_id])
                    row = cur.fetchone()
                    if row:
                        snap_comision, snap_dias, snap_limit = row
            except Exception as exc:
                log.warning("BCPA snapshot failed: %s", exc)

            # ── 3. Atomic: invalidar asignación previa + crear la nueva ──
            with transaction.atomic():
                BrandClientPricingAssignment.objects.filter(
                    brand_id=brand_id, cliente_id=cliente_id, is_active=True,
                ).update(is_active=False)

                # created_by_id / updated_by_id (tolerante a AnonymousUser)
                user_id = getattr(request.user, "id", None)
                if user_id:
                    data["created_by_id"] = str(user_id)
                    data["updated_by_id"] = str(user_id)

                ser = self.get_serializer(data=data)
                ser.is_valid(raise_exception=True)
                instance = ser.save()

                # Snapshot vía direct field assignment (read_only_fields o no
                # del serializer). Solo seteamos campos efectivamente leídos.
                snapshot_updates = {}
                if snap_comision is not None:
                    snapshot_updates["comision_pct_snapshot"] = snap_comision
                if snap_dias is not None:
                    snapshot_updates["credito_dias_snapshot"] = snap_dias
                if snap_limit is not None:
                    snapshot_updates["credito_limit_snapshot"] = snap_limit
                if snapshot_updates:
                    for k, v in snapshot_updates.items():
                        setattr(instance, k, v)
                    instance.save(update_fields=list(snapshot_updates.keys()))

            return Response(self.get_serializer(instance).data, status=201)

        except Exception as exc:
            log.exception("BCPA create failed")
            return Response(
                {"detail": f"Error al crear asignación: {exc.__class__.__name__}: {exc}"},
                status=500,
            )

    def perform_destroy(self, instance):
        """Soft delete (los archivos permanecen en MinIO)."""
        instance.is_active = False
        instance.save(update_fields=["is_active", "updated_at"])

    # ── POST /api/commercial/brand-client-pricing/<id>/upload-file/ ─────
    @action(detail=True, methods=["post"], url_path="upload-file",
            parser_classes=[MultiPartParser, FormParser])
    def upload_file(self, request, pk=None):
        """Sube un Excel COMEX (Tabela de preços v6+), parsea y puebla
        pricing.pricelist_version + pricing.grade_item.
        """
        try:
            bcpa = BrandClientPricingAssignment.objects.get(pk=pk, is_active=True)
        except BrandClientPricingAssignment.DoesNotExist:
            return Response({"detail": "Asignación no encontrada."}, status=404)

        f = request.FILES.get("file") or request.data.get("file")
        if not f:
            return Response({"detail": "file es obligatorio (multipart)."}, status=400)

        try:
            object_key = f"commercial/pricing/{bcpa.brand_id}/{bcpa.cliente_id}/{uuid.uuid4()}.xlsx"
            bcpa.file_object_key  = object_key
            bcpa.file_name        = getattr(f, "name", "price_list.xlsx")[:255]
            bcpa.file_size_bytes  = getattr(f, "size", None)
            # file_mime es VARCHAR(64). El MIME de xlsx es 71 chars
            # ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            # así que truncamos a 64.
            _mime = getattr(f, "content_type", None)
            bcpa.file_mime        = _mime[:64] if _mime else None
            bcpa.file_uploaded_at = timezone.now()
            # file_uploaded_by es UUIDField — solo seteamos si parece UUID
            uid = getattr(request.user, "id", None)
            try:
                bcpa.file_uploaded_by = uuid.UUID(str(uid)) if uid else None
            except Exception:
                bcpa.file_uploaded_by = None
            bcpa.save()
        except Exception as exc:
            log.exception("upload_file metadata save failed")
            return Response({
                "detail": f"Error al guardar metadata: {exc.__class__.__name__}: {exc}",
            }, status=500)

        try:
            import openpyxl  # type: ignore
        except ImportError:
            return Response({"detail": "openpyxl no instalado.",
                             "assignment": self.get_serializer(bcpa).data}, status=500)
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
        except Exception as exc:
            return Response({"detail": f"Excel inválido: {exc}",
                             "assignment": self.get_serializer(bcpa).data}, status=400)

        sheet = "Tabela de Preços"
        if sheet not in wb.sheetnames:
            return Response({"detail": f"Hoja '{sheet}' no encontrada. Hojas: {wb.sheetnames}",
                             "assignment": self.get_serializer(bcpa).data}, status=400)
        tp = wb[sheet]

        # Normalizar brand_id (puede llegar como UUID o string desde managed=False)
        try:
            brand_uuid = bcpa.brand_id if hasattr(bcpa.brand_id, 'hex') else uuid.UUID(str(bcpa.brand_id))
        except Exception:
            brand_uuid = uuid.uuid4()

        plv_id = uuid.uuid4()
        codigo = f"COMEX-{brand_uuid.hex[:8]}-{int(timezone.now().timestamp())}"
        skus_imported = 0
        skus_skipped = 0
        sample = []

        try:
         with transaction.atomic():
             # ── Desactivar PLVs previas de ESTA asignación cuyo rango
             #    de fechas SE SOLAPA con el nuevo upload ──
             # Política temporal:
             #   · Re-upload con MISMAS fechas (mismo rango)        → solapa → desactiva
             #     vieja. Reemplazo idempotente.
             #   · Upload con fechas FUTURAS (no solapa con activa) → coexisten.
             #     La vigente hoy gana al resolver; cuando expira, la futura
             #     toma el relevo automáticamente (filtrado por fecha en
             #     resolve_client_price y resolved-prices).
             #
             # Lógica de solapamiento entre ranges A=[a1,a2] y B=[b1,b2]:
             #     overlap ⇔ a1 <= b2 AND b1 <= a2  (con null = infinito)
             new_from = bcpa.fecha_inicio or timezone.now().date()
             new_to   = bcpa.fecha_fin  # None ⇒ +infinity

             # existing.valid_to >= new_from  OR  existing.valid_to IS NULL
             cond_a = Q(valid_to__gte=new_from) | Q(valid_to__isnull=True)
             # existing.valid_from <= new_to  (si new_to es None, condición trivial)
             cond_b = Q(valid_from__lte=new_to) if new_to is not None else Q()

             stale_plvs = list(PriceListVersion.objects.filter(
                 brand_id=bcpa.brand_id,
                 is_active=True,
                 metadata__assignment_id=str(bcpa.id),
             ).filter(cond_a & cond_b).values_list("id", flat=True))

             if stale_plvs:
                 GradeItem.objects.filter(
                     pricelist_version_id__in=stale_plvs,
                     is_active=True,
                 ).update(is_active=False)
                 PriceListVersion.objects.filter(
                     id__in=stale_plvs,
                 ).update(is_active=False)
                 log.info(
                     "upload_file: desactivadas %d PLVs previas de BCPA %s "
                     "(solapan con %s..%s)",
                     len(stale_plvs), bcpa.id, new_from, new_to or "∞",
                 )

             PriceListVersion.objects.create(
                 id=plv_id, brand_id=bcpa.brand_id, codigo=codigo,
                 nombre=bcpa.file_name or "COMEX upload",
                 descripcion=f"Upload via BCPA {bcpa.id}", currency="USD",
                 valid_from=bcpa.fecha_inicio or timezone.now().date(),
                 valid_to=bcpa.fecha_fin, storage_key=object_key, source="UPLOAD",
                 uploaded_by_id=getattr(request.user, "id", None),
                 metadata={"assignment_id": str(bcpa.id)},
             )
             grade_items = []
             for row_idx in range(2, tp.max_row + 1):
                 sku_raw = tp.cell(row=row_idx, column=1).value
                 price_raw = tp.cell(row=row_idx, column=10).value
                 if sku_raw is None or price_raw is None:
                     skus_skipped += 1; continue
                 try:
                     sku = str(sku_raw).strip()
                     price = Decimal(str(price_raw))
                 except Exception:
                     skus_skipped += 1; continue
                 if not sku or price < 0:
                     skus_skipped += 1; continue
                 product_name = (tp.cell(row=row_idx, column=2).value
                                 or tp.cell(row=row_idx, column=4).value or "")
                 ncm = tp.cell(row=row_idx, column=6).value
                 ca  = tp.cell(row=row_idx, column=7).value
                 centro = tp.cell(row=row_idx, column=9).value
                 grade_items.append(GradeItem(
                     id=uuid.uuid4(), pricelist_version_id=plv_id,
                     brand_id=bcpa.brand_id, product_sku=sku,
                     product_name=str(product_name)[:240], unit_price_usd=price,
                     grade_moq_total=0, size_multipliers={}, tags=[],
                     metadata={"ncm": str(ncm) if ncm is not None else None,
                               "ca":  str(ca)  if ca  is not None else None,
                               "centro_facturacion": str(centro) if centro is not None else None,
                               "row_excel": row_idx},
                 ))
                 if len(sample) < 5: sample.append(sku)
                 skus_imported += 1
             if grade_items:
                 GradeItem.objects.bulk_create(grade_items, batch_size=200)

             # ── Auto-clear CPAs override del cliente para los SKUs subidos ──
             # Si el operador sube un Excel nuevo, asume que esos precios son
             # los nuevos. Cualquier override manual previo (introducido desde
             # el detalle del producto u otro flujo) queda obsoleto y se desactiva.
             # Usamos SOFT-delete (is_active=FALSE) para preservar trazabilidad
             # — los CPAs viejos quedan en la BD pero ya no aplican en el waterfall.
             cpas_cleared = 0
             if grade_items:
                 imported_skus = list({gi.product_sku for gi in grade_items})
                 # Hacemos en chunks de 1000 para evitar query gigante con un Excel
                 # grande (ANY() con miles de SKUs en un IN clause es OK pero por
                 # las dudas).
                 chunk = 1000
                 for i in range(0, len(imported_skus), chunk):
                     chunk_skus = imported_skus[i:i + chunk]
                     n = ClientAssignment.objects.filter(
                         client_id=bcpa.cliente_id,
                         brand_id=bcpa.brand_id,
                         brand_sku__in=chunk_skus,
                         is_active=True,
                     ).update(is_active=False)
                     cpas_cleared += n
                 if cpas_cleared:
                     log.info(
                         "upload_file: desactivados %d CPA overrides para "
                         "cliente %s / marca %s (SKUs del Excel)",
                         cpas_cleared, bcpa.cliente_id, bcpa.brand_id,
                     )

         return Response({"assignment": self.get_serializer(bcpa).data,
                          "pricelist_version_id": str(plv_id),
                          "pricelist_codigo": codigo,
                          "skus_imported": skus_imported,
                          "skus_skipped": skus_skipped,
                          "cpas_cleared": cpas_cleared,
                          "sample": sample}, status=200)
        except Exception as exc:
            log.exception("upload_file failed")
            return Response({
                "detail": f"{exc.__class__.__name__}: {exc}",
                "assignment": self.get_serializer(bcpa).data,
                "skus_imported": skus_imported,
                "skus_skipped": skus_skipped,
            }, status=500)


# =====================================================================
class BrandClientsSummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, brand_id):
        is_admin = _is_admin(request.user)

        # 1. Todos los clientes activos (ligero — no filtramos por ventas
        #    por ahora; en fase 2 podemos enlazar con expedientes.expediente
        #    WHERE brand_id = ...).
        with connection.cursor() as cur:
            # Sprint Parent-Child: incluimos parent_id para que el FE
            # pueda renderizar la jerarquia con sangria.
            cur.execute("""
                SELECT id, razon_social, nombre_comercial, pais_iso2,
                       estado, dias_credito,
                       credito_aprobado, comision_pct,
                       parent_id
                  FROM clientes.cliente
                 WHERE is_active = TRUE
                 ORDER BY razon_social
                 LIMIT 500
            """)
            cols = [c[0] for c in cur.description]
            rows = [dict(zip(cols, r)) for r in cur.fetchall()]

        # 2. Asignaciones vigentes para esta marca.
        bcpas = {
            str(a.cliente_id): a
            for a in BrandClientPricingAssignment.objects.filter(
                brand_id=brand_id, is_active=True,
            )
        }

        def _assignment_dict(a):
            if not a:
                return None
            return {
                "id":                str(a.id),
                "file_name":         a.file_name,
                "file_size_bytes":   a.file_size_bytes,
                "file_uploaded_at":  a.file_uploaded_at.isoformat() if a.file_uploaded_at else None,
                "fecha_inicio":      a.fecha_inicio.isoformat() if a.fecha_inicio else None,
                "fecha_fin":         a.fecha_fin.isoformat() if a.fecha_fin else None,
                "sobre_precio_pct":  str(a.sobre_precio_pct)   if a.sobre_precio_pct  is not None else None,
                "pronto_pago_dias":  a.pronto_pago_dias,
                "pronto_pago_pct":   str(a.pronto_pago_pct)    if a.pronto_pago_pct   is not None else None,
                "volumen_pct":       str(a.volumen_pct)        if a.volumen_pct      is not None else None,
                "volumen_min_units": a.volumen_min_units,
                "updated_at":        a.updated_at.isoformat()  if a.updated_at else None,
            }

        out = []
        for r in rows:
            assignment = bcpas.get(str(r["id"]))
            card = {
                "cliente_id":       str(r["id"]),
                "razon_social":     r["razon_social"],
                "nombre_comercial": r["nombre_comercial"],
                "pais_iso2":        r["pais_iso2"],
                "estado":           r["estado"],
                "dias_credito":     r["dias_credito"],
                # Parent-Child (sprint 2026-04-29) — FE sangra subsidiarias.
                "parent_id":        str(r["parent_id"]) if r.get("parent_id") else None,
                # CEO-ONLY — enmascarado para no-admin
                "credito_limit_usd": str(r["credito_aprobado"]) if is_admin and r["credito_aprobado"] is not None else None,
                "comision_pct":      str(r["comision_pct"])    if is_admin and r["comision_pct"]    is not None else None,
                "assignment":        _assignment_dict(assignment),
            }
            out.append(card)

        return Response({
            "brand_id": str(brand_id),
            "is_admin": is_admin,
            "is_admin": is_admin,
            "count":    len(out),
            "clients":  out,
        }, status=200)


# =====================================================================
# Helpers waterfall — fórmula COMEX + modificadores BCPA
# =====================================================================
def _comex_factor_comision(comision_pct):
    import math
    base = Decimal("1.0183")
    bc = PricingConstant.objects.filter(slug="base_commission_rate", is_active=True).first()
    if bc:
        base = Decimal(str(bc.value))
    return Decimal(f"{math.pow(float(base), float(100 * comision_pct)):.6f}")


def _comex_factor_indice(dias, mercado="ME"):
    pi = PaymentIndex.objects.filter(dias=dias, is_active=True).first()
    if not pi:
        return Decimal("1.0")
    return Decimal(str(pi.factor_me if mercado == "ME" else pi.factor_mi))


def _resolve_price_for_assignment(gi, bcpa):
    precio_base = Decimal(str(gi.unit_price_usd))
    comision = Decimal(str(bcpa.comision_pct_snapshot or 0))
    fc = _comex_factor_comision(comision)
    # D6 del Excel = días de pago para el factor índice. Prioridad:
    #   1. pronto_pago_dias (override explícito de la asignación)
    #   2. credito_dias_snapshot (snapshot de días de crédito del cliente)
    #   3. 0 (contado, factor_me = 1.0)
    if bcpa.pronto_pago_dias is not None:
        dias = int(bcpa.pronto_pago_dias)
    elif bcpa.credito_dias_snapshot is not None:
        dias = int(bcpa.credito_dias_snapshot)
    else:
        dias = 0
    fi = _comex_factor_indice(dias, mercado="ME")
    # Precio calculadora redondeado a 2 decimales (centavos USD) — coincide
    # con la "Preço Líquido" del Excel (14,96 no 14,9647).
    precio_calculadora = (precio_base * fc * fi).quantize(Decimal("0.01"))
    sp_pct  = Decimal(str(bcpa.sobre_precio_pct or 0))
    pp_pct  = Decimal(str(bcpa.pronto_pago_pct or 0))
    vol_pct = Decimal(str(bcpa.volumen_pct or 0))
    factor_mod = (Decimal("1") + sp_pct) * (Decimal("1") - pp_pct) * (Decimal("1") - vol_pct)
    precio_final = (precio_calculadora * factor_mod).quantize(Decimal("0.01"))
    return {
        "sku":                gi.product_sku,
        "product_name":       gi.product_name,
        "precio_base_usd":    str(precio_base),
        "comision_pct":       str(comision),
        "factor_comision":    str(fc),
        "pronto_pago_dias":   dias,
        "factor_indice":      str(fi),
        "precio_calculadora": str(precio_calculadora),
        "sobre_precio_pct":   str(sp_pct),
        "pronto_pago_pct":    str(pp_pct),
        "volumen_pct":        str(vol_pct),
        "factor_modificadores": str(factor_mod.quantize(Decimal("0.000001"))),
        "precio_final_usd":   str(precio_final),
    }


# =====================================================================
# GET /api/commercial/brand-client-pricing/<id>/resolved-prices/
# =====================================================================
class ResolvedPricesByAssignmentView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            bcpa = BrandClientPricingAssignment.objects.get(pk=pk, is_active=True)
        except BrandClientPricingAssignment.DoesNotExist:
            return Response({"detail": "Asignación no encontrada."}, status=404)

        # Selección de la PLV ganadora (mismo criterio que resolve_client_price):
        #   1. PLVs ACTIVE y vigentes HOY.
        #   2. Closed-window (valid_to IS NOT NULL) gana sobre open-window.
        #   3. Empate → created_at DESC (la más reciente).
        # Solo mostramos los SKUs de esa PLV — antes mostraba todos de todas
        # las activas, lo que producía duplicados (mismo SKU N veces) y daba
        # la impresión de que el sistema no actualizaba los precios.
        today = timezone.now().date()
        winning_plv = (
            PriceListVersion.objects.filter(
                brand_id=bcpa.brand_id, is_active=True,
                valid_from__lte=today,
            ).filter(
                Q(valid_to__isnull=True) | Q(valid_to__gte=today)
            ).annotate(
                _has_end=Case(
                    When(valid_to__isnull=False, then=Value(0)),
                    default=Value(1),
                    output_field=IntegerField(),
                )
            ).order_by("_has_end", "-created_at").first()
        )
        if winning_plv:
            items = GradeItem.objects.filter(
                pricelist_version_id=winning_plv.id, is_active=True,
            ).order_by("product_sku")
        else:
            items = GradeItem.objects.none()

        limit = int(request.query_params.get("limit", 1000))
        sku_filter = request.query_params.get("sku")
        if sku_filter:
            items = items.filter(product_sku=sku_filter)

        out = [_resolve_price_for_assignment(gi, bcpa) for gi in items[:limit]]
        return Response({
            "assignment_id":  str(bcpa.id),
            "brand_id":       str(bcpa.brand_id),
            "cliente_id":     str(bcpa.cliente_id),
            "count":          len(out),
            "items":          out,
            # Info de la PLV ganadora (transparencia para debugging y UI):
            "winning_plv": {
                "id":         str(winning_plv.id),
                "codigo":     winning_plv.codigo,
                "valid_from": winning_plv.valid_from.isoformat() if winning_plv.valid_from else None,
                "valid_to":   winning_plv.valid_to.isoformat()   if winning_plv.valid_to   else None,
                "created_at": winning_plv.created_at.isoformat() if winning_plv.created_at else None,
                "selected_by": (
                    "closed_window" if winning_plv.valid_to else "latest_open_window"
                ),
            } if winning_plv else None,
        }, status=200)


# =====================================================================
# GET /api/commercial/products/<sku>/clients-pricing/
# =====================================================================
class ProductClientsPricingView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, sku):
        is_admin = _is_admin(request.user)
        gi = (
            GradeItem.objects.filter(product_sku=str(sku), is_active=True)
            .order_by("-updated_at").first()
        )
        if not gi:
            return Response({
                "sku": sku, "found": False,
                "detail": "SKU sin grade_item activo.",
                "clients": [],
            }, status=200)
        bcpas = list(BrandClientPricingAssignment.objects.filter(
            brand_id=gi.brand_id, is_active=True,
        ))
        cliente_ids = [str(b.cliente_id) for b in bcpas]
        clientes_map = {}
        if cliente_ids:
            with connection.cursor() as cur:
                cur.execute("""
                    SELECT id, razon_social, nombre_comercial, pais_iso2
                      FROM clientes.cliente
                     WHERE is_active = TRUE
                       AND id::text = ANY(%s)
                """, [cliente_ids])
                for row in cur.fetchall():
                    clientes_map[str(row[0])] = {
                        "razon_social":     row[1],
                        "nombre_comercial": row[2],
                        "pais_iso2":        row[3],
                    }

        out_clients = []
        for bcpa in bcpas:
            cli_id = str(bcpa.cliente_id)
            cli = clientes_map.get(cli_id, {})
            resolved = _resolve_price_for_assignment(gi, bcpa)
            out_clients.append({
                "cliente_id":         cli_id,
                "razon_social":       cli.get("razon_social"),
                "nombre_comercial":   cli.get("nombre_comercial"),
                "pais_iso2":          cli.get("pais_iso2"),
                "assignment_id":      str(bcpa.id),
                "precio_calculadora": resolved["precio_calculadora"],
                "precio_final_usd":   resolved["precio_final_usd"],
                "breakdown":          resolved if is_admin else None,
            })

        return Response({
            "sku":              str(sku),
            "brand_id":         str(gi.brand_id),
            "product_name":     gi.product_name,
            "precio_base_usd":  str(gi.unit_price_usd),
            "is_admin":         is_admin,
            "count":            len(out_clients),
            "clients":          out_clients,
        }, status=200)


# =====================================================================
# MarluvasExchangeRateView · cotización USD/BRL en vivo
# ---------------------------------------------------------------------
# Proxy con fallback en cadena hacia 2 proveedores de FX y cache Redis
# (60 min). El simulador de precios Marluvas consume este endpoint
# para resaltar la banda cambial vigente.
#
# Cadena de upstreams (en orden):
#   1. AwesomeAPI BR (mercado BR, datos minuto a minuto)
#   2. Frankfurter (ECB, datos diarios, sin rate limit)
#   3. Último valor cacheado en Redis (stale)
#   4. rate=null + error
#
# GET /api/commercial/exchange-rate/usd-brl/
#   → { rate, bid, ask, high, low, varBid, timestamp, source, cached }
# =====================================================================
class MarluvasExchangeRateView(APIView):
    """Proxy + cache + fallback chain para USD/BRL."""
    permission_classes = [IsAuthenticated]

    CACHE_KEY = "commercial:fx:usd-brl"
    CACHE_TTL = 60 * 60                       # 60 minutos (era 15)
    UPSTREAM_TIMEOUT = 6                       # segundos

    URL_AWESOME    = "https://economia.awesomeapi.com.br/last/USD-BRL"
    URL_FRANKFURTER = "https://api.frankfurter.app/latest?from=USD&to=BRL"

    # ---------- Helpers de fetch por upstream -------------------------
    @classmethod
    def _fetch_awesomeapi(cls):
        """Devuelve payload normalizado o levanta excepción."""
        resp = requests.get(cls.URL_AWESOME, timeout=cls.UPSTREAM_TIMEOUT)
        resp.raise_for_status()
        raw = resp.json()

        # AwesomeAPI a veces devuelve HTTP 200 con body de error:
        # { "status": 429, "code": "QuotaExceeded", ... }
        # Detectamos eso antes de intentar parsear como cotización.
        if isinstance(raw, dict) and raw.get("code") and not raw.get("USDBRL"):
            raise RuntimeError(f"AwesomeAPI body error: {raw.get('code')} — {raw.get('message')}")

        data = raw.get("USDBRL") or {}
        bid = data.get("bid")
        if bid in (None, ""):
            raise RuntimeError("AwesomeAPI: bid vacío en respuesta")
        bid_f = float(bid)
        return {
            "rate":      bid_f,
            "bid":       bid_f,
            "ask":       float(data["ask"])    if data.get("ask")    not in (None, "") else None,
            "high":      float(data["high"])   if data.get("high")   not in (None, "") else None,
            "low":       float(data["low"])    if data.get("low")    not in (None, "") else None,
            "varBid":    float(data["varBid"]) if data.get("varBid") not in (None, "") else None,
            "timestamp": data.get("create_date"),
            "source":    "AwesomeAPI BR",
            "cached":    False,
        }

    @classmethod
    def _fetch_frankfurter(cls):
        """Fallback: ECB via Frankfurter. Sin rate limit."""
        resp = requests.get(cls.URL_FRANKFURTER, timeout=cls.UPSTREAM_TIMEOUT)
        resp.raise_for_status()
        raw = resp.json()
        rate = (raw.get("rates") or {}).get("BRL")
        if rate in (None, ""):
            raise RuntimeError("Frankfurter: rate BRL vacío")
        rate_f = float(rate)
        return {
            "rate":      rate_f,
            "bid":       rate_f,
            "ask":       rate_f,
            "high":      None,
            "low":       None,
            "varBid":    None,
            "timestamp": raw.get("date"),
            "source":    "Frankfurter (ECB)",
            "cached":    False,
        }

    # ---------- View ---------------------------------------------------
    def get(self, request):
        force = str(request.query_params.get("refresh", "")).lower() in ("1", "true", "yes")

        if not force:
            cached = cache.get(self.CACHE_KEY)
            if cached:
                payload = dict(cached)
                payload["cached"] = True
                return Response(payload, status=200)

        # Intentar upstreams en orden — el primero que devuelva válido gana.
        errors = []
        for fetcher_name, fetcher in [
            ("awesomeapi",  self._fetch_awesomeapi),
            ("frankfurter", self._fetch_frankfurter),
        ]:
            try:
                payload = fetcher()
                cache.set(self.CACHE_KEY, payload, timeout=self.CACHE_TTL)
                log.info("MarluvasExchangeRateView: ok via %s (rate=%s)", fetcher_name, payload.get("rate"))
                return Response(payload, status=200)
            except Exception as exc:  # noqa: BLE001
                err_msg = f"{fetcher_name}: {exc}"
                errors.append(err_msg)
                log.warning("MarluvasExchangeRateView upstream %s failed: %s", fetcher_name, exc)

        # Si todos los upstreams fallaron → último valor cacheado (stale).
        stale = cache.get(self.CACHE_KEY)
        if stale:
            payload = dict(stale)
            payload["cached"] = True
            payload["error"] = "Todos los upstreams sin respuesta; usando último valor cacheado."
            payload["upstream_errors"] = errors
            return Response(payload, status=200)

        # Sin cache previo: usar fallback hardcoded y cachear por 5 minutos
        # para evitar saturar el servidor con peticiones bloqueantes repetitivas.
        fallback_rate = 5.10
        fallback_payload = {
            "rate":             fallback_rate,
            "bid":              fallback_rate,
            "ask":              fallback_rate,
            "high":             None,
            "low":              None,
            "varBid":           None,
            "timestamp":        None,
            "source":           "Hardcoded Fallback",
            "cached":           False,
            "error":            "Todos los upstreams (AwesomeAPI, Frankfurter) fallaron. Usando fallback hardcoded.",
            "upstream_errors":  errors,
        }
        try:
            cache.set(self.CACHE_KEY, fallback_payload, timeout=300)
        except Exception:
            pass
        return Response(fallback_payload, status=200)


# =====================================================================
# MarluvasExchangeRateHistoryView · serie histórica USD/BRL (Frankfurter)
# ---------------------------------------------------------------------
# Proxy de la serie temporal diaria de Frankfurter (ECB) para alimentar
# la gráfica de línea/área del Cronograma (/cronograma → tab Análisis).
# Frankfurter expone rangos: /{start}..{end}?from=USD&to=BRL → {rates:{
#   "YYYY-MM-DD": {"BRL": 5.10}, ...}}. Cacheamos 6h en Redis porque la
# serie sólo cambia una vez al día (días hábiles).
#
# GET /api/commercial/exchange-rate/usd-brl/history/?days=180
#   → { series:[{date, rate}], count, start, end,
#       stats:{min,max,avg,std,last,first,change_pct}, source, cached }
# =====================================================================
class MarluvasExchangeRateHistoryView(APIView):
    """Proxy + cache para la serie histórica USD/BRL (ECB/Frankfurter)."""
    permission_classes = [IsAuthenticated]

    CACHE_TTL = 60 * 60 * 6                    # 6 horas
    UPSTREAM_TIMEOUT = 8                       # segundos
    MIN_DAYS = 7
    MAX_DAYS = 365

    @classmethod
    def _fetch_series(cls, days):
        import datetime as _dt
        end = _dt.date.today()
        start = end - _dt.timedelta(days=days)
        url = (
            f"https://api.frankfurter.app/{start.isoformat()}.."
            f"{end.isoformat()}?from=USD&to=BRL"
        )
        resp = requests.get(url, timeout=cls.UPSTREAM_TIMEOUT)
        resp.raise_for_status()
        raw = resp.json() or {}
        rates = raw.get("rates") or {}
        series = []
        for date_str in sorted(rates.keys()):
            brl = (rates[date_str] or {}).get("BRL")
            if brl in (None, ""):
                continue
            series.append({"date": date_str, "rate": round(float(brl), 4)})
        return series

    @staticmethod
    def _stats(series):
        if not series:
            return None
        vals = [p["rate"] for p in series]
        n = len(vals)
        avg = sum(vals) / n
        var = sum((v - avg) ** 2 for v in vals) / n
        first, last = vals[0], vals[-1]
        change_pct = ((last - first) / first * 100.0) if first else 0.0
        return {
            "min":        round(min(vals), 4),
            "max":        round(max(vals), 4),
            "avg":        round(avg, 4),
            "std":        round(var ** 0.5, 4),
            "first":      round(first, 4),
            "last":       round(last, 4),
            "change_pct": round(change_pct, 2),
        }

    def get(self, request):
        try:
            days = int(request.query_params.get("days", 180))
        except (TypeError, ValueError):
            days = 180
        days = max(self.MIN_DAYS, min(self.MAX_DAYS, days))
        force = str(request.query_params.get("refresh", "")).lower() in ("1", "true", "yes")
        cache_key = f"commercial:fx:usd-brl:history:{days}"

        if not force:
            cached = cache.get(cache_key)
            if cached:
                payload = dict(cached)
                payload["cached"] = True
                return Response(payload, status=200)

        try:
            series = self._fetch_series(days)
            payload = {
                "series":  series,
                "count":   len(series),
                "start":   series[0]["date"] if series else None,
                "end":     series[-1]["date"] if series else None,
                "stats":   self._stats(series),
                "source":  "Frankfurter (ECB)",
                "cached":  False,
            }
            cache.set(cache_key, payload, timeout=self.CACHE_TTL)
            return Response(payload, status=200)
        except Exception as exc:  # noqa: BLE001
            log.warning("MarluvasExchangeRateHistoryView upstream failed: %s", exc)
            stale = cache.get(cache_key)
            if stale:
                payload = dict(stale)
                payload["cached"] = True
                payload["error"] = "Upstream sin respuesta; usando último valor cacheado."
                return Response(payload, status=200)
            return Response(
                {"series": [], "count": 0, "stats": None,
                 "source": "Frankfurter (ECB)", "cached": False,
                 "error": f"No se pudo obtener la serie histórica: {exc}"},
                status=200,
            )


# =====================================================================
# UsdCrcExchangeRateView · cotización USD/CRC (colón costarricense) en vivo
# ---------------------------------------------------------------------
# Mismo patrón que MarluvasExchangeRateView (proxy + cache 60min + cadena
# de fallback). La liquidación de movimientos / factura lo consume para
# mostrar montos en colones. Frankfurter (ECB) NO tiene CRC → la cadena es
# AwesomeAPI → open.er-api.com → cache stale → fallback.
# GET /api/commercial/exchange-rate/usd-crc/  → { rate, source, cached, ... }
# =====================================================================
class UsdCrcExchangeRateView(APIView):
    """Proxy + cache + fallback chain para USD/CRC."""
    permission_classes = [IsAuthenticated]

    CACHE_KEY = "commercial:fx:usd-crc"
    CACHE_TTL = 60 * 60
    UPSTREAM_TIMEOUT = 6

    URL_AWESOME = "https://economia.awesomeapi.com.br/last/USD-CRC"
    URL_ERAPI   = "https://open.er-api.com/v6/latest/USD"

    @classmethod
    def _fetch_awesomeapi(cls):
        resp = requests.get(cls.URL_AWESOME, timeout=cls.UPSTREAM_TIMEOUT)
        resp.raise_for_status()
        raw = resp.json()
        if isinstance(raw, dict) and raw.get("code") and not raw.get("USDCRC"):
            raise RuntimeError(f"AwesomeAPI body error: {raw.get('code')}")
        data = raw.get("USDCRC") or {}
        bid = data.get("bid")
        if bid in (None, ""):
            raise RuntimeError("AwesomeAPI: bid vacío")
        b = float(bid)
        return {
            "rate": b, "bid": b,
            "ask":  float(data["ask"])    if data.get("ask")    not in (None, "") else None,
            "high": float(data["high"])   if data.get("high")   not in (None, "") else None,
            "low":  float(data["low"])    if data.get("low")    not in (None, "") else None,
            "varBid": float(data["varBid"]) if data.get("varBid") not in (None, "") else None,
            "timestamp": data.get("create_date"), "source": "AwesomeAPI", "cached": False,
        }

    @classmethod
    def _fetch_erapi(cls):
        resp = requests.get(cls.URL_ERAPI, timeout=cls.UPSTREAM_TIMEOUT)
        resp.raise_for_status()
        raw = resp.json()
        rate = (raw.get("rates") or {}).get("CRC")
        if rate in (None, ""):
            raise RuntimeError("open.er-api: rate CRC vacío")
        r = float(rate)
        return {
            "rate": r, "bid": r, "ask": r, "high": None, "low": None, "varBid": None,
            "timestamp": raw.get("time_last_update_utc"), "source": "open.er-api.com", "cached": False,
        }

    def get(self, request):
        force = str(request.query_params.get("refresh", "")).lower() in ("1", "true", "yes")
        if not force:
            cached = cache.get(self.CACHE_KEY)
            if cached:
                payload = dict(cached); payload["cached"] = True
                return Response(payload, status=200)
        errors = []
        for name, fetcher in [("awesomeapi", self._fetch_awesomeapi), ("er-api", self._fetch_erapi)]:
            try:
                payload = fetcher()
                cache.set(self.CACHE_KEY, payload, timeout=self.CACHE_TTL)
                return Response(payload, status=200)
            except Exception as exc:  # noqa: BLE001
                errors.append(f"{name}: {exc}")
                log.warning("UsdCrcExchangeRateView upstream %s failed: %s", name, exc)
        stale = cache.get(self.CACHE_KEY)
        if stale:
            payload = dict(stale); payload["cached"] = True
            payload["upstream_errors"] = errors
            return Response(payload, status=200)
        fallback = {
            "rate": 505.0, "bid": 505.0, "ask": 505.0, "high": None, "low": None, "varBid": None,
            "timestamp": None, "source": "Hardcoded Fallback", "cached": False,
            "error": "Upstreams USD/CRC fallaron; usando fallback.", "upstream_errors": errors,
        }
        try:
            cache.set(self.CACHE_KEY, fallback, timeout=300)
        except Exception:
            pass
        return Response(fallback, status=200)


# =====================================================================
# MarluvasClientEnabledSkusView · SKUs habilitados por cliente
# ---------------------------------------------------------------------
# Source of truth: `productos.producto.especificaciones->'visibility'`.
# Los toggles "Excepciones por cliente" del detalle del producto se
# persisten en ese JSON con la forma:
#   {
#     "visibility": {
#       "visible_to_all":   <bool>,
#       "client_overrides": { "<cliente_uuid_str>": true, ... }
#     }
#   }
#
# Un SKU está habilitado para un cliente si:
#   · visible_to_all === true                       (catálogo público), o
#   · client_overrides[cliente_id_str] === true     (excepción explícita)
#
# Se usa raw SQL (mismo patrón que apps/expedientes/views_simplified_wizard.py)
# porque Django JSONField __contains no expresa bien el OR cruzado.
#
# GET /api/commercial/clients/<uuid:cliente_id>/enabled-skus/?brand_id=<uuid>
#   → { cliente_id, brand_id, count, skus: [string], source,
#        breakdown: {visible_to_all, overrides} }
# =====================================================================
class MarluvasClientEnabledSkusView(APIView):
    """GET SKUs habilitados para un cliente en una marca (vía Producto.especificaciones)."""
    permission_classes = [IsAuthenticated]

    SQL = """
        SELECT sku,
               COALESCE(
                 (especificaciones #>> '{visibility,visible_to_all}')::boolean,
                 FALSE
               ) AS visible_to_all,
               COALESCE(
                 (especificaciones #>> ARRAY['visibility','client_overrides', %(cid)s])::boolean,
                 FALSE
               ) AS overridden
          FROM productos.producto
         WHERE sku IS NOT NULL
           AND COALESCE(is_active, TRUE) = TRUE
           AND (%(brand_id)s IS NULL OR marca_id = %(brand_id)s)
           AND (
             COALESCE((especificaciones #>> '{visibility,visible_to_all}')::boolean, FALSE) = TRUE
             OR
             COALESCE((especificaciones #>> ARRAY['visibility','client_overrides', %(cid)s])::boolean, FALSE) = TRUE
           )
         ORDER BY sku
    """

    def get(self, request, cliente_id):
        brand_id = request.query_params.get("brand_id") or None
        cid_str = str(cliente_id)

        params = {"cid": cid_str, "brand_id": brand_id}

        rows = []
        try:
            with connection.cursor() as cur:
                cur.execute(self.SQL, params)
                rows = cur.fetchall()
        except Exception as exc:  # noqa: BLE001 — degradación: log y devuelve vacío
            log.warning("MarluvasClientEnabledSkusView failed: %s", exc)
            return Response({
                "cliente_id": cid_str,
                "brand_id":   brand_id,
                "count":      0,
                "skus":       [],
                "source":     "error",
                "error":      str(exc),
            }, status=200)

        skus = [r[0] for r in rows]
        visible_to_all_count = sum(1 for r in rows if r[1])
        overridden_count     = sum(1 for r in rows if r[2])

        return Response({
            "cliente_id": cid_str,
            "brand_id":   brand_id,
            "count":      len(skus),
            "skus":       skus,
            "source":     "producto_visibility" if skus else "empty",
            "breakdown": {
                "visible_to_all": visible_to_all_count,
                "overrides":      overridden_count,
            },
        }, status=200)


# =====================================================================
# MarluvasSaveSimulationView · persistencia de simulación de precios
# ---------------------------------------------------------------------
# Guarda el estado actual del simulador Marluvas para un par
# (brand_id, cliente_id). Implementa "snapshot por reemplazo":
# desactiva todas las rows activas previas del par y reinserta una
# row por cada SKU del payload con `activo === true`. Los SKUs con
# `activo === false` no se reinsertan pero su row previa queda
# desactivada (auditoría preservada).
#
# Respeta el UNIQUE INDEX parcial
#   (brand_id, cliente_id, sku) WHERE is_active=TRUE
# gracias al bloque transaccional: desactivar → insertar dentro del
# mismo @transaction.atomic.
#
# POST /api/commercial/marluvas/save-simulation/
#   body: {
#     brand_id, cliente_id,
#     fecha_inicio (YYYY-MM-DD|null), fecha_fin (YYYY-MM-DD|null),
#     skus: [{sku, brl_override, com_pct, ajuste_usd,
#             sobreprecio_pct, activo}]
#   }
#   → { saved, brand_id, cliente_id, snapshot_at }
# =====================================================================
class MarluvasSaveSimulationView(APIView):
    """POST · snapshot por reemplazo de la grilla de precios Marluvas."""
    permission_classes = [IsAuthenticated]

    @staticmethod
    def _to_decimal(value, *, allow_null=False):
        """Convierte JSON value a Decimal preservando precisión.

        - None / "" → None si allow_null, si no Decimal("0").
        - str/float/int → Decimal(str(value))
        - Cualquier otro tipo o no-numérico → ValueError.
        """
        if value is None or value == "":
            if allow_null:
                return None
            return Decimal("0")
        try:
            return Decimal(str(value))
        except (ArithmeticError, ValueError, TypeError) as exc:
            raise ValueError(f"valor decimal inválido: {value!r}") from exc

    @staticmethod
    def _parse_uuid(value, field_name):
        if not value:
            raise ValueError(f"{field_name} es requerido")
        try:
            return uuid.UUID(str(value))
        except (ValueError, AttributeError, TypeError) as exc:
            raise ValueError(f"{field_name} no es un UUID válido") from exc

    @staticmethod
    def _parse_date_optional(value, field_name):
        """None | '' → None. String YYYY-MM-DD → date. Otro → ValueError."""
        if value in (None, ""):
            return None
        from django.utils.dateparse import parse_date
        parsed = parse_date(str(value))
        if parsed is None:
            raise ValueError(f"{field_name} debe ser YYYY-MM-DD o null")
        return parsed

    # ------------------------------------------------------------------
    # _parse_sizes_pricing · normaliza overrides por talla (JSONB) Fase 3
    #
    # Shape esperado:
    #   { "<talla_uuid>": {
    #         "matrix": {"<bandaId>": {"<plazoDias>": <float>}},
    #         "anchor": {"bandaId": <int 1..12>, "plazoDias": <int 8|30|60|90>}
    #     } }
    #
    # Entradas inválidas se descartan silenciosamente (no fallan el SKU).
    # ------------------------------------------------------------------
    @staticmethod
    def _parse_sizes_pricing(raw):
        if not isinstance(raw, dict):
            return {}
        VALID_PLAZOS = {8, 30, 60, 90}
        out = {}
        for talla_key, payload in raw.items():
            try:
                talla_uuid = str(uuid.UUID(str(talla_key)))
            except (ValueError, AttributeError, TypeError):
                continue
            if not isinstance(payload, dict):
                continue

            # matrix: mismo patrón que prices_matrix.
            matrix_clean = {}
            matrix_raw = payload.get("matrix")
            if isinstance(matrix_raw, dict):
                for banda_key, plazos in matrix_raw.items():
                    if not isinstance(plazos, dict):
                        continue
                    plazos_clean = {}
                    for plazo_key, price in plazos.items():
                        try:
                            plazos_clean[str(plazo_key)] = round(float(price), 4)
                        except (TypeError, ValueError):
                            continue
                    if plazos_clean:
                        matrix_clean[str(banda_key)] = plazos_clean

            # anchor: bandaId ∈ 1..12, plazoDias ∈ {8,30,60,90}.
            anchor_clean = None
            anchor_raw = payload.get("anchor")
            if isinstance(anchor_raw, dict):
                try:
                    b = int(anchor_raw.get("bandaId"))
                    p = int(anchor_raw.get("plazoDias"))
                    if 1 <= b <= 12 and p in VALID_PLAZOS:
                        anchor_clean = {"bandaId": b, "plazoDias": p}
                except (TypeError, ValueError):
                    pass

            entry = {}
            if matrix_clean:
                entry["matrix"] = matrix_clean
            if anchor_clean:
                entry["anchor"] = anchor_clean
            if entry:
                out[talla_uuid] = entry
        return out

    # ------------------------------------------------------------------
    # _parse_custom_plazos · normaliza plazos personalizados por banda (Fase 4)
    #
    # Shape esperado:
    #   { "<bandaId 1..12>": [
    #       {"dias": <int 1..3650>, "factor": <float 0..10>},
    #       ...
    #     ] }
    #
    # Bandas inválidas / plazos inválidos se descartan silenciosamente.
    # ------------------------------------------------------------------
    @staticmethod
    def _parse_custom_plazos(raw):
        if not isinstance(raw, dict):
            return {}
        out = {}
        for banda_key, plazos_list in raw.items():
            try:
                banda_id = int(banda_key)
            except (TypeError, ValueError):
                continue
            if not (1 <= banda_id <= 12):
                continue
            if not isinstance(plazos_list, list):
                continue
            clean_list = []
            seen_dias = set()
            for p in plazos_list:
                if not isinstance(p, dict):
                    continue
                try:
                    dias   = int(p.get("dias"))
                    factor = float(p.get("factor"))
                except (TypeError, ValueError):
                    continue
                if not (1 <= dias <= 3650):
                    continue
                if not (0 < factor <= 10):
                    continue
                if dias in seen_dias:
                    continue
                seen_dias.add(dias)
                clean_list.append({"dias": dias, "factor": round(factor, 6)})
            # Ordenamos descendente por días (más largo primero — display order).
            clean_list.sort(key=lambda x: x["dias"], reverse=True)
            if clean_list:
                out[str(banda_id)] = clean_list
        return out

    @classmethod
    def _build_price_rows(cls, skus_payload, *, brand_id, cliente_id,
                          custom_plazos, fecha_ini, fecha_fin):
        # Construye (sin persistir) las filas MarluvasClientSkuPricing a partir
        # del payload. Cada fila lleva ._anchor_for_history para la bitacora.
        # Lanza ValueError ante shape invalido. Compartido por snapshot/upsert.
        from .models import MarluvasClientSkuPricing

        rows = []
        for idx, item in enumerate(skus_payload):
            if not isinstance(item, dict):
                raise ValueError(f"skus[{idx}] debe ser objeto.")
            if not item.get("activo"):
                continue
            sku = (item.get("sku") or "").strip()
            if not sku:
                raise ValueError(f"skus[{idx}].sku es requerido.")

            matrix_raw = item.get("prices_matrix")
            if isinstance(matrix_raw, dict):
                prices_matrix = {}
                for banda_key, plazos in matrix_raw.items():
                    if not isinstance(plazos, dict):
                        continue
                    plazos_clean = {}
                    for plazo_key, price in plazos.items():
                        try:
                            plazos_clean[str(plazo_key)] = round(float(price), 4)
                        except (TypeError, ValueError):
                            continue
                    if plazos_clean:
                        prices_matrix[str(banda_key)] = plazos_clean
            else:
                prices_matrix = {}

            sizes_pricing = cls._parse_sizes_pricing(item.get("sizes_pricing"))

            anchor_raw = item.get("anchor")
            anchor_clean = None
            if isinstance(anchor_raw, dict):
                try:
                    b = int(anchor_raw.get("bandaId"))
                    p = int(anchor_raw.get("plazoDias"))
                    if 1 <= b <= 12 and 1 <= p <= 3650:
                        anchor_clean = {"bandaId": b, "plazoDias": p}
                except (TypeError, ValueError):
                    pass

            row = MarluvasClientSkuPricing(
                brand_id        = brand_id,
                cliente_id      = cliente_id,
                sku             = sku,
                brl_override    = cls._to_decimal(item.get("brl_override"), allow_null=True),
                com_pct         = cls._to_decimal(item.get("com_pct")),
                ajuste_usd      = cls._to_decimal(item.get("ajuste_usd")),
                sobreprecio_pct = cls._to_decimal(item.get("sobreprecio_pct")),
                prices_matrix   = prices_matrix,
                sizes_pricing   = sizes_pricing,
                custom_plazos   = custom_plazos,
                is_active       = True,
                fecha_inicio    = fecha_ini,
                fecha_fin       = fecha_fin,
            )
            row._anchor_for_history = anchor_clean
            rows.append(row)
        return rows

    @staticmethod
    def _count_cells(rows):
        total_cells = 0
        for row in rows:
            m = row.prices_matrix or {}
            if isinstance(m, dict):
                for plazos in m.values():
                    if isinstance(plazos, dict):
                        total_cells += len(plazos)
        return total_cells

    @classmethod
    def _write_price_history(cls, *, brand_id, cliente_id, rows, custom_plazos,
                             banda_vigente_id, fecha_ini, fecha_fin, notas,
                             user_id, total_cells):
        # Bitacora best-effort: nunca debe tumbar el guardado.
        try:
            from .models import (
                MarluvasPriceHistoryEvent, MarluvasPriceHistorySku,
            )
            history_event = MarluvasPriceHistoryEvent.objects.create(
                brand_id           = brand_id,
                cliente_id         = cliente_id,
                created_by_user_id = user_id,
                fecha_inicio       = fecha_ini,
                fecha_fin          = fecha_fin,
                custom_plazos      = custom_plazos or {},
                sku_count          = len(rows),
                cells_count        = total_cells,
                banda_vigente_id   = banda_vigente_id,
                notas              = notas or None,
            )
            history_skus = []
            for row in rows:
                history_skus.append(MarluvasPriceHistorySku(
                    event_id        = history_event.id,
                    sku             = row.sku,
                    brl_override    = row.brl_override,
                    com_pct         = row.com_pct,
                    ajuste_usd      = row.ajuste_usd,
                    sobreprecio_pct = row.sobreprecio_pct,
                    anchor          = getattr(row, "_anchor_for_history", None) or None,
                    prices_matrix   = row.prices_matrix or {},
                    sizes_pricing   = getattr(row, "sizes_pricing", None) or {},
                    activo          = row.is_active,
                ))
            if history_skus:
                MarluvasPriceHistorySku.objects.bulk_create(history_skus, batch_size=200)
        except Exception as hist_exc:  # noqa: BLE001
            log.warning(
                "price history insert failed (brand=%s cliente=%s): %s",
                brand_id, cliente_id, hist_exc,
            )

    @staticmethod
    def _enabled_skus_for_client(brand_id, cliente_id, skus):
        """Subconjunto de `skus` que siguen habilitados para el cliente.

        Misma fuente de verdad que MarluvasClientEnabledSkusView:
        productos.producto.especificaciones->'visibility' (visible_to_all
        OR client_overrides[cliente]). Devuelve set(); ante fallo de DB
        devuelve vacío para no bloquear el save (degradación segura).
        """
        if not skus:
            return set()
        cid_str = str(cliente_id)
        rows = []
        try:
            with connection.cursor() as cur:
                cur.execute("""
                    SELECT sku
                      FROM productos.producto
                     WHERE sku = ANY(%(skus)s)
                       AND COALESCE(is_active, TRUE) = TRUE
                       AND (%(brand_id)s IS NULL OR marca_id = %(brand_id)s)
                       AND (
                         COALESCE(
                           (especificaciones #>> '{visibility,visible_to_all}')::boolean,
                           FALSE
                         ) = TRUE
                         OR
                         COALESCE(
                           (especificaciones #>> ARRAY['visibility','client_overrides', %(cid)s])::boolean,
                           FALSE
                         ) = TRUE
                       )
                """, {
                    "skus": list(skus),
                    "brand_id": str(brand_id) if brand_id else None,
                    "cid": cid_str,
                })
                rows = [r[0] for r in cur.fetchall()]
        except Exception as exc:  # noqa: BLE001
            log.warning("_enabled_skus_for_client failed (brand=%s cliente=%s): %s",
                        brand_id, cliente_id, exc)
        return set(rows)

    @classmethod
    def persist_snapshot(cls, *, brand_id, cliente_id, skus_payload,
                         custom_plazos, banda_vigente_id,
                         fecha_ini, fecha_fin, notas, user_id):
        # Snapshot por REEMPLAZO: desactiva los SKUs activos del par
        # (marca, cliente) y reinserta los enviados. Devuelve
        # {saved, cells, size_overrides_saved}.
        #
        # Guard anti-borrado accidental (2026-08-11): el frontend del
        # simulador manda solo los SKUs cargados en su estado. Si por
        # cualquier razon el estado llega parcial (filtro enabled-skus,
        # carga incompleta), un REPLACE total desactivaria silenciosamente
        # las matrices de los SKUs omitidos. Para evitarlo, los SKUs
        # activos que NO figuran en el payload (ni activos ni inactivos)
        # y que siguen habilitados para el cliente se PRESERVAN: no se
        # desactivan ni se borran. Los SKUs que sí vienen en el payload
        # (aunque sea con activo=false) mantienen la semántica REPLACE.
        from .models import MarluvasClientSkuPricing

        rows = cls._build_price_rows(
            skus_payload, brand_id=brand_id, cliente_id=cliente_id,
            custom_plazos=custom_plazos, fecha_ini=fecha_ini, fecha_fin=fecha_fin,
        )

        payload_sku_ids = {
            (item.get("sku") or "").strip()
            for item in skus_payload
            if isinstance(item, dict) and (item.get("sku") or "").strip()
        }
        preserve_skus = set()
        if payload_sku_ids:
            try:
                active_skus = set(
                    MarluvasClientSkuPricing.objects
                        .filter(brand_id=brand_id, cliente_id=cliente_id, is_active=True)
                        .values_list("sku", flat=True)
                )
                omitted = active_skus - payload_sku_ids
                if omitted:
                    preserve_skus = cls._enabled_skus_for_client(
                        brand_id, cliente_id, omitted,
                    )
                    if preserve_skus:
                        log.warning(
                            "persist_snapshot preserva %d SKU(s) habilitados no incluidos "
                            "en el payload (brand=%s cliente=%s): %s",
                            len(preserve_skus), brand_id, cliente_id,
                            sorted(preserve_skus),
                        )
            except Exception as exc:  # noqa: BLE001
                log.warning("persist_snapshot guard skipped (%s): %s", exc, exc)
                preserve_skus = set()

        with transaction.atomic():
            deact_qs = (MarluvasClientSkuPricing.objects
                        .filter(brand_id=brand_id, cliente_id=cliente_id, is_active=True))
            if preserve_skus:
                deact_qs = deact_qs.exclude(sku__in=preserve_skus)
            deact_qs.update(is_active=False)
            if rows:
                MarluvasClientSkuPricing.objects.bulk_create(rows)

        total_cells = cls._count_cells(rows)
        size_overrides_saved = sum(
            len(getattr(row, "sizes_pricing", None) or {}) for row in rows
        )
        cls._write_price_history(
            brand_id=brand_id, cliente_id=cliente_id, rows=rows,
            custom_plazos=custom_plazos, banda_vigente_id=banda_vigente_id,
            fecha_ini=fecha_ini, fecha_fin=fecha_fin, notas=notas,
            user_id=user_id, total_cells=total_cells,
        )
        return {
            "saved":                len(rows),
            "cells":                total_cells,
            "size_overrides_saved": size_overrides_saved,
        }

    @classmethod
    def persist_upsert_many(cls, *, brand_id, cliente_id, skus_payload,
                            custom_plazos, banda_vigente_id,
                            fecha_ini, fecha_fin, notas, user_id):
        # MERGE: upsert de SOLO los SKUs enviados, preservando el resto de la
        # lista del cliente. Por cada SKU desactiva su fila activa previa y
        # reinserta la nueva; los SKUs NO enviados quedan intactos.
        from .models import MarluvasClientSkuPricing

        rows = cls._build_price_rows(
            skus_payload, brand_id=brand_id, cliente_id=cliente_id,
            custom_plazos=custom_plazos, fecha_ini=fecha_ini, fecha_fin=fecha_fin,
        )
        skus_seen = [r.sku for r in rows]
        with transaction.atomic():
            if skus_seen:
                (MarluvasClientSkuPricing.objects
                    .filter(brand_id=brand_id, cliente_id=cliente_id,
                            sku__in=skus_seen, is_active=True)
                    .update(is_active=False))
            if rows:
                MarluvasClientSkuPricing.objects.bulk_create(rows)

        total_cells = cls._count_cells(rows)
        size_overrides_saved = sum(
            len(getattr(row, "sizes_pricing", None) or {}) for row in rows
        )
        cls._write_price_history(
            brand_id=brand_id, cliente_id=cliente_id, rows=rows,
            custom_plazos=custom_plazos, banda_vigente_id=banda_vigente_id,
            fecha_ini=fecha_ini, fecha_fin=fecha_fin, notas=notas,
            user_id=user_id, total_cells=total_cells,
        )
        return {
            "saved":                len(rows),
            "cells":                total_cells,
            "size_overrides_saved": size_overrides_saved,
        }

    def post(self, request):
        # Import diferido para no romper boot si el modelo aún no se cargó.
        from .models import MarluvasClientSkuPricing

        data = request.data or {}

        # --- Validación de campos top-level ------------------------------
        try:
            brand_id   = self._parse_uuid(data.get("brand_id"),   "brand_id")
            cliente_id = self._parse_uuid(data.get("cliente_id"), "cliente_id")
            fecha_ini  = self._parse_date_optional(data.get("fecha_inicio"), "fecha_inicio")
            fecha_fin  = self._parse_date_optional(data.get("fecha_fin"),   "fecha_fin")
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=400)

        skus_payload = data.get("skus")
        if not isinstance(skus_payload, list):
            return Response(
                {"detail": "skus debe ser un array (puede ser vacío)."},
                status=400,
            )

        # Fase 4 · custom_plazos es TOP-LEVEL del payload (mismo valor
        # para todos los SKUs del cliente-marca). Lo aplicamos a cada row.
        custom_plazos = self._parse_custom_plazos(data.get("custom_plazos"))

        # F6.1 · 2026-05-21 · Banda vigente (según TC del día) al momento
        # del save. Opcional — si no viene o es inválida, queda NULL y el
        # visor del historial hace fallback al default (banda 6).
        banda_vigente_raw = data.get("banda_vigente_id")
        banda_vigente_id = None
        if banda_vigente_raw is not None:
            try:
                bv = int(banda_vigente_raw)
                if 1 <= bv <= 12:
                    banda_vigente_id = bv
            except (TypeError, ValueError):
                banda_vigente_id = None

        # --- Persistencia (core reutilizable) ----------------------------
        # Sprint 2026-07-16 · el armado de filas + snapshot atomico +
        # bitacora vive ahora en persist_snapshot() para compartirlo con la
        # carga masiva por marca. El comportamiento por-cliente es identico.
        user_obj = getattr(request, "user", None)
        user_id = None
        if user_obj is not None and getattr(user_obj, "is_authenticated", False):
            raw = getattr(user_obj, "id", None) or getattr(user_obj, "pk", None)
            try:
                user_id = uuid.UUID(str(raw)) if raw else None
            except (ValueError, AttributeError, TypeError):
                user_id = None

        try:
            res = self.persist_snapshot(
                brand_id         = brand_id,
                cliente_id       = cliente_id,
                skus_payload     = skus_payload,
                custom_plazos    = custom_plazos,
                banda_vigente_id = banda_vigente_id,
                fecha_ini        = fecha_ini,
                fecha_fin        = fecha_fin,
                notas            = (data.get("notas") if isinstance(data, dict) else None),
                user_id          = user_id,
            )
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=400)
        except Exception as exc:  # noqa: BLE001
            log.warning(
                "MarluvasSaveSimulationView failed (brand=%s cliente=%s): %s",
                brand_id, cliente_id, exc,
            )
            return Response(
                {"detail": f"No se pudo guardar la simulacion: {exc}"},
                status=500,
            )

        return Response({
            "saved":                res["saved"],
            "cells":                res["cells"],
            "size_overrides_saved": res["size_overrides_saved"],
            "brand_id":             str(brand_id),
            "cliente_id":           str(cliente_id),
            "snapshot_at":          timezone.now().isoformat(),
        }, status=200)


# =====================================================================
# MarluvasSaveSimulationBulkView - carga masiva por marca (Sprint 2026-07-16)
# =====================================================================
# POST /api/commercial/marluvas/save-simulation-bulk/
#
# Sube UNA vez la lista (parseada en el front) y genera+guarda la matriz de
# precios para VARIOS clientes en una sola llamada. Cada cliente aporta su
# propio `skus` (matriz ya calculada con SU comision en el front). Por cada
# cliente el backend: (1) persiste el snapshot de precios (persist_snapshot)
# y (2) upserta su BrandClientPricingAssignment con la metadata del archivo
# + vigencia compartida, para que la card muestre "Archivo activo" sin volver
# a subir el Excel por cliente.
#
# Payload:
#   { brand_id, fecha_inicio?, fecha_fin?, custom_plazos?, banda_vigente_id?,
#     file_name?, file_size_bytes?,
#     clients: [ { cliente_id, skus:[...], notas? }, ... ] }
# =====================================================================
def _bulk_upsert_bcpa(*, brand_id, cliente_id, fecha_ini, fecha_fin,
                      file_name, file_size_bytes, notas, user_id):
    # Reemplaza (soft-delete) la asignacion activa previa del par
    # (marca, cliente) y crea una nueva con metadata de archivo + vigencia.
    # Copia el snapshot de comision/dias/limite del cliente.
    from .models import BrandClientPricingAssignment

    snap_comision = snap_dias = snap_limit = None
    try:
        with connection.cursor() as cur:
            cur.execute(
                "SELECT comision_pct, dias_credito, credito_aprobado "
                "FROM clientes.cliente WHERE id = %s AND is_active = TRUE LIMIT 1",
                [str(cliente_id)],
            )
            row = cur.fetchone()
            if row:
                snap_comision, snap_dias, snap_limit = row
    except Exception as exc:  # noqa: BLE001
        log.warning("bulk BCPA snapshot failed (%s): %s", cliente_id, exc)

    with transaction.atomic():
        BrandClientPricingAssignment.objects.filter(
            brand_id=brand_id, cliente_id=cliente_id, is_active=True,
        ).update(is_active=False)

        bcpa = BrandClientPricingAssignment(
            id           = uuid.uuid4(),
            brand_id     = brand_id,
            cliente_id   = cliente_id,
            fecha_inicio = fecha_ini or timezone.now().date(),
            fecha_fin    = fecha_fin,
            notas        = notas or None,
            is_active    = True,
        )
        if file_name:
            bcpa.file_name = str(file_name)[:255]
            bcpa.file_uploaded_at = timezone.now()
            if user_id:
                bcpa.file_uploaded_by = user_id
        if file_size_bytes is not None:
            try:
                bcpa.file_size_bytes = int(file_size_bytes)
            except (TypeError, ValueError):
                pass
        if snap_comision is not None:
            bcpa.comision_pct_snapshot = snap_comision
        if snap_dias is not None:
            bcpa.credito_dias_snapshot = snap_dias
        if snap_limit is not None:
            bcpa.credito_limit_snapshot = snap_limit
        if user_id:
            bcpa.created_by_id = user_id
            bcpa.updated_by_id = user_id
        bcpa.save()
    return bcpa.id


class MarluvasSaveSimulationBulkView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        data = request.data or {}
        try:
            brand_id  = MarluvasSaveSimulationView._parse_uuid(data.get("brand_id"), "brand_id")
            fecha_ini = MarluvasSaveSimulationView._parse_date_optional(data.get("fecha_inicio"), "fecha_inicio")
            fecha_fin = MarluvasSaveSimulationView._parse_date_optional(data.get("fecha_fin"), "fecha_fin")
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=400)

        clients = data.get("clients")
        if not isinstance(clients, list) or not clients:
            return Response({"detail": "clients debe ser un array no vacio."}, status=400)

        custom_plazos = MarluvasSaveSimulationView._parse_custom_plazos(data.get("custom_plazos"))

        banda_vigente_id = None
        bv_raw = data.get("banda_vigente_id")
        if bv_raw is not None:
            try:
                bv = int(bv_raw)
                if 1 <= bv <= 12:
                    banda_vigente_id = bv
            except (TypeError, ValueError):
                banda_vigente_id = None

        file_name = data.get("file_name")
        file_size_bytes = data.get("file_size_bytes")
        mode = str(data.get("mode") or "merge").lower()

        user_obj = getattr(request, "user", None)
        user_id = None
        if user_obj is not None and getattr(user_obj, "is_authenticated", False):
            raw = getattr(user_obj, "id", None) or getattr(user_obj, "pk", None)
            try:
                user_id = uuid.UUID(str(raw)) if raw else None
            except (ValueError, AttributeError, TypeError):
                user_id = None

        results = []
        saved_clients = 0
        for entry in clients:
            if not isinstance(entry, dict):
                results.append({"ok": False, "error": "entrada de cliente invalida"})
                continue
            try:
                cliente_id = MarluvasSaveSimulationView._parse_uuid(entry.get("cliente_id"), "cliente_id")
            except ValueError as exc:
                results.append({"ok": False, "error": str(exc)})
                continue

            skus_payload = entry.get("skus")
            if not isinstance(skus_payload, list):
                results.append({"cliente_id": str(cliente_id), "ok": False,
                                "error": "skus debe ser un array"})
                continue

            # Cliente sin SKUs seleccionados: NO se toca (ni precios ni BCPA).
            if not skus_payload:
                results.append({"cliente_id": str(cliente_id), "ok": True,
                                "saved": 0, "skipped": True})
                continue

            notas = entry.get("notas") or f"[Marluvas v7 bulk - {len(skus_payload)} SKUs]"
            # mode: "merge" (default) upsert solo los SKUs enviados sin borrar el
            # resto; "replace" hace snapshot completo por reemplazo.
            persist_fn = (MarluvasSaveSimulationView.persist_snapshot
                          if mode == "replace"
                          else MarluvasSaveSimulationView.persist_upsert_many)
            try:
                res = persist_fn(
                    brand_id         = brand_id,
                    cliente_id       = cliente_id,
                    skus_payload     = skus_payload,
                    custom_plazos    = custom_plazos,
                    banda_vigente_id = banda_vigente_id,
                    fecha_ini        = fecha_ini,
                    fecha_fin        = fecha_fin,
                    notas            = notas,
                    user_id          = user_id,
                )
            except ValueError as exc:
                results.append({"cliente_id": str(cliente_id), "ok": False, "error": str(exc)})
                continue
            except Exception as exc:  # noqa: BLE001
                log.warning("bulk persist failed (cliente=%s): %s", cliente_id, exc)
                results.append({"cliente_id": str(cliente_id), "ok": False, "error": str(exc)})
                continue

            assignment_id = None
            try:
                assignment_id = _bulk_upsert_bcpa(
                    brand_id=brand_id, cliente_id=cliente_id,
                    fecha_ini=fecha_ini, fecha_fin=fecha_fin,
                    file_name=file_name, file_size_bytes=file_size_bytes,
                    notas=notas, user_id=user_id,
                )
            except Exception as exc:  # noqa: BLE001
                log.warning("bulk BCPA upsert failed (cliente=%s): %s", cliente_id, exc)

            saved_clients += 1
            results.append({
                "cliente_id":    str(cliente_id),
                "ok":            True,
                "saved":         res["saved"],
                "cells":         res["cells"],
                "assignment_id": str(assignment_id) if assignment_id else None,
            })

        return Response({
            "brand_id":       str(brand_id),
            "total_clients":  len(clients),
            "saved_clients":  saved_clients,
            "results":        results,
        }, status=200)


# =====================================================================
# MarluvasLoadSimulationView · cargar último snapshot vigente
# ---------------------------------------------------------------------
# Devuelve la grilla de precios Marluvas vigente (is_active=TRUE)
# para un par (brand_id, cliente_id). Si no hay snapshot previo,
# responde 200 con `source="empty"` y `skus=[]` para que el frontend
# inicialice la matriz con los defaults del catálogo.
#
# GET /api/commercial/marluvas/load-simulation/?brand_id=X&cliente_id=Y
#   → { brand_id, cliente_id, fecha_inicio, fecha_fin,
#       skus: [...], source: "db"|"empty", count }
# =====================================================================
class MarluvasLoadSimulationView(APIView):
    """GET · último snapshot vigente de pricing Marluvas para (brand, cliente)."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from .models import MarluvasClientSkuPricing

        brand_id_raw   = request.query_params.get("brand_id")
        cliente_id_raw = request.query_params.get("cliente_id")

        try:
            brand_id   = uuid.UUID(str(brand_id_raw))   if brand_id_raw   else None
            cliente_id = uuid.UUID(str(cliente_id_raw)) if cliente_id_raw else None
        except (ValueError, AttributeError, TypeError):
            return Response(
                {"detail": "brand_id y cliente_id deben ser UUIDs válidos."},
                status=400,
            )
        if not brand_id or not cliente_id:
            return Response(
                {"detail": "brand_id y cliente_id son requeridos."},
                status=400,
            )

        try:
            qs = (MarluvasClientSkuPricing.objects
                  .filter(brand_id=brand_id,
                          cliente_id=cliente_id,
                          is_active=True)
                  .order_by("sku"))
            rows = list(qs)
        except Exception as exc:  # noqa: BLE001
            log.warning(
                "MarluvasLoadSimulationView failed (brand=%s cliente=%s): %s",
                brand_id, cliente_id, exc,
            )
            return Response(
                {"detail": f"No se pudo cargar la simulación: {exc}"},
                status=500,
            )

        if not rows:
            return Response({
                "brand_id":     str(brand_id),
                "cliente_id":   str(cliente_id),
                "fecha_inicio": None,
                "fecha_fin":    None,
                "skus":         [],
                "source":       "empty",
                "count":        0,
            }, status=200)

        first = rows[0]

        # Contador de celdas (12 bandas × 4 plazos por SKU = 48 esperadas).
        # Útil para que el frontend valide que el snapshot está completo.
        total_cells = 0
        for r in rows:
            m = r.prices_matrix or {}
            if isinstance(m, dict):
                for plazos in m.values():
                    if isinstance(plazos, dict):
                        total_cells += len(plazos)

        skus_out = [{
            "sku":             r.sku,
            "brl_override":    str(r.brl_override) if r.brl_override is not None else None,
            "com_pct":         str(r.com_pct),
            "ajuste_usd":      str(r.ajuste_usd),
            "sobreprecio_pct": str(r.sobreprecio_pct),
            "prices_matrix":   r.prices_matrix or {},
            "sizes_pricing":   getattr(r, "sizes_pricing", None) or {},
            "activo":          True,
            "bcpa_id":         str(r.bcpa_id) if r.bcpa_id else None,
        } for r in rows]

        # Fase 4 · custom_plazos es top-level del response (vale igual para
        # todos los SKUs del cliente-marca). Lo tomamos del primer row.
        custom_plazos_out = getattr(first, "custom_plazos", None) or {}

        return Response({
            "brand_id":     str(brand_id),
            "cliente_id":   str(cliente_id),
            "fecha_inicio": first.fecha_inicio.isoformat() if first.fecha_inicio else None,
            "fecha_fin":    first.fecha_fin.isoformat()    if first.fecha_fin    else None,
            "skus":         skus_out,
            "custom_plazos": custom_plazos_out,
            "source":       "db",
            "count":        len(rows),
            "cells":        total_cells,   # típicamente N_skus × 48
        }, status=200)


# =====================================================================
# MarluvasProductClientsMatrixView · matriz por cliente, dado un SKU
# ---------------------------------------------------------------------
# Vista inversa al simulador: en el detalle del producto, listar todos
# los clientes habilitados que tienen un row en
# pricing.marluvas_client_sku_pricing para ese SKU, y devolver la
# matriz guardada de cada uno.
#
# GET /api/commercial/marluvas/product-clients-matrix/?sku=X&brand_id=Y
#   → { sku, brand_id, count,
#       clients: [{cliente_id, razon_social, nombre_comercial, pais_iso2,
#                  brl_override, com_pct, ajuste_usd, sobreprecio_pct,
#                  prices_matrix, fecha_inicio, fecha_fin, updated_at}] }
#
# Si no hay rows para ese SKU → 200 con clients=[].
# =====================================================================
class MarluvasProductClientsMatrixView(APIView):
    """GET · matrices por cliente para un SKU específico."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from .models import MarluvasClientSkuPricing

        sku      = request.query_params.get("sku")
        brand_id = request.query_params.get("brand_id") or None
        if not sku:
            return Response({"detail": "sku es requerido."}, status=400)

        qs = MarluvasClientSkuPricing.objects.filter(sku=sku, is_active=True)
        if brand_id:
            try:
                brand_uuid = uuid.UUID(str(brand_id))
            except (ValueError, AttributeError, TypeError):
                return Response({"detail": "brand_id no es UUID válido."}, status=400)
            qs = qs.filter(brand_id=brand_uuid)

        rows = list(qs.order_by("cliente_id"))

        # Lookup masivo de clientes para razón social / país.
        cliente_ids = [str(r.cliente_id) for r in rows]
        clientes_map = {}
        if cliente_ids:
            try:
                with connection.cursor() as cur:
                    cur.execute("""
                        SELECT id, razon_social, nombre_comercial, pais_iso2
                          FROM clientes.cliente
                         WHERE is_active = TRUE
                           AND id::text = ANY(%s)
                    """, [cliente_ids])
                    for r in cur.fetchall():
                        clientes_map[str(r[0])] = {
                            "razon_social":     r[1],
                            "nombre_comercial": r[2],
                            "pais_iso2":        r[3],
                        }
            except Exception as exc:  # noqa: BLE001
                log.warning("MarluvasProductClientsMatrixView clientes lookup failed: %s", exc)

        clients_out = []
        for r in rows:
            cli = clientes_map.get(str(r.cliente_id), {})
            clients_out.append({
                "cliente_id":       str(r.cliente_id),
                "razon_social":     cli.get("razon_social"),
                "nombre_comercial": cli.get("nombre_comercial"),
                "pais_iso2":        cli.get("pais_iso2"),
                "brl_override":     str(r.brl_override) if r.brl_override is not None else None,
                "com_pct":          str(r.com_pct),
                "ajuste_usd":       str(r.ajuste_usd),
                "sobreprecio_pct":  str(r.sobreprecio_pct),
                "prices_matrix":    r.prices_matrix or {},
                "sizes_pricing":    getattr(r, "sizes_pricing", None) or {},
                # Fase 4 · plazos custom por banda (global por cliente-marca).
                # El frontend lo usa para mostrar columnas dinámicas en la
                # matriz por cliente del detalle de producto.
                "custom_plazos":    getattr(r, "custom_plazos", None) or {},
                "fecha_inicio":     r.fecha_inicio.isoformat() if r.fecha_inicio else None,
                "fecha_fin":        r.fecha_fin.isoformat()    if r.fecha_fin    else None,
                "updated_at":       r.updated_at.isoformat()   if r.updated_at   else None,
                "bcpa_id":          str(r.bcpa_id) if r.bcpa_id else None,
            })

        return Response({
            "sku":      str(sku),
            "brand_id": str(brand_id) if brand_id else None,
            "count":    len(clients_out),
            "clients":  clients_out,
        }, status=200)


# =====================================================================
# MarluvasUpsertSkuView · actualización de UN solo SKU para UN cliente
# ---------------------------------------------------------------------
# Diferencia clave con MarluvasSaveSimulationView:
#   · Save-simulation: snapshot por REEMPLAZO — desactiva TODOS los rows
#     activos del par (brand, cliente) y reinserta los del payload. Sirve
#     para "guardar todo el simulador" desde la vista cliente-marca.
#   · Upsert-sku: snapshot por PARCHE — desactiva sólo el row activo de
#     (brand, cliente, sku) y crea uno nuevo. NO toca otros SKUs del
#     mismo cliente. Sirve para editar desde el detalle del producto
#     (donde sólo modificás UN SKU para UN cliente).
#
# POST /api/commercial/marluvas/upsert-sku/
#   body: { brand_id, cliente_id, sku,
#           brl_override, com_pct, ajuste_usd, sobreprecio_pct,
#           prices_matrix, fecha_inicio?, fecha_fin? }
#   → { saved: 1, cells, brand_id, cliente_id, sku, snapshot_at }
# =====================================================================
class MarluvasUpsertSkuView(APIView):
    """POST · upsert atómico de UN row (brand, cliente, sku) sin tocar otros."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        from .models import MarluvasClientSkuPricing

        data = request.data or {}

        try:
            brand_id   = MarluvasSaveSimulationView._parse_uuid(data.get("brand_id"),   "brand_id")
            cliente_id = MarluvasSaveSimulationView._parse_uuid(data.get("cliente_id"), "cliente_id")
            fecha_ini  = MarluvasSaveSimulationView._parse_date_optional(data.get("fecha_inicio"), "fecha_inicio")
            fecha_fin  = MarluvasSaveSimulationView._parse_date_optional(data.get("fecha_fin"),   "fecha_fin")
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=400)

        sku = (data.get("sku") or "").strip()
        if not sku:
            return Response({"detail": "sku es requerido."}, status=400)

        try:
            brl_override    = MarluvasSaveSimulationView._to_decimal(data.get("brl_override"), allow_null=True)
            com_pct         = MarluvasSaveSimulationView._to_decimal(data.get("com_pct"))
            ajuste_usd      = MarluvasSaveSimulationView._to_decimal(data.get("ajuste_usd"))
            sobreprecio_pct = MarluvasSaveSimulationView._to_decimal(data.get("sobreprecio_pct"))
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=400)

        # Normalizar prices_matrix (mismo patrón que save-simulation).
        matrix_raw = data.get("prices_matrix")
        prices_matrix = {}
        if isinstance(matrix_raw, dict):
            for banda_key, plazos in matrix_raw.items():
                if not isinstance(plazos, dict):
                    continue
                plazos_clean = {}
                for plazo_key, price in plazos.items():
                    try:
                        plazos_clean[str(plazo_key)] = round(float(price), 4)
                    except (TypeError, ValueError):
                        continue
                if plazos_clean:
                    prices_matrix[str(banda_key)] = plazos_clean

        sizes_pricing = MarluvasSaveSimulationView._parse_sizes_pricing(data.get("sizes_pricing"))

        new_row = MarluvasClientSkuPricing(
            brand_id        = brand_id,
            cliente_id      = cliente_id,
            sku             = sku,
            brl_override    = brl_override,
            com_pct         = com_pct,
            ajuste_usd      = ajuste_usd,
            sobreprecio_pct = sobreprecio_pct,
            prices_matrix   = prices_matrix,
            sizes_pricing   = sizes_pricing,
            is_active       = True,
            fecha_inicio    = fecha_ini,
            fecha_fin       = fecha_fin,
        )

        try:
            with transaction.atomic():
                # Desactivar SOLO el row activo previo de este triple
                # (brand, cliente, sku). Los demás SKUs del cliente quedan intactos.
                (MarluvasClientSkuPricing.objects
                    .filter(brand_id=brand_id,
                            cliente_id=cliente_id,
                            sku=sku,
                            is_active=True)
                    .update(is_active=False))
                new_row.save()
        except Exception as exc:  # noqa: BLE001
            log.warning(
                "MarluvasUpsertSkuView failed (brand=%s cliente=%s sku=%s): %s",
                brand_id, cliente_id, sku, exc,
            )
            return Response(
                {"detail": f"No se pudo guardar el override: {exc}"},
                status=500,
            )

        # Cells contadas del payload aceptado.
        cells = 0
        for plazos in prices_matrix.values():
            if isinstance(plazos, dict):
                cells += len(plazos)

        return Response({
            "saved":                1,
            "cells":                cells,
            "size_overrides_saved": len(sizes_pricing),
            "brand_id":             str(brand_id),
            "cliente_id":           str(cliente_id),
            "sku":                  sku,
            "snapshot_at":          timezone.now().isoformat(),
        }, status=200)
